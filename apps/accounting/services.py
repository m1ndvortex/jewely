"""
Accounting services for jewelry shop management.

This module provides services for integrating with django-ledger and
automating accounting operations for jewelry businesses.
"""

import logging
from datetime import date
from decimal import Decimal
from typing import Dict, Optional

from django.contrib.auth import get_user_model
from django.db import models, transaction

from django_ledger.models import (
    AccountModel,
    ChartOfAccountModel,
    EntityModel,
    JournalEntryModel,
    TransactionModel,
)

from apps.core.models import Tenant

from .models import (
    AccountingConfiguration,
    JewelryChartOfAccounts,
    JewelryEntity,
    JournalEntryTemplate,
    JournalEntryTemplateLine,
)

User = get_user_model()
logger = logging.getLogger(__name__)


class AccountingService:
    """
    Service class for handling accounting operations.
    """

    @staticmethod
    def setup_tenant_accounting(tenant: Tenant, user: User) -> JewelryEntity:
        """
        Set up accounting for a new tenant.
        Creates EntityModel, chart of accounts, and configuration.
        """
        # Check if accounting already exists
        try:
            existing_entity = JewelryEntity.objects.get(tenant=tenant)
            logger.info(f"Accounting already exists for tenant: {tenant.company_name}")
            return existing_entity
        except JewelryEntity.DoesNotExist:
            pass

        with transaction.atomic():
            # Create django-ledger entity
            entity = EntityModel.add_root(
                name=tenant.company_name,
                slug=tenant.slug,
                admin=user,
                hidden=False,
                address_1="",  # Required field
                accrual_method=True,
                fy_start_month=1,
            )

            # Create default entity unit using django-ledger's API
            from django_ledger.models import EntityUnitModel, LedgerModel

            EntityUnitModel.add_root(
                entity=entity, name=f"{tenant.company_name} - Main Unit", slug=f"{tenant.slug}-main"
            )

            # Create a ledger for the entity unit
            LedgerModel.objects.create(
                entity=entity, name=f"{tenant.company_name} - General Ledger"
            )

            # Create jewelry entity wrapper
            jewelry_entity = JewelryEntity.objects.create(
                tenant=tenant,
                ledger_entity=entity,
                fiscal_year_start_month=1,  # January
                default_currency="USD",
            )

            # Create chart of accounts
            AccountingService._create_chart_of_accounts(entity)

            # Create accounting configuration
            AccountingConfiguration.objects.create(
                tenant=tenant, use_automatic_journal_entries=True, inventory_valuation_method="FIFO"
            )

            # Create journal entry templates
            AccountingService._create_journal_templates()

            logger.info(f"Accounting setup completed for tenant: {tenant.company_name}")
            return jewelry_entity

    @staticmethod
    def _create_chart_of_accounts(entity: EntityModel) -> None:
        """
        Create standard chart of accounts for jewelry business.
        """
        # Get predefined chart of accounts
        chart_templates = JewelryChartOfAccounts.objects.filter(is_active=True)

        if not chart_templates.exists():
            # Create default chart of accounts if not exists
            AccountingService._create_default_chart_templates()
            chart_templates = JewelryChartOfAccounts.objects.filter(is_active=True)

        # Create chart of accounts for the entity
        coa = ChartOfAccountModel.objects.create(entity=entity, slug=f"{entity.slug}-coa")

        # Create accounts based on templates
        for template in chart_templates:
            account_role = AccountingService._get_account_role(template.account_type)

            # Use add_root for tree structure
            AccountModel.add_root(
                coa_model=coa,
                code=template.account_code,
                name=template.name,
                role=account_role,
                balance_type=AccountingService._get_balance_type(template.account_type),
                active=template.is_active,
            )

    @staticmethod
    def _create_default_chart_templates() -> None:
        """
        Create default chart of accounts templates for jewelry business.
        """
        default_accounts = [
            # Assets (1000-1999)
            ("1001", "Cash - Checking Account", "CASH", "Primary checking account"),
            ("1002", "Cash - Credit Card Processing", "CASH", "Credit card merchant account"),
            ("1003", "Cash - Petty Cash", "CASH", "Petty cash fund"),
            ("1100", "Accounts Receivable", "ACCOUNTS_RECEIVABLE", "Customer receivables"),
            ("1200", "Inventory - Finished Goods", "INVENTORY", "Finished jewelry inventory"),
            ("1201", "Inventory - Raw Materials", "INVENTORY", "Gold, silver, gems inventory"),
            ("1202", "Inventory - Work in Process", "INVENTORY", "Custom orders in progress"),
            ("1300", "Equipment - Store Fixtures", "EQUIPMENT", "Display cases, lighting"),
            ("1301", "Equipment - Tools", "EQUIPMENT", "Jewelry making tools"),
            ("1302", "Equipment - Security System", "EQUIPMENT", "Security cameras, safes"),
            ("1400", "Prepaid Insurance", "PREPAID_EXPENSES", "Prepaid insurance premiums"),
            ("1401", "Prepaid Rent", "PREPAID_EXPENSES", "Prepaid rent expenses"),
            # Liabilities (2000-2999)
            ("2001", "Accounts Payable", "ACCOUNTS_PAYABLE", "Supplier payables"),
            ("2002", "Accrued Wages", "ACCRUED_EXPENSES", "Unpaid employee wages"),
            ("2003", "Sales Tax Payable", "SALES_TAX_PAYABLE", "Collected sales tax"),
            ("2100", "Business Loan", "LOANS_PAYABLE", "Bank loans"),
            ("2101", "Equipment Financing", "LOANS_PAYABLE", "Equipment loans"),
            # Equity (3000-3999)
            ("3001", "Owner's Equity", "OWNERS_EQUITY", "Owner's capital investment"),
            ("3002", "Retained Earnings", "RETAINED_EARNINGS", "Accumulated profits"),
            # Revenue (4000-4999)
            ("4001", "Jewelry Sales", "JEWELRY_SALES", "Sales of finished jewelry"),
            ("4002", "Repair Services", "REPAIR_REVENUE", "Jewelry repair revenue"),
            ("4003", "Custom Orders", "CUSTOM_ORDER_REVENUE", "Custom jewelry revenue"),
            ("4004", "Gift Card Sales", "GIFT_CARD_REVENUE", "Gift card sales"),
            # Expenses (5000-5999)
            ("5001", "Cost of Goods Sold", "COST_OF_GOODS_SOLD", "Direct cost of inventory sold"),
            ("5100", "Rent Expense", "RENT_EXPENSE", "Store rent"),
            ("5101", "Utilities", "UTILITIES_EXPENSE", "Electricity, water, gas"),
            ("5102", "Insurance", "INSURANCE_EXPENSE", "Business insurance"),
            ("5103", "Marketing & Advertising", "MARKETING_EXPENSE", "Promotional expenses"),
            ("5200", "Wages & Salaries", "EMPLOYEE_WAGES", "Employee compensation"),
            ("5201", "Professional Fees", "PROFESSIONAL_FEES", "Legal, accounting fees"),
            ("5202", "Bank Fees", "BANK_FEES", "Banking and credit card fees"),
            ("5300", "Depreciation Expense", "DEPRECIATION_EXPENSE", "Equipment depreciation"),
        ]

        for code, name, account_type, description in default_accounts:
            JewelryChartOfAccounts.objects.get_or_create(
                account_code=code,
                defaults={
                    "name": name,
                    "account_type": account_type,
                    "description": description,
                    "is_active": True,
                },
            )

    @staticmethod
    def _create_journal_templates() -> None:
        """
        Create journal entry templates for common transactions.
        """
        # Cash Sale Template
        cash_sale_template, created = JournalEntryTemplate.objects.get_or_create(
            template_type="CASH_SALE",
            defaults={
                "name": "Cash Sale",
                "description": "Journal entry for cash sales",
                "is_active": True,
            },
        )

        if created:
            # Debit Cash, Credit Sales Revenue
            JournalEntryTemplateLine.objects.create(
                template=cash_sale_template,
                account_code="1001",  # Cash - Checking
                debit_credit="DEBIT",
                amount_field="total",
                description_template="Cash sale #{sale_number}",
                order=1,
            )
            JournalEntryTemplateLine.objects.create(
                template=cash_sale_template,
                account_code="4001",  # Jewelry Sales
                debit_credit="CREDIT",
                amount_field="subtotal",
                description_template="Cash sale #{sale_number}",
                order=2,
            )
            # Sales tax if applicable
            JournalEntryTemplateLine.objects.create(
                template=cash_sale_template,
                account_code="2003",  # Sales Tax Payable
                debit_credit="CREDIT",
                amount_field="tax",
                description_template="Sales tax on sale #{sale_number}",
                order=3,
            )

        # Card Sale Template
        card_sale_template, created = JournalEntryTemplate.objects.get_or_create(
            template_type="CARD_SALE",
            defaults={
                "name": "Credit Card Sale",
                "description": "Journal entry for credit card sales",
                "is_active": True,
            },
        )

        if created:
            # Debit Credit Card Processing, Credit Sales Revenue
            JournalEntryTemplateLine.objects.create(
                template=card_sale_template,
                account_code="1002",  # Credit Card Processing
                debit_credit="DEBIT",
                amount_field="total",
                description_template="Card sale #{sale_number}",
                order=1,
            )
            JournalEntryTemplateLine.objects.create(
                template=card_sale_template,
                account_code="4001",  # Jewelry Sales
                debit_credit="CREDIT",
                amount_field="subtotal",
                description_template="Card sale #{sale_number}",
                order=2,
            )
            JournalEntryTemplateLine.objects.create(
                template=card_sale_template,
                account_code="2003",  # Sales Tax Payable
                debit_credit="CREDIT",
                amount_field="tax",
                description_template="Sales tax on sale #{sale_number}",
                order=3,
            )

        # Inventory Purchase Template
        purchase_template, created = JournalEntryTemplate.objects.get_or_create(
            template_type="INVENTORY_PURCHASE",
            defaults={
                "name": "Inventory Purchase",
                "description": "Journal entry for inventory purchases",
                "is_active": True,
            },
        )

        if created:
            # Debit Inventory, Credit Accounts Payable
            JournalEntryTemplateLine.objects.create(
                template=purchase_template,
                account_code="1200",  # Inventory
                debit_credit="DEBIT",
                amount_field="total_amount",
                description_template="Inventory purchase - PO #{po_number}",
                order=1,
            )
            JournalEntryTemplateLine.objects.create(
                template=purchase_template,
                account_code="2001",  # Accounts Payable
                debit_credit="CREDIT",
                amount_field="total_amount",
                description_template="Amount owed to supplier - PO #{po_number}",
                order=2,
            )

        # Expense Payment Template
        expense_template, created = JournalEntryTemplate.objects.get_or_create(
            template_type="EXPENSE_PAYMENT",
            defaults={
                "name": "Expense Payment",
                "description": "Journal entry for expense payments",
                "is_active": True,
            },
        )

        if created:
            # Debit Expense Account, Credit Cash
            JournalEntryTemplateLine.objects.create(
                template=expense_template,
                account_code="5400",  # Other Expenses (default)
                debit_credit="DEBIT",
                amount_field="amount",
                description_template="Expense: {description}",
                order=1,
            )
            JournalEntryTemplateLine.objects.create(
                template=expense_template,
                account_code="1001",  # Cash
                debit_credit="CREDIT",
                amount_field="amount",
                description_template="Cash payment for: {description}",
                order=2,
            )

    @staticmethod
    def _get_account_role(account_type: str) -> str:
        """
        Map account type to django-ledger account role.
        """
        role_mapping = {
            "CASH": "asset_ca_cash",
            "INVENTORY": "asset_ca_inv",
            "ACCOUNTS_RECEIVABLE": "asset_ca_recv",
            "EQUIPMENT": "asset_ppe_equip",
            "PREPAID_EXPENSES": "asset_ca_prepaid",
            "ACCOUNTS_PAYABLE": "lia_cl_acc_payable",
            "ACCRUED_EXPENSES": "lia_cl_other",
            "LOANS_PAYABLE": "lia_ltl_notes",
            "SALES_TAX_PAYABLE": "lia_cl_taxes_payable",
            "OWNERS_EQUITY": "eq_capital",
            "RETAINED_EARNINGS": "eq_adjustment",
            "JEWELRY_SALES": "in_operational",
            "REPAIR_REVENUE": "in_operational",
            "CUSTOM_ORDER_REVENUE": "in_operational",
            "GIFT_CARD_REVENUE": "in_operational",
            "COST_OF_GOODS_SOLD": "cogs_regular",
            "RENT_EXPENSE": "ex_regular",
            "UTILITIES_EXPENSE": "ex_regular",
            "INSURANCE_EXPENSE": "ex_regular",
            "MARKETING_EXPENSE": "ex_regular",
            "EMPLOYEE_WAGES": "ex_regular",
            "PROFESSIONAL_FEES": "ex_regular",
            "BANK_FEES": "ex_regular",
            "DEPRECIATION_EXPENSE": "ex_depreciation",
        }
        return role_mapping.get(account_type, "asset_ca_cash")

    @staticmethod
    def _get_balance_type(account_type: str) -> str:
        """
        Get balance type (debit/credit) for account type.
        """
        debit_accounts = [
            "CASH",
            "INVENTORY",
            "ACCOUNTS_RECEIVABLE",
            "EQUIPMENT",
            "PREPAID_EXPENSES",
            "COST_OF_GOODS_SOLD",
            "RENT_EXPENSE",
            "UTILITIES_EXPENSE",
            "INSURANCE_EXPENSE",
            "MARKETING_EXPENSE",
            "EMPLOYEE_WAGES",
            "PROFESSIONAL_FEES",
            "BANK_FEES",
            "DEPRECIATION_EXPENSE",
        ]

        return "debit" if account_type in debit_accounts else "credit"

    @staticmethod
    def create_sale_journal_entry(sale, user: User) -> Optional[JournalEntryModel]:
        """
        Create journal entry for a sale transaction.

        Creates double-entry bookkeeping entries for sales:
        - Debit: Cash/Card account (total amount)
        - Credit: Sales Revenue account (subtotal)
        - Credit: Sales Tax Payable account (tax amount)
        - Debit: Cost of Goods Sold (COGS)
        - Credit: Inventory (COGS)
        """
        try:
            # Get tenant's accounting entity
            jewelry_entity = JewelryEntity.objects.get(tenant=sale.tenant)
            entity = jewelry_entity.ledger_entity
            entity_unit = entity.entityunitmodel_set.first()

            if not entity_unit:
                logger.error(f"No entity unit found for tenant {sale.tenant.company_name}")
                return None

            # Get accounting configuration
            config = AccountingConfiguration.objects.get(tenant=sale.tenant)

            if not config.use_automatic_journal_entries:
                logger.info(
                    f"Automatic journal entries disabled for tenant {sale.tenant.company_name}"
                )
                return None

            with transaction.atomic():
                # Get the ledger for this entity
                ledger = entity.ledgermodel_set.first()
                if not ledger:
                    logger.error(f"No ledger found for entity {entity}")
                    return None

                # Create journal entry (unposted initially)
                journal_entry = JournalEntryModel.objects.create(
                    ledger=ledger,
                    description=f"Sale #{sale.sale_number}",
                    posted=False,
                )

                # 1. Record the sale revenue and payment
                AccountingService._create_sale_revenue_entries(journal_entry, sale, config)

                # 2. Record cost of goods sold and inventory reduction
                AccountingService._create_cogs_entries(journal_entry, sale, config)

                # Post the journal entry
                journal_entry.posted = True
                journal_entry.save()

                logger.info(f"Journal entry created for sale {sale.sale_number}")
                return journal_entry

        except Exception as e:
            logger.error(f"Failed to create journal entry for sale {sale.sale_number}: {str(e)}")
            return None

    @staticmethod
    def _create_sale_revenue_entries(
        journal_entry: JournalEntryModel, sale, config: AccountingConfiguration
    ):
        """Create revenue-related journal entries for a sale."""
        entity = journal_entry.ledger.entity

        # Debit: Cash or Card account (total amount received)
        if sale.payment_method.upper() == "CASH":
            cash_account = AccountModel.objects.get(
                coa_model__entity=entity, code=config.default_cash_account
            )
            TransactionModel.objects.create(
                journal_entry=journal_entry,
                account=cash_account,
                amount=sale.total,
                tx_type="debit",
                description=f"Cash received for sale #{sale.sale_number}",
            )
        elif sale.payment_method.upper() == "CARD":
            card_account = AccountModel.objects.get(
                coa_model__entity=entity, code=config.default_card_account
            )
            TransactionModel.objects.create(
                journal_entry=journal_entry,
                account=card_account,
                amount=sale.total,
                tx_type="debit",
                description=f"Card payment for sale #{sale.sale_number}",
            )
        elif sale.payment_method.upper() == "STORE_CREDIT":
            # For store credit, we debit accounts receivable or a store credit liability account
            # For now, use cash account as placeholder
            cash_account = AccountModel.objects.get(
                coa_model__entity=entity, code=config.default_cash_account
            )
            TransactionModel.objects.create(
                journal_entry=journal_entry,
                account=cash_account,
                amount=sale.total,
                tx_type="debit",
                description=f"Store credit used for sale #{sale.sale_number}",
            )

        # Credit: Sales Revenue (subtotal amount)
        if sale.subtotal > 0:
            sales_account = AccountModel.objects.get(
                coa_model__entity=entity, code=config.default_sales_account
            )
            TransactionModel.objects.create(
                journal_entry=journal_entry,
                account=sales_account,
                amount=sale.subtotal,
                tx_type="credit",
                description=f"Sales revenue for sale #{sale.sale_number}",
            )

        # Credit: Sales Tax Payable (tax amount)
        if sale.tax > 0:
            tax_account = AccountModel.objects.get(
                coa_model__entity=entity, code=config.default_tax_account
            )
            TransactionModel.objects.create(
                journal_entry=journal_entry,
                account=tax_account,
                amount=sale.tax,
                tx_type="credit",
                description=f"Sales tax collected for sale #{sale.sale_number}",
            )

    @staticmethod
    def _create_cogs_entries(
        journal_entry: JournalEntryModel, sale, config: AccountingConfiguration
    ):
        """Create cost of goods sold entries for a sale."""
        entity = journal_entry.ledger.entity

        # Calculate total COGS for all items in the sale
        total_cogs = Decimal("0.00")

        for sale_item in sale.items.all():
            item_cogs = sale_item.inventory_item.cost_price * sale_item.quantity
            total_cogs += item_cogs

        if total_cogs > 0:
            # Debit: Cost of Goods Sold
            cogs_account = AccountModel.objects.get(
                coa_model__entity=entity, code=config.default_cogs_account
            )
            TransactionModel.objects.create(
                journal_entry=journal_entry,
                account=cogs_account,
                amount=total_cogs,
                tx_type="debit",
                description=f"Cost of goods sold for sale #{sale.sale_number}",
            )

            # Credit: Inventory
            inventory_account = AccountModel.objects.get(
                coa_model__entity=entity, code=config.default_inventory_account
            )
            TransactionModel.objects.create(
                journal_entry=journal_entry,
                account=inventory_account,
                amount=total_cogs,
                tx_type="credit",
                description=f"Inventory reduction for sale #{sale.sale_number}",
            )

    @staticmethod
    def create_purchase_journal_entry(purchase_order, user: User) -> Optional[JournalEntryModel]:
        """
        Create journal entry for inventory purchase.

        Creates double-entry bookkeeping entries for purchases:
        - Debit: Inventory (purchase amount)
        - Credit: Accounts Payable (purchase amount)
        """
        try:
            # Get tenant's accounting entity
            jewelry_entity = JewelryEntity.objects.get(tenant=purchase_order.tenant)
            entity = jewelry_entity.ledger_entity

            # Get accounting configuration
            config = AccountingConfiguration.objects.get(tenant=purchase_order.tenant)

            if not config.use_automatic_journal_entries:
                logger.info(
                    f"Automatic journal entries disabled for tenant {purchase_order.tenant.company_name}"
                )
                return None

            with transaction.atomic():
                # Get the ledger for this entity
                ledger = entity.ledgermodel_set.first()
                if not ledger:
                    logger.error(f"No ledger found for entity {entity}")
                    return None

                # Create journal entry
                journal_entry = JournalEntryModel.objects.create(
                    ledger=ledger,
                    description=f"Purchase Order #{purchase_order.po_number}",
                    posted=False,
                )

                # Debit: Inventory
                inventory_account = AccountModel.objects.get(
                    coa_model__entity=entity, code=config.default_inventory_account
                )
                TransactionModel.objects.create(
                    journal_entry=journal_entry,
                    account=inventory_account,
                    amount=purchase_order.total_amount,
                    tx_type="debit",
                    description=f"Inventory purchase - PO #{purchase_order.po_number}",
                )

                # Credit: Accounts Payable
                payable_account = AccountModel.objects.get(
                    coa_model__entity=entity, code="2001"
                )  # Accounts Payable
                TransactionModel.objects.create(
                    journal_entry=journal_entry,
                    account=payable_account,
                    amount=purchase_order.total_amount,
                    tx_type="credit",
                    description=f"Amount owed to supplier - PO #{purchase_order.po_number}",
                )

                # Post the journal entry
                journal_entry.posted = True
                journal_entry.save()

                logger.info(f"Journal entry created for purchase order {purchase_order.po_number}")
                return journal_entry

        except Exception as e:
            logger.error(
                f"Failed to create journal entry for purchase order {purchase_order.po_number}: {str(e)}"
            )
            return None

    @staticmethod
    def create_payment_journal_entry(payment, user: User) -> Optional[JournalEntryModel]:
        """
        Create journal entry for payment to supplier.

        Creates double-entry bookkeeping entries for payments:
        - Debit: Accounts Payable (payment amount)
        - Credit: Cash/Bank (payment amount)
        """
        try:
            # Get tenant's accounting entity
            jewelry_entity = JewelryEntity.objects.get(tenant=payment.tenant)
            entity = jewelry_entity.ledger_entity

            # Get accounting configuration
            config = AccountingConfiguration.objects.get(tenant=payment.tenant)

            if not config.use_automatic_journal_entries:
                logger.info(
                    f"Automatic journal entries disabled for tenant {payment.tenant.company_name}"
                )
                return None

            with transaction.atomic():
                # Get the ledger for this entity
                ledger = entity.ledgermodel_set.first()
                if not ledger:
                    logger.error(f"No ledger found for entity {entity}")
                    return None

                # Create journal entry
                journal_entry = JournalEntryModel.objects.create(
                    ledger=ledger,
                    description=f"Payment #{payment.payment_number}",
                    posted=False,
                )

                # Debit: Accounts Payable (reducing the liability)
                payable_account = AccountModel.objects.get(
                    coa_model__entity=entity, code="2001"
                )  # Accounts Payable
                TransactionModel.objects.create(
                    journal_entry=journal_entry,
                    account=payable_account,
                    amount=payment.amount,
                    tx_type="debit",
                    description=f"Payment to supplier - {payment.supplier.name}",
                )

                # Credit: Cash/Bank (reducing the asset)
                if payment.payment_method.upper() == "CASH":
                    cash_account = AccountModel.objects.get(
                        coa_model__entity=entity, code=config.default_cash_account
                    )
                    account_description = "Cash payment"
                else:
                    # Assume bank/check payment
                    cash_account = AccountModel.objects.get(
                        coa_model__entity=entity, code=config.default_cash_account
                    )
                    account_description = f"{payment.payment_method} payment"

                TransactionModel.objects.create(
                    journal_entry=journal_entry,
                    account=cash_account,
                    amount=payment.amount,
                    tx_type="credit",
                    description=f"{account_description} to {payment.supplier.name}",
                )

                # Post the journal entry
                journal_entry.posted = True
                journal_entry.save()

                logger.info(f"Journal entry created for payment {payment.payment_number}")
                return journal_entry

        except Exception as e:
            logger.error(
                f"Failed to create journal entry for payment {payment.payment_number}: {str(e)}"
            )
            return None

    @staticmethod
    def create_expense_journal_entry(expense, user: User) -> Optional[JournalEntryModel]:
        """
        Create journal entry for business expense.

        Creates double-entry bookkeeping entries for expenses:
        - Debit: Expense Account (expense amount)
        - Credit: Cash/Bank (expense amount)
        """
        try:
            # Get tenant's accounting entity
            jewelry_entity = JewelryEntity.objects.get(tenant=expense.tenant)
            entity = jewelry_entity.ledger_entity

            # Get accounting configuration
            config = AccountingConfiguration.objects.get(tenant=expense.tenant)

            if not config.use_automatic_journal_entries:
                logger.info(
                    f"Automatic journal entries disabled for tenant {expense.tenant.company_name}"
                )
                return None

            with transaction.atomic():
                # Get the ledger for this entity
                ledger = entity.ledgermodel_set.first()
                if not ledger:
                    logger.error(f"No ledger found for entity {entity}")
                    return None

                # Create journal entry
                journal_entry = JournalEntryModel.objects.create(
                    ledger=ledger,
                    description=f"Expense: {expense.description}",
                    posted=False,
                )

                # Debit: Expense Account
                expense_account_code = AccountingService._get_expense_account_code(expense.category)
                expense_account = AccountModel.objects.get(
                    coa_model__entity=entity, code=expense_account_code
                )
                TransactionModel.objects.create(
                    journal_entry=journal_entry,
                    account=expense_account,
                    amount=expense.amount,
                    tx_type="debit",
                    description=f"{expense.category}: {expense.description}",
                )

                # Credit: Cash/Bank
                if expense.payment_method.upper() == "CASH":
                    cash_account = AccountModel.objects.get(
                        coa_model__entity=entity, code=config.default_cash_account
                    )
                    account_description = "Cash expense"
                else:
                    # Assume bank/check payment
                    cash_account = AccountModel.objects.get(
                        coa_model__entity=entity, code=config.default_cash_account
                    )
                    account_description = f"{expense.payment_method} expense"

                TransactionModel.objects.create(
                    journal_entry=journal_entry,
                    account=cash_account,
                    amount=expense.amount,
                    tx_type="credit",
                    description=f"{account_description}: {expense.description}",
                )

                # Post the journal entry
                journal_entry.posted = True
                journal_entry.save()

                logger.info(f"Journal entry created for expense {expense.description}")
                return journal_entry

        except Exception as e:
            logger.error(
                f"Failed to create journal entry for expense {expense.description}: {str(e)}"
            )
            return None

    @staticmethod
    def _get_expense_account_code(expense_category: str) -> str:
        """
        Map expense category to appropriate account code.
        """
        category_mapping = {
            "RENT": "5100",  # Rent Expense
            "UTILITIES": "5101",  # Utilities
            "INSURANCE": "5102",  # Insurance
            "MARKETING": "5103",  # Marketing & Advertising
            "WAGES": "5200",  # Wages & Salaries
            "PROFESSIONAL": "5201",  # Professional Fees
            "BANK_FEES": "5202",  # Bank Fees
            "OFFICE": "5300",  # Office Expenses (general)
            "TRAVEL": "5301",  # Travel Expenses
            "SUPPLIES": "5302",  # Office Supplies
            "EQUIPMENT": "5303",  # Equipment Expenses
            "OTHER": "5400",  # Other Expenses
        }
        return category_mapping.get(expense_category.upper(), "5400")  # Default to Other Expenses

    @staticmethod
    def get_financial_reports(tenant: Tenant, start_date: date, end_date: date) -> Dict:
        """
        Generate financial reports for a tenant.
        """
        try:
            jewelry_entity = JewelryEntity.objects.get(tenant=tenant)
            entity = jewelry_entity.ledger_entity

            # Get balance sheet statement
            balance_sheet = AccountingService._generate_balance_sheet(entity, end_date)

            # Get income statement
            income_statement = AccountingService._generate_income_statement(
                entity, start_date, end_date
            )

            # Get cash flow statement
            cash_flow = AccountingService._generate_cash_flow_statement(
                entity, start_date, end_date
            )

            # Get trial balance
            trial_balance = AccountingService._generate_trial_balance(entity, end_date)

            return {
                "balance_sheet": balance_sheet,
                "income_statement": income_statement,
                "cash_flow": cash_flow,
                "trial_balance": trial_balance,
                "period": {"start_date": start_date, "end_date": end_date},
            }

        except Exception as e:
            logger.error(
                f"Failed to generate financial reports for tenant {tenant.company_name}: {str(e)}"
            )
            return {}

    @staticmethod
    def _generate_balance_sheet(entity: EntityModel, as_of_date: date) -> Dict:  # noqa: C901
        """
        Generate balance sheet report.
        """
        try:
            from django_ledger.models import AccountModel

            coa = entity.chartofaccountmodel_set.first()
            if not coa:
                return {}

            accounts = AccountModel.objects.filter(coa_model=coa, active=True)

            # Initialize balance sheet structure
            balance_sheet = {
                "assets": {
                    "current_assets": [],
                    "fixed_assets": [],
                    "total_assets": Decimal("0.00"),
                },
                "liabilities": {
                    "current_liabilities": [],
                    "long_term_liabilities": [],
                    "total_liabilities": Decimal("0.00"),
                },
                "equity": {"equity_accounts": [], "total_equity": Decimal("0.00")},
                "as_of_date": as_of_date,
            }

            # Categorize accounts and calculate balances
            for account in accounts:
                balance = AccountingService._get_account_balance_for_date(account, as_of_date)

                if balance == 0:
                    continue

                account_data = {"code": account.code, "name": account.name, "balance": balance}

                # Categorize by account role
                if account.role in [
                    "asset_ca_cash",
                    "asset_ca_recv",
                    "asset_ca_inv",
                    "asset_ca_prepaid",
                ]:
                    balance_sheet["assets"]["current_assets"].append(account_data)
                    balance_sheet["assets"]["total_assets"] += balance
                elif account.role in ["asset_ppe_equip", "asset_ppe_building", "asset_intangible"]:
                    balance_sheet["assets"]["fixed_assets"].append(account_data)
                    balance_sheet["assets"]["total_assets"] += balance
                elif account.role in ["lia_cl_acc_payable", "lia_cl_taxes_payable", "lia_cl_other"]:
                    balance_sheet["liabilities"]["current_liabilities"].append(account_data)
                    balance_sheet["liabilities"]["total_liabilities"] += balance
                elif account.role in ["lia_ltl_notes", "lia_ltl_mortgage"]:
                    balance_sheet["liabilities"]["long_term_liabilities"].append(account_data)
                    balance_sheet["liabilities"]["total_liabilities"] += balance
                elif account.role in ["eq_capital", "eq_adjustment", "eq_retained_earnings"]:
                    balance_sheet["equity"]["equity_accounts"].append(account_data)
                    balance_sheet["equity"]["total_equity"] += balance

            return balance_sheet

        except Exception as e:
            logger.error(f"Failed to generate balance sheet: {str(e)}")
            return {}

    @staticmethod
    def _generate_income_statement(  # noqa: C901
        entity: EntityModel, start_date: date, end_date: date
    ) -> Dict:
        """
        Generate income statement (P&L) report.
        """
        try:
            from django_ledger.models import AccountModel

            coa = entity.chartofaccountmodel_set.first()
            if not coa:
                return {}

            accounts = AccountModel.objects.filter(coa_model=coa, active=True)

            # Initialize income statement structure
            income_statement = {
                "revenue": {
                    "operating_revenue": [],
                    "other_revenue": [],
                    "total_revenue": Decimal("0.00"),
                },
                "expenses": {
                    "cost_of_goods_sold": [],
                    "operating_expenses": [],
                    "other_expenses": [],
                    "total_expenses": Decimal("0.00"),
                },
                "net_income": Decimal("0.00"),
                "period": {"start_date": start_date, "end_date": end_date},
            }

            # Calculate period balances for revenue and expense accounts
            for account in accounts:
                period_balance = AccountingService._get_account_period_balance(
                    account, start_date, end_date
                )

                if period_balance == 0:
                    continue

                account_data = {
                    "code": account.code,
                    "name": account.name,
                    "balance": period_balance,
                }

                # Categorize by account role
                if account.role in ["in_operational", "in_sales"]:
                    income_statement["revenue"]["operating_revenue"].append(account_data)
                    income_statement["revenue"]["total_revenue"] += period_balance
                elif account.role in ["in_other", "in_interest"]:
                    income_statement["revenue"]["other_revenue"].append(account_data)
                    income_statement["revenue"]["total_revenue"] += period_balance
                elif account.role in ["cogs_regular", "cogs_other"]:
                    income_statement["expenses"]["cost_of_goods_sold"].append(account_data)
                    income_statement["expenses"]["total_expenses"] += period_balance
                elif account.role in ["ex_regular", "ex_depreciation"]:
                    income_statement["expenses"]["operating_expenses"].append(account_data)
                    income_statement["expenses"]["total_expenses"] += period_balance
                elif account.role in ["ex_other", "ex_interest"]:
                    income_statement["expenses"]["other_expenses"].append(account_data)
                    income_statement["expenses"]["total_expenses"] += period_balance

            # Calculate net income
            income_statement["net_income"] = (
                income_statement["revenue"]["total_revenue"]
                - income_statement["expenses"]["total_expenses"]
            )

            return income_statement

        except Exception as e:
            logger.error(f"Failed to generate income statement: {str(e)}")
            return {}

    @staticmethod
    def _generate_cash_flow_statement(
        entity: EntityModel, start_date: date, end_date: date
    ) -> Dict:
        """
        Generate cash flow statement.
        """
        try:
            from django_ledger.models import AccountModel

            coa = entity.chartofaccountmodel_set.first()
            if not coa:
                return {}

            # Initialize cash flow statement structure
            cash_flow = {
                "operating_activities": {
                    "net_income": Decimal("0.00"),
                    "adjustments": [],
                    "working_capital_changes": [],
                    "net_cash_from_operations": Decimal("0.00"),
                },
                "investing_activities": {
                    "activities": [],
                    "net_cash_from_investing": Decimal("0.00"),
                },
                "financing_activities": {
                    "activities": [],
                    "net_cash_from_financing": Decimal("0.00"),
                },
                "net_change_in_cash": Decimal("0.00"),
                "cash_beginning": Decimal("0.00"),
                "cash_ending": Decimal("0.00"),
                "period": {"start_date": start_date, "end_date": end_date},
            }

            # Get net income from income statement
            income_statement = AccountingService._generate_income_statement(
                entity, start_date, end_date
            )
            cash_flow["operating_activities"]["net_income"] = income_statement.get(
                "net_income", Decimal("0.00")
            )

            # Get cash accounts
            cash_accounts = AccountModel.objects.filter(
                coa_model=coa, role__in=["asset_ca_cash"], active=True
            )

            # Calculate cash balances
            cash_beginning = Decimal("0.00")
            cash_ending = Decimal("0.00")

            for account in cash_accounts:
                cash_beginning += AccountingService._get_account_balance_for_date(
                    account, start_date
                )
                cash_ending += AccountingService._get_account_balance_for_date(account, end_date)

            cash_flow["cash_beginning"] = cash_beginning
            cash_flow["cash_ending"] = cash_ending
            cash_flow["net_change_in_cash"] = cash_ending - cash_beginning

            # For now, set operating cash flow equal to net income
            # In a full implementation, we would add back depreciation and adjust for working capital changes
            cash_flow["operating_activities"]["net_cash_from_operations"] = cash_flow[
                "operating_activities"
            ]["net_income"]

            return cash_flow

        except Exception as e:
            logger.error(f"Failed to generate cash flow statement: {str(e)}")
            return {}

    @staticmethod
    def _generate_trial_balance(entity: EntityModel, as_of_date: date) -> Dict:
        """
        Generate trial balance report.
        """
        try:
            from django_ledger.models import AccountModel

            coa = entity.chartofaccountmodel_set.first()
            if not coa:
                return {}

            accounts = AccountModel.objects.filter(coa_model=coa, active=True).order_by("code")

            trial_balance = {
                "accounts": [],
                "total_debits": Decimal("0.00"),
                "total_credits": Decimal("0.00"),
                "as_of_date": as_of_date,
                "is_balanced": False,
            }

            for account in accounts:
                balance = AccountingService._get_account_balance_for_date(account, as_of_date)

                if balance == 0:
                    continue

                # Determine if balance should be shown as debit or credit
                if account.balance_type == "debit":
                    debit_balance = balance if balance > 0 else Decimal("0.00")
                    credit_balance = abs(balance) if balance < 0 else Decimal("0.00")
                else:
                    credit_balance = balance if balance > 0 else Decimal("0.00")
                    debit_balance = abs(balance) if balance < 0 else Decimal("0.00")

                account_data = {
                    "code": account.code,
                    "name": account.name,
                    "debit_balance": debit_balance,
                    "credit_balance": credit_balance,
                }

                trial_balance["accounts"].append(account_data)
                trial_balance["total_debits"] += debit_balance
                trial_balance["total_credits"] += credit_balance

            # Check if trial balance is balanced
            trial_balance["is_balanced"] = (
                trial_balance["total_debits"] == trial_balance["total_credits"]
            )

            return trial_balance

        except Exception as e:
            logger.error(f"Failed to generate trial balance: {str(e)}")
            return {}

    @staticmethod
    def _get_account_balance_for_date(account, as_of_date: date) -> Decimal:
        """
        Get account balance as of a specific date.
        """
        try:
            from django_ledger.models import TransactionModel

            # Get all transactions for this account up to the date
            transactions = TransactionModel.objects.filter(
                account=account, journal_entry__posted=True
            )

            balance = Decimal("0.00")
            for txn in transactions:
                if txn.tx_type == "debit":
                    balance += txn.amount
                else:  # credit
                    balance -= txn.amount

            # Adjust balance based on account's normal balance type
            if account.balance_type == "credit":
                balance = -balance

            return balance

        except Exception as e:
            logger.error(f"Failed to get account balance for {account.code}: {str(e)}")
            return Decimal("0.00")

    @staticmethod
    def _get_account_period_balance(account, start_date: date, end_date: date) -> Decimal:
        """
        Get account balance for a specific period (for income statement).
        """
        try:
            from django_ledger.models import TransactionModel

            # Get all transactions for this account within the period
            transactions = TransactionModel.objects.filter(
                account=account,
                journal_entry__posted=True,
            )

            balance = Decimal("0.00")
            for txn in transactions:
                if txn.tx_type == "debit":
                    balance += txn.amount
                else:  # credit
                    balance -= txn.amount

            # For revenue and expense accounts, we want the absolute activity
            # Revenue accounts (credit normal) should show positive for credits
            # Expense accounts (debit normal) should show positive for debits
            if account.role in ["in_operational", "in_sales", "in_other", "in_interest"]:
                # Revenue accounts - return negative of balance to show credits as positive
                return -balance
            elif account.role in [
                "cogs_regular",
                "cogs_other",
                "ex_regular",
                "ex_depreciation",
                "ex_other",
                "ex_interest",
            ]:
                # Expense accounts - return balance as-is to show debits as positive
                return balance
            else:
                return balance

        except Exception as e:
            logger.error(f"Failed to get account period balance for {account.code}: {str(e)}")
            return Decimal("0.00")

    @staticmethod
    def get_account_balance(tenant: Tenant, account_code: str, as_of_date: date = None) -> Decimal:
        """
        Get account balance for a specific account.
        """
        try:
            jewelry_entity = JewelryEntity.objects.get(tenant=tenant)
            entity = jewelry_entity.ledger_entity

            AccountModel.objects.get(coa_model__entity=entity, code=account_code)

            if as_of_date is None:
                as_of_date = date.today()

            # For now, return 0 as balance calculation requires transactions
            # This will be implemented in task 7.2 when we add journal entries
            return Decimal("0.00")

        except Exception as e:
            logger.error(f"Failed to get account balance for {account_code}: {str(e)}")
            return Decimal("0")

    @staticmethod
    def export_financial_reports_to_pdf(  # noqa: C901
        tenant: Tenant, start_date: date, end_date: date
    ) -> bytes:
        """
        Export financial reports to PDF format.
        """
        try:
            from io import BytesIO

            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            # Get financial reports data
            reports = AccountingService.get_financial_reports(tenant, start_date, end_date)

            if not reports:
                raise ValueError("No financial reports data available")

            # Create PDF buffer
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5 * inch)

            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=16,
                spaceAfter=30,
                alignment=1,  # Center alignment
            )
            heading_style = ParagraphStyle(
                "CustomHeading", parent=styles["Heading2"], fontSize=14, spaceAfter=12
            )

            # Build PDF content
            story = []

            # Title
            story.append(Paragraph(f"{tenant.company_name} - Financial Reports", title_style))
            story.append(
                Paragraph(
                    f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
                    styles["Normal"],
                )
            )
            story.append(Spacer(1, 20))

            # Balance Sheet
            if reports.get("balance_sheet"):
                story.append(Paragraph("Balance Sheet", heading_style))
                story.append(Paragraph(f"As of {end_date.strftime('%B %d, %Y')}", styles["Normal"]))
                story.append(Spacer(1, 12))

                balance_sheet = reports["balance_sheet"]

                # Assets
                story.append(Paragraph("ASSETS", styles["Heading3"]))

                # Current Assets
                if balance_sheet["assets"]["current_assets"]:
                    story.append(Paragraph("Current Assets:", styles["Normal"]))
                    asset_data = [["Account", "Amount"]]
                    for asset in balance_sheet["assets"]["current_assets"]:
                        asset_data.append(
                            [f"{asset['code']} - {asset['name']}", f"${asset['balance']:,.2f}"]
                        )

                    asset_table = Table(asset_data, colWidths=[4 * inch, 2 * inch])
                    asset_table.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("FONTSIZE", (0, 0), (-1, 0), 12),
                                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                            ]
                        )
                    )
                    story.append(asset_table)
                    story.append(Spacer(1, 12))

                # Total Assets
                story.append(
                    Paragraph(
                        f"Total Assets: ${balance_sheet['assets']['total_assets']:,.2f}",
                        styles["Heading4"],
                    )
                )
                story.append(Spacer(1, 20))

            # Income Statement
            if reports.get("income_statement"):
                story.append(Paragraph("Income Statement", heading_style))
                story.append(
                    Paragraph(
                        f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
                        styles["Normal"],
                    )
                )
                story.append(Spacer(1, 12))

                income_statement = reports["income_statement"]

                # Revenue
                if income_statement["revenue"]["operating_revenue"]:
                    story.append(Paragraph("REVENUE", styles["Heading3"]))
                    revenue_data = [["Account", "Amount"]]
                    for revenue in income_statement["revenue"]["operating_revenue"]:
                        revenue_data.append(
                            [
                                f"{revenue['code']} - {revenue['name']}",
                                f"${revenue['balance']:,.2f}",
                            ]
                        )

                    revenue_table = Table(revenue_data, colWidths=[4 * inch, 2 * inch])
                    revenue_table.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("FONTSIZE", (0, 0), (-1, 0), 12),
                                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                            ]
                        )
                    )
                    story.append(revenue_table)
                    story.append(Spacer(1, 12))

                # Net Income
                story.append(
                    Paragraph(
                        f"Net Income: ${income_statement['net_income']:,.2f}", styles["Heading4"]
                    )
                )
                story.append(Spacer(1, 20))

            # Trial Balance
            if reports.get("trial_balance"):
                story.append(Paragraph("Trial Balance", heading_style))
                story.append(Paragraph(f"As of {end_date.strftime('%B %d, %Y')}", styles["Normal"]))
                story.append(Spacer(1, 12))

                trial_balance = reports["trial_balance"]

                if trial_balance["accounts"]:
                    tb_data = [["Account Code", "Account Name", "Debit", "Credit"]]
                    for account in trial_balance["accounts"]:
                        tb_data.append(
                            [
                                account["code"],
                                account["name"],
                                (
                                    f"${account['debit_balance']:,.2f}"
                                    if account["debit_balance"] > 0
                                    else ""
                                ),
                                (
                                    f"${account['credit_balance']:,.2f}"
                                    if account["credit_balance"] > 0
                                    else ""
                                ),
                            ]
                        )

                    # Add totals row
                    tb_data.append(
                        [
                            "TOTALS",
                            "",
                            f"${trial_balance['total_debits']:,.2f}",
                            f"${trial_balance['total_credits']:,.2f}",
                        ]
                    )

                    tb_table = Table(tb_data, colWidths=[1 * inch, 3 * inch, 1 * inch, 1 * inch])
                    tb_table.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                                ("FONTSIZE", (0, 0), (-1, 0), 10),
                                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                ("BACKGROUND", (0, 1), (-1, -2), colors.beige),
                                ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                            ]
                        )
                    )
                    story.append(tb_table)

            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Failed to export financial reports to PDF: {str(e)}")
            raise

    @staticmethod
    def close_fiscal_year(tenant: Tenant, fiscal_year_end: date, user: User) -> Dict:  # noqa: C901
        """
        Close fiscal year by transferring revenue and expense balances to retained earnings.

        This creates closing entries to:
        1. Close all revenue accounts to retained earnings
        2. Close all expense accounts to retained earnings
        3. Calculate and record net income for the year
        """
        try:
            jewelry_entity = JewelryEntity.objects.get(tenant=tenant)
            entity = jewelry_entity.ledger_entity

            # Get accounting configuration (for future use)
            # config = AccountingConfiguration.objects.get(tenant=tenant)

            # Calculate fiscal year start date
            fiscal_year_start = date(
                fiscal_year_end.year, jewelry_entity.fiscal_year_start_month, 1
            )
            if fiscal_year_end.month < jewelry_entity.fiscal_year_start_month:
                fiscal_year_start = date(
                    fiscal_year_end.year - 1, jewelry_entity.fiscal_year_start_month, 1
                )

            with transaction.atomic():
                # Get the ledger for this entity
                ledger = entity.ledgermodel_set.first()
                if not ledger:
                    logger.error(f"No ledger found for entity {entity}")
                    return {"success": False, "error": "No ledger found"}

                # Get all accounts
                coa = entity.chartofaccountmodel_set.first()
                accounts = AccountModel.objects.filter(coa_model=coa, active=True)

                # Calculate net income for the year
                income_statement = AccountingService._generate_income_statement(
                    entity, fiscal_year_start, fiscal_year_end
                )
                net_income = income_statement.get("net_income", Decimal("0.00"))

                # Create closing journal entry
                closing_entry = JournalEntryModel.objects.create(
                    ledger=ledger,
                    description=f"Fiscal Year End Closing - {fiscal_year_end.year}",
                    posted=False,
                )

                total_revenue_closed = Decimal("0.00")
                total_expenses_closed = Decimal("0.00")

                # Close revenue accounts (debit revenue, credit retained earnings)
                for account in accounts:
                    if account.role in ["in_operational", "in_sales", "in_other", "in_interest"]:
                        period_balance = AccountingService._get_account_period_balance(
                            account, fiscal_year_start, fiscal_year_end
                        )

                        if period_balance > 0:  # Has revenue to close
                            # Debit the revenue account to zero it out
                            TransactionModel.objects.create(
                                journal_entry=closing_entry,
                                account=account,
                                amount=period_balance,
                                tx_type="debit",
                                description=f"Close revenue account {account.code}",
                            )
                            total_revenue_closed += period_balance

                # Close expense accounts (credit expense, debit retained earnings)
                for account in accounts:
                    if account.role in [
                        "cogs_regular",
                        "cogs_other",
                        "ex_regular",
                        "ex_depreciation",
                        "ex_other",
                        "ex_interest",
                    ]:
                        period_balance = AccountingService._get_account_period_balance(
                            account, fiscal_year_start, fiscal_year_end
                        )

                        if period_balance > 0:  # Has expenses to close
                            # Credit the expense account to zero it out
                            TransactionModel.objects.create(
                                journal_entry=closing_entry,
                                account=account,
                                amount=period_balance,
                                tx_type="credit",
                                description=f"Close expense account {account.code}",
                            )
                            total_expenses_closed += period_balance

                # Transfer net income to retained earnings
                retained_earnings_account = AccountModel.objects.filter(
                    coa_model=coa, code="3002"  # Retained Earnings
                ).first()

                if retained_earnings_account and net_income != 0:
                    if net_income > 0:  # Profit - credit retained earnings
                        TransactionModel.objects.create(
                            journal_entry=closing_entry,
                            account=retained_earnings_account,
                            amount=net_income,
                            tx_type="credit",
                            description=f"Net income for fiscal year {fiscal_year_end.year}",
                        )
                    else:  # Loss - debit retained earnings
                        TransactionModel.objects.create(
                            journal_entry=closing_entry,
                            account=retained_earnings_account,
                            amount=abs(net_income),
                            tx_type="debit",
                            description=f"Net loss for fiscal year {fiscal_year_end.year}",
                        )

                # Post the closing entry
                closing_entry.posted = True
                closing_entry.save()

                logger.info(
                    f"Fiscal year {fiscal_year_end.year} closed for tenant {tenant.company_name}"
                )

                return {
                    "success": True,
                    "fiscal_year_end": fiscal_year_end,
                    "net_income": net_income,
                    "total_revenue_closed": total_revenue_closed,
                    "total_expenses_closed": total_expenses_closed,
                    "closing_entry_id": closing_entry.pk,
                }

        except Exception as e:
            logger.error(f"Failed to close fiscal year for tenant {tenant.company_name}: {str(e)}")
            return {"success": False, "error": str(e)}

    @staticmethod
    def export_financial_reports_to_excel(  # noqa: C901
        tenant: Tenant, start_date: date, end_date: date
    ) -> bytes:
        """
        Export financial reports to Excel format.
        """
        try:
            from io import BytesIO

            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            # Get financial reports data
            reports = AccountingService.get_financial_reports(tenant, start_date, end_date)

            if not reports:
                raise ValueError("No financial reports data available")

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Define styles
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            currency_format = "#,##0.00"

            # Balance Sheet worksheet
            if reports.get("balance_sheet"):
                bs_ws = wb.create_sheet("Balance Sheet")
                balance_sheet = reports["balance_sheet"]

                # Title
                bs_ws["A1"] = f"{tenant.company_name} - Balance Sheet"
                bs_ws["A1"].font = title_font
                bs_ws["A2"] = f"As of {end_date.strftime('%B %d, %Y')}"

                row = 4

                # Assets
                bs_ws[f"A{row}"] = "ASSETS"
                bs_ws[f"A{row}"].font = header_font
                row += 1

                if balance_sheet["assets"]["current_assets"]:
                    bs_ws[f"A{row}"] = "Current Assets:"
                    bs_ws[f"A{row}"].font = Font(bold=True)
                    row += 1

                    for asset in balance_sheet["assets"]["current_assets"]:
                        bs_ws[f"A{row}"] = f"{asset['code']} - {asset['name']}"
                        bs_ws[f"B{row}"] = float(asset["balance"])
                        bs_ws[f"B{row}"].number_format = currency_format
                        row += 1

                bs_ws[f"A{row}"] = "Total Assets"
                bs_ws[f"A{row}"].font = header_font
                bs_ws[f"B{row}"] = float(balance_sheet["assets"]["total_assets"])
                bs_ws[f"B{row}"].number_format = currency_format
                bs_ws[f"B{row}"].font = header_font

            # Income Statement worksheet
            if reports.get("income_statement"):
                is_ws = wb.create_sheet("Income Statement")
                income_statement = reports["income_statement"]

                # Title
                is_ws["A1"] = f"{tenant.company_name} - Income Statement"
                is_ws["A1"].font = title_font
                is_ws["A2"] = (
                    f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
                )

                row = 4

                # Revenue
                is_ws[f"A{row}"] = "REVENUE"
                is_ws[f"A{row}"].font = header_font
                row += 1

                if income_statement["revenue"]["operating_revenue"]:
                    for revenue in income_statement["revenue"]["operating_revenue"]:
                        is_ws[f"A{row}"] = f"{revenue['code']} - {revenue['name']}"
                        is_ws[f"B{row}"] = float(revenue["balance"])
                        is_ws[f"B{row}"].number_format = currency_format
                        row += 1

                is_ws[f"A{row}"] = "Total Revenue"
                is_ws[f"A{row}"].font = Font(bold=True)
                is_ws[f"B{row}"] = float(income_statement["revenue"]["total_revenue"])
                is_ws[f"B{row}"].number_format = currency_format
                is_ws[f"B{row}"].font = Font(bold=True)
                row += 2

                # Expenses
                is_ws[f"A{row}"] = "EXPENSES"
                is_ws[f"A{row}"].font = header_font
                row += 1

                if income_statement["expenses"]["cost_of_goods_sold"]:
                    for expense in income_statement["expenses"]["cost_of_goods_sold"]:
                        is_ws[f"A{row}"] = f"{expense['code']} - {expense['name']}"
                        is_ws[f"B{row}"] = float(expense["balance"])
                        is_ws[f"B{row}"].number_format = currency_format
                        row += 1

                if income_statement["expenses"]["operating_expenses"]:
                    for expense in income_statement["expenses"]["operating_expenses"]:
                        is_ws[f"A{row}"] = f"{expense['code']} - {expense['name']}"
                        is_ws[f"B{row}"] = float(expense["balance"])
                        is_ws[f"B{row}"].number_format = currency_format
                        row += 1

                is_ws[f"A{row}"] = "Total Expenses"
                is_ws[f"A{row}"].font = Font(bold=True)
                is_ws[f"B{row}"] = float(income_statement["expenses"]["total_expenses"])
                is_ws[f"B{row}"].number_format = currency_format
                is_ws[f"B{row}"].font = Font(bold=True)
                row += 2

                # Net Income
                is_ws[f"A{row}"] = "Net Income"
                is_ws[f"A{row}"].font = header_font
                is_ws[f"B{row}"] = float(income_statement["net_income"])
                is_ws[f"B{row}"].number_format = currency_format
                is_ws[f"B{row}"].font = header_font

            # Trial Balance worksheet
            if reports.get("trial_balance"):
                tb_ws = wb.create_sheet("Trial Balance")
                trial_balance = reports["trial_balance"]

                # Title
                tb_ws["A1"] = f"{tenant.company_name} - Trial Balance"
                tb_ws["A1"].font = title_font
                tb_ws["A2"] = f"As of {end_date.strftime('%B %d, %Y')}"

                # Headers
                tb_ws["A4"] = "Account Code"
                tb_ws["B4"] = "Account Name"
                tb_ws["C4"] = "Debit"
                tb_ws["D4"] = "Credit"

                for col in ["A4", "B4", "C4", "D4"]:
                    tb_ws[col].font = header_font

                row = 5

                if trial_balance["accounts"]:
                    for account in trial_balance["accounts"]:
                        tb_ws[f"A{row}"] = account["code"]
                        tb_ws[f"B{row}"] = account["name"]

                        if account["debit_balance"] > 0:
                            tb_ws[f"C{row}"] = float(account["debit_balance"])
                            tb_ws[f"C{row}"].number_format = currency_format

                        if account["credit_balance"] > 0:
                            tb_ws[f"D{row}"] = float(account["credit_balance"])
                            tb_ws[f"D{row}"].number_format = currency_format

                        row += 1

                    # Totals
                    tb_ws[f"A{row}"] = "TOTALS"
                    tb_ws[f"A{row}"].font = header_font
                    tb_ws[f"C{row}"] = float(trial_balance["total_debits"])
                    tb_ws[f"C{row}"].number_format = currency_format
                    tb_ws[f"C{row}"].font = header_font
                    tb_ws[f"D{row}"] = float(trial_balance["total_credits"])
                    tb_ws[f"D{row}"].number_format = currency_format
                    tb_ws[f"D{row}"].font = header_font

            # Auto-adjust column widths
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (TypeError, AttributeError):
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Failed to export financial reports to Excel: {str(e)}")
            raise
            if not reports:
                raise ValueError("No financial reports data available")

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Define styles
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            currency_format = "#,##0.00"

            # Balance Sheet worksheet
            if reports.get("balance_sheet"):
                bs_ws = wb.create_sheet("Balance Sheet")
                bs_ws["A1"] = f"{tenant.company_name} - Balance Sheet"
                bs_ws["A1"].font = title_font
                bs_ws["A2"] = f"As of {end_date.strftime('%B %d, %Y')}"

                row = 4
                balance_sheet = reports["balance_sheet"]

                # Assets
                bs_ws[f"A{row}"] = "ASSETS"
                bs_ws[f"A{row}"].font = header_font
                row += 1

                # Current Assets
                if balance_sheet["assets"]["current_assets"]:
                    bs_ws[f"A{row}"] = "Current Assets:"
                    bs_ws[f"A{row}"].font = header_font
                    row += 1

                    for asset in balance_sheet["assets"]["current_assets"]:
                        bs_ws[f"A{row}"] = f"{asset['code']} - {asset['name']}"
                        bs_ws[f"B{row}"] = float(asset["balance"])
                        bs_ws[f"B{row}"].number_format = currency_format
                        row += 1

                # Total Assets
                bs_ws[f"A{row}"] = "Total Assets"
                bs_ws[f"A{row}"].font = header_font
                bs_ws[f"B{row}"] = float(balance_sheet["assets"]["total_assets"])
                bs_ws[f"B{row}"].number_format = currency_format
                bs_ws[f"B{row}"].font = header_font

            # Income Statement worksheet
            if reports.get("income_statement"):
                is_ws = wb.create_sheet("Income Statement")
                is_ws["A1"] = f"{tenant.company_name} - Income Statement"
                is_ws["A1"].font = title_font
                is_ws["A2"] = (
                    f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
                )

                row = 4
                income_statement = reports["income_statement"]

                # Revenue
                if income_statement["revenue"]["operating_revenue"]:
                    is_ws[f"A{row}"] = "REVENUE"
                    is_ws[f"A{row}"].font = header_font
                    row += 1

                    for revenue in income_statement["revenue"]["operating_revenue"]:
                        is_ws[f"A{row}"] = f"{revenue['code']} - {revenue['name']}"
                        is_ws[f"B{row}"] = float(revenue["balance"])
                        is_ws[f"B{row}"].number_format = currency_format
                        row += 1

                # Net Income
                row += 1
                is_ws[f"A{row}"] = "Net Income"
                is_ws[f"A{row}"].font = header_font
                is_ws[f"B{row}"] = float(income_statement["net_income"])
                is_ws[f"B{row}"].number_format = currency_format
                is_ws[f"B{row}"].font = header_font

            # Trial Balance worksheet
            if reports.get("trial_balance"):
                tb_ws = wb.create_sheet("Trial Balance")
                tb_ws["A1"] = f"{tenant.company_name} - Trial Balance"
                tb_ws["A1"].font = title_font
                tb_ws["A2"] = f"As of {end_date.strftime('%B %d, %Y')}"

                # Headers
                tb_ws["A4"] = "Account Code"
                tb_ws["B4"] = "Account Name"
                tb_ws["C4"] = "Debit"
                tb_ws["D4"] = "Credit"

                for col in ["A4", "B4", "C4", "D4"]:
                    tb_ws[col].font = header_font

                row = 5
                trial_balance = reports["trial_balance"]

                for account in trial_balance["accounts"]:
                    tb_ws[f"A{row}"] = account["code"]
                    tb_ws[f"B{row}"] = account["name"]

                    if account["debit_balance"] > 0:
                        tb_ws[f"C{row}"] = float(account["debit_balance"])
                        tb_ws[f"C{row}"].number_format = currency_format

                    if account["credit_balance"] > 0:
                        tb_ws[f"D{row}"] = float(account["credit_balance"])
                        tb_ws[f"D{row}"].number_format = currency_format

                    row += 1

                # Totals
                tb_ws[f"A{row}"] = "TOTALS"
                tb_ws[f"A{row}"].font = header_font
                tb_ws[f"C{row}"] = float(trial_balance["total_debits"])
                tb_ws[f"C{row}"].number_format = currency_format
                tb_ws[f"C{row}"].font = header_font
                tb_ws[f"D{row}"] = float(trial_balance["total_credits"])
                tb_ws[f"D{row}"].number_format = currency_format
                tb_ws[f"D{row}"].font = header_font

            # Auto-adjust column widths
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Failed to export financial reports to Excel: {str(e)}")
            raise


# ============================================================================
# Invoice Service (Task 3.5)
# ============================================================================


class InvoiceService:
    """
    Service class for handling invoice operations and automatic journal entries.

    Provides methods for creating invoices, recording payments, applying credit memos,
    and generating automatic journal entries for accounts receivable.

    Requirements: 3.1, 3.2, 3.3, 3.4, 3.6, 3.7, 3.8
    """

    @staticmethod
    @transaction.atomic
    def create_invoice_journal_entry(invoice, user: User) -> Optional[JournalEntryModel]:
        """
        Create automatic journal entry when invoice is created.

        Journal Entry:
        DR Accounts Receivable (Asset)    invoice.total
        CR Revenue (Revenue)               invoice.subtotal
        CR Sales Tax Payable (Liability)   invoice.tax

        Requirements: 3.2
        """
        try:
            # Get the entity for this tenant
            jewelry_entity = JewelryEntity.objects.get(tenant=invoice.tenant)
            entity = jewelry_entity.ledger_entity

            # Get the chart of accounts
            coa = ChartOfAccountModel.objects.filter(entity=entity).first()
            if not coa:
                logger.error(f"No chart of accounts found for tenant: {invoice.tenant}")
                return None

            # Get required accounts
            # Accounts Receivable (Asset)
            ar_account = AccountModel.objects.filter(
                coa_model=coa,
                role__istartswith="asset_ca_",
                name__icontains="receivable",
                active=True,
            ).first()

            # Revenue account
            revenue_account = AccountModel.objects.filter(
                coa_model=coa, role__istartswith="in_sales", active=True
            ).first()

            # Sales Tax Payable (Liability)
            tax_account = AccountModel.objects.filter(
                coa_model=coa, role__istartswith="lia_cl_", name__icontains="tax", active=True
            ).first()

            if not ar_account or not revenue_account:
                logger.error(
                    f"Required accounts not found for invoice journal entry. "
                    f"AR: {ar_account}, Revenue: {revenue_account}"
                )
                return None

            # Get the ledger
            ledger = entity.ledgermodel_set.first()
            if not ledger:
                logger.error(f"No ledger found for entity: {entity}")
                return None

            # Create journal entry
            je = JournalEntryModel.objects.create(
                ledger=ledger,
                description=f"Invoice {invoice.invoice_number} - {invoice.customer}",
                date=invoice.invoice_date,
                activity="op",  # Operating activity
                origin="invoice",
                posted=True,  # Auto-post invoice journal entries
            )

            # Create transactions
            # DR Accounts Receivable
            TransactionModel.objects.create(
                journal_entry=je,
                account=ar_account,
                tx_type="debit",
                amount=invoice.total,
                description=f"Invoice {invoice.invoice_number}",
            )

            # CR Revenue
            TransactionModel.objects.create(
                journal_entry=je,
                account=revenue_account,
                tx_type="credit",
                amount=invoice.subtotal,
                description=f"Sales revenue - Invoice {invoice.invoice_number}",
            )

            # CR Sales Tax Payable (if tax > 0)
            if invoice.tax > Decimal("0.00") and tax_account:
                TransactionModel.objects.create(
                    journal_entry=je,
                    account=tax_account,
                    tx_type="credit",
                    amount=invoice.tax,
                    description=f"Sales tax - Invoice {invoice.invoice_number}",
                )

            # Link journal entry to invoice
            invoice.journal_entry = je
            invoice.save(update_fields=["journal_entry"])

            # Log to audit trail
            from apps.core.audit_models import AuditLog

            AuditLog.objects.create(
                tenant=invoice.tenant,
                user=user,
                category="ACCOUNTING",
                action="CREATE_INVOICE_JE",
                severity="INFO",
                description=f"Created journal entry for invoice {invoice.invoice_number}",
                before_value=None,
                after_value=f"JE: {je.uuid}, Amount: ${invoice.total:,.2f}",
            )

            logger.info(
                f"Created journal entry {je.uuid} for invoice {invoice.invoice_number} "
                f"(${invoice.total:,.2f})"
            )

            return je

        except Exception as e:
            logger.error(
                f"Failed to create journal entry for invoice {invoice.invoice_number}: {str(e)}"
            )
            raise

    @staticmethod
    @transaction.atomic
    def create_payment_journal_entry(payment, user: User) -> Optional[JournalEntryModel]:
        """
        Create automatic journal entry when payment is received.

        Journal Entry:
        DR Cash/Bank (Asset)                payment.amount
        CR Accounts Receivable (Asset)     payment.amount

        Requirements: 3.4
        """
        try:
            invoice = payment.invoice
            tenant = payment.tenant

            # Get the entity for this tenant
            jewelry_entity = JewelryEntity.objects.get(tenant=tenant)
            entity = jewelry_entity.ledger_entity

            # Get the chart of accounts
            coa = ChartOfAccountModel.objects.filter(entity=entity).first()
            if not coa:
                logger.error(f"No chart of accounts found for tenant: {tenant}")
                return None

            # Get required accounts
            # Cash/Bank account (Asset)
            cash_account = AccountModel.objects.filter(
                coa_model=coa, role__istartswith="asset_ca_cash", active=True
            ).first()

            # Accounts Receivable (Asset)
            ar_account = AccountModel.objects.filter(
                coa_model=coa,
                role__istartswith="asset_ca_",
                name__icontains="receivable",
                active=True,
            ).first()

            if not cash_account or not ar_account:
                logger.error(
                    f"Required accounts not found for payment journal entry. "
                    f"Cash: {cash_account}, AR: {ar_account}"
                )
                return None

            # Get the ledger
            ledger = entity.ledgermodel_set.first()
            if not ledger:
                logger.error(f"No ledger found for entity: {entity}")
                return None

            # Create journal entry
            je = JournalEntryModel.objects.create(
                ledger=ledger,
                description=f"Payment for Invoice {invoice.invoice_number} - {invoice.customer}",
                date=payment.payment_date,
                activity="op",  # Operating activity
                origin="payment",
                posted=True,  # Auto-post payment journal entries
            )

            # Create transactions
            # DR Cash/Bank
            TransactionModel.objects.create(
                journal_entry=je,
                account=cash_account,
                tx_type="debit",
                amount=payment.amount,
                description=f"Payment received - Invoice {invoice.invoice_number}",
            )

            # CR Accounts Receivable
            TransactionModel.objects.create(
                journal_entry=je,
                account=ar_account,
                tx_type="credit",
                amount=payment.amount,
                description=f"Payment applied - Invoice {invoice.invoice_number}",
            )

            # Link journal entry to payment
            payment.journal_entry = je
            payment.save(update_fields=["journal_entry"])

            # Log to audit trail
            from apps.core.audit_models import AuditLog

            AuditLog.objects.create(
                tenant=tenant,
                user=user,
                category="ACCOUNTING",
                action="CREATE_PAYMENT_JE",
                severity="INFO",
                description=f"Created journal entry for payment on invoice {invoice.invoice_number}",
                before_value=None,
                after_value=f"JE: {je.uuid}, Amount: ${payment.amount:,.2f}",
            )

            logger.info(
                f"Created journal entry {je.uuid} for payment on invoice {invoice.invoice_number} "
                f"(${payment.amount:,.2f})"
            )

            return je

        except Exception as e:
            logger.error(
                f"Failed to create journal entry for payment on invoice {payment.invoice.invoice_number}: {str(e)}"
            )
            raise

    @staticmethod
    @transaction.atomic
    def create_credit_memo_journal_entry(credit_memo, user: User) -> Optional[JournalEntryModel]:
        """
        Create automatic journal entry when credit memo is created.

        Journal Entry:
        DR Sales Returns/Allowances (Contra-Revenue)    credit_memo.amount
        CR Accounts Receivable (Asset)                  credit_memo.amount

        Requirements: 3.6
        """
        try:
            tenant = credit_memo.tenant

            # Get the entity for this tenant
            jewelry_entity = JewelryEntity.objects.get(tenant=tenant)
            entity = jewelry_entity.ledger_entity

            # Get the chart of accounts
            coa = ChartOfAccountModel.objects.filter(entity=entity).first()
            if not coa:
                logger.error(f"No chart of accounts found for tenant: {tenant}")
                return None

            # Get required accounts
            # Sales Returns/Allowances (Contra-Revenue) - try to find or use revenue account
            returns_account = AccountModel.objects.filter(
                coa_model=coa, name__icontains="return", active=True
            ).first()

            if not returns_account:
                # Fallback to revenue account if no returns account exists
                returns_account = AccountModel.objects.filter(
                    coa_model=coa, role__istartswith="in_sales", active=True
                ).first()

            # Accounts Receivable (Asset)
            ar_account = AccountModel.objects.filter(
                coa_model=coa,
                role__istartswith="asset_ca_",
                name__icontains="receivable",
                active=True,
            ).first()

            if not returns_account or not ar_account:
                logger.error(
                    f"Required accounts not found for credit memo journal entry. "
                    f"Returns: {returns_account}, AR: {ar_account}"
                )
                return None

            # Get the ledger
            ledger = entity.ledgermodel_set.first()
            if not ledger:
                logger.error(f"No ledger found for entity: {entity}")
                return None

            # Create journal entry
            je = JournalEntryModel.objects.create(
                ledger=ledger,
                description=f"Credit Memo {credit_memo.credit_memo_number} - {credit_memo.customer}",
                date=credit_memo.credit_date,
                activity="op",  # Operating activity
                origin="credit_memo",
                posted=True,  # Auto-post credit memo journal entries
            )

            # Create transactions
            # DR Sales Returns/Allowances
            TransactionModel.objects.create(
                journal_entry=je,
                account=returns_account,
                tx_type="debit",
                amount=credit_memo.amount,
                description=f"Credit memo {credit_memo.credit_memo_number} - {credit_memo.reason}",
            )

            # CR Accounts Receivable
            TransactionModel.objects.create(
                journal_entry=je,
                account=ar_account,
                tx_type="credit",
                amount=credit_memo.amount,
                description=f"Credit memo {credit_memo.credit_memo_number}",
            )

            # Link journal entry to credit memo
            credit_memo.journal_entry = je
            credit_memo.save(update_fields=["journal_entry"])

            # Log to audit trail
            from apps.core.audit_models import AuditLog

            AuditLog.objects.create(
                tenant=tenant,
                user=user,
                category="ACCOUNTING",
                action="CREATE_CREDIT_MEMO_JE",
                severity="INFO",
                description=f"Created journal entry for credit memo {credit_memo.credit_memo_number}",
                before_value=None,
                after_value=f"JE: {je.uuid}, Amount: ${credit_memo.amount:,.2f}",
            )

            logger.info(
                f"Created journal entry {je.uuid} for credit memo {credit_memo.credit_memo_number} "
                f"(${credit_memo.amount:,.2f})"
            )

            return je

        except Exception as e:
            logger.error(
                f"Failed to create journal entry for credit memo {credit_memo.credit_memo_number}: {str(e)}"
            )
            raise

    @staticmethod
    def check_customer_credit_limit(customer, additional_amount: Decimal = Decimal("0.00")) -> Dict:
        """
        Check if customer is within credit limit.

        Returns dict with:
        - within_limit: bool
        - current_outstanding: Decimal
        - credit_limit: Decimal
        - available_credit: Decimal
        - warning_message: str (if applicable)

        Requirements: 3.7
        """
        from .invoice_models import Invoice

        try:
            # Get customer's credit limit
            credit_limit = getattr(customer, "credit_limit", Decimal("0.00"))

            # If no credit limit set, allow unlimited credit
            if credit_limit <= Decimal("0.00"):
                return {
                    "within_limit": True,
                    "current_outstanding": Decimal("0.00"),
                    "credit_limit": Decimal("0.00"),
                    "available_credit": None,  # Unlimited
                    "warning_message": None,
                }

            # Calculate current outstanding balance
            outstanding = Invoice.objects.filter(
                customer=customer, status__in=["SENT", "PARTIALLY_PAID", "OVERDUE"]
            ).aggregate(total=models.Sum(models.F("total") - models.F("amount_paid")))[
                "total"
            ] or Decimal(
                "0.00"
            )

            # Add the additional amount (for new invoice)
            total_outstanding = outstanding + additional_amount

            # Calculate available credit
            available_credit = credit_limit - total_outstanding

            # Determine if within limit
            within_limit = total_outstanding <= credit_limit

            # Generate warning message if needed
            warning_message = None
            if not within_limit:
                warning_message = (
                    f"Customer exceeds credit limit. "
                    f"Outstanding: ${total_outstanding:,.2f}, "
                    f"Limit: ${credit_limit:,.2f}, "
                    f"Over by: ${abs(available_credit):,.2f}"
                )
            elif available_credit < credit_limit * Decimal("0.1"):  # Less than 10% available
                warning_message = (
                    f"Customer approaching credit limit. "
                    f"Available credit: ${available_credit:,.2f} "
                    f"of ${credit_limit:,.2f}"
                )

            return {
                "within_limit": within_limit,
                "current_outstanding": outstanding,
                "credit_limit": credit_limit,
                "available_credit": available_credit,
                "warning_message": warning_message,
            }

        except Exception as e:
            logger.error(f"Failed to check credit limit for customer {customer}: {str(e)}")
            return {
                "within_limit": True,  # Allow on error
                "current_outstanding": Decimal("0.00"),
                "credit_limit": Decimal("0.00"),
                "available_credit": None,
                "warning_message": f"Error checking credit limit: {str(e)}",
            }

    @staticmethod
    def get_aged_receivables(tenant: Tenant, as_of_date: Optional[date] = None) -> Dict:
        """
        Generate aged receivables report.

        Groups outstanding invoices by aging buckets:
        - Current (not yet due)
        - 1-30 days overdue
        - 31-60 days overdue
        - 61-90 days overdue
        - 90+ days overdue

        Requirements: 3.5
        """
        from .invoice_models import Invoice

        if as_of_date is None:
            as_of_date = date.today()

        try:
            # Get all unpaid invoices for tenant
            invoices = Invoice.objects.filter(
                tenant=tenant, status__in=["SENT", "PARTIALLY_PAID", "OVERDUE"]
            ).select_related("customer")

            # Initialize buckets
            buckets = {
                "current": {"invoices": [], "total": Decimal("0.00")},
                "1_30": {"invoices": [], "total": Decimal("0.00")},
                "31_60": {"invoices": [], "total": Decimal("0.00")},
                "61_90": {"invoices": [], "total": Decimal("0.00")},
                "90_plus": {"invoices": [], "total": Decimal("0.00")},
            }

            # Categorize invoices
            for invoice in invoices:
                amount_due = invoice.amount_due
                days_overdue = (as_of_date - invoice.due_date).days

                invoice_data = {
                    "invoice": invoice,
                    "invoice_number": invoice.invoice_number,
                    "customer": invoice.customer,
                    "invoice_date": invoice.invoice_date,
                    "due_date": invoice.due_date,
                    "amount_due": amount_due,
                    "days_overdue": max(0, days_overdue),
                }

                if days_overdue < 0:
                    # Not yet due
                    buckets["current"]["invoices"].append(invoice_data)
                    buckets["current"]["total"] += amount_due
                elif days_overdue <= 30:
                    buckets["1_30"]["invoices"].append(invoice_data)
                    buckets["1_30"]["total"] += amount_due
                elif days_overdue <= 60:
                    buckets["31_60"]["invoices"].append(invoice_data)
                    buckets["31_60"]["total"] += amount_due
                elif days_overdue <= 90:
                    buckets["61_90"]["invoices"].append(invoice_data)
                    buckets["61_90"]["total"] += amount_due
                else:
                    buckets["90_plus"]["invoices"].append(invoice_data)
                    buckets["90_plus"]["total"] += amount_due

            # Calculate grand total
            grand_total = sum(bucket["total"] for bucket in buckets.values())

            return {
                "as_of_date": as_of_date,
                "buckets": buckets,
                "grand_total": grand_total,
                "invoice_count": invoices.count(),
            }

        except Exception as e:
            logger.error(f"Failed to generate aged receivables report: {str(e)}")
            raise

    @staticmethod
    def get_customer_statement(customer, start_date: date, end_date: date) -> Dict:
        """
        Generate customer statement showing all transactions.

        Requirements: 3.8
        """
        from .invoice_models import CreditMemo, Invoice, InvoicePayment

        try:
            # Get all invoices for customer in date range
            invoices = Invoice.objects.filter(
                customer=customer, invoice_date__range=[start_date, end_date]
            ).order_by("invoice_date")

            # Get all payments in date range
            payments = (
                InvoicePayment.objects.filter(
                    invoice__customer=customer, payment_date__range=[start_date, end_date]
                )
                .select_related("invoice")
                .order_by("payment_date")
            )

            # Get all credit memos in date range
            credit_memos = CreditMemo.objects.filter(
                customer=customer, credit_date__range=[start_date, end_date]
            ).order_by("credit_date")

            # Calculate beginning balance (invoices before start_date)
            beginning_balance = Invoice.objects.filter(
                customer=customer, invoice_date__lt=start_date
            ).aggregate(total=models.Sum(models.F("total") - models.F("amount_paid")))[
                "total"
            ] or Decimal(
                "0.00"
            )

            # Build transaction list
            transactions = []

            for invoice in invoices:
                transactions.append(
                    {
                        "date": invoice.invoice_date,
                        "type": "Invoice",
                        "reference": invoice.invoice_number,
                        "description": f"Invoice {invoice.invoice_number}",
                        "amount": invoice.total,
                        "balance_change": invoice.total,
                    }
                )

            for payment in payments:
                transactions.append(
                    {
                        "date": payment.payment_date,
                        "type": "Payment",
                        "reference": payment.reference_number or "-",
                        "description": f"Payment for Invoice {payment.invoice.invoice_number}",
                        "amount": payment.amount,
                        "balance_change": -payment.amount,
                    }
                )

            for credit_memo in credit_memos:
                transactions.append(
                    {
                        "date": credit_memo.credit_date,
                        "type": "Credit Memo",
                        "reference": credit_memo.credit_memo_number,
                        "description": f"Credit: {credit_memo.reason}",
                        "amount": credit_memo.amount,
                        "balance_change": -credit_memo.amount,
                    }
                )

            # Sort transactions by date
            transactions.sort(key=lambda x: x["date"])

            # Calculate running balance
            running_balance = beginning_balance
            for txn in transactions:
                running_balance += txn["balance_change"]
                txn["running_balance"] = running_balance

            # Calculate ending balance
            ending_balance = running_balance

            return {
                "customer": customer,
                "start_date": start_date,
                "end_date": end_date,
                "beginning_balance": beginning_balance,
                "transactions": transactions,
                "ending_balance": ending_balance,
                "total_invoiced": sum(t["amount"] for t in transactions if t["type"] == "Invoice"),
                "total_paid": sum(t["amount"] for t in transactions if t["type"] == "Payment"),
                "total_credits": sum(
                    t["amount"] for t in transactions if t["type"] == "Credit Memo"
                ),
            }

        except Exception as e:
            logger.error(f"Failed to generate customer statement: {str(e)}")
            raise


class BankReconciliationService:
    """
    Service class for handling bank reconciliation operations.

    Provides methods for starting reconciliations, marking transactions as reconciled,
    auto-matching transactions, and completing reconciliations with proper audit trails.
    """

    @staticmethod
    def get_unreconciled_transactions(bank_account, as_of_date=None):
        """
        Get all unreconciled transactions for a bank account.

        Args:
            bank_account: BankAccount instance
            as_of_date: Optional date to filter transactions up to

        Returns:
            QuerySet of unreconciled BankTransaction objects
        """
        from .bank_models import BankTransaction

        queryset = BankTransaction.objects.filter(bank_account=bank_account, is_reconciled=False)

        if as_of_date:
            queryset = queryset.filter(transaction_date__lte=as_of_date)

        return queryset.order_by("transaction_date")

    @staticmethod
    def calculate_book_balance(bank_account, as_of_date):
        """
        Calculate the book balance for a bank account as of a specific date.

        Args:
            bank_account: BankAccount instance
            as_of_date: Date to calculate balance as of

        Returns:
            Decimal: Book balance as of the date
        """
        from .bank_models import BankTransaction

        # Start with opening balance
        balance = bank_account.opening_balance

        # Add all transactions up to the date
        transactions = BankTransaction.objects.filter(
            bank_account=bank_account, transaction_date__lte=as_of_date
        )

        for txn in transactions:
            if txn.transaction_type == "CREDIT":
                balance += txn.amount
            else:  # DEBIT
                balance -= txn.amount

        return balance

    @staticmethod
    def start_reconciliation(
        bank_account,
        statement_date,
        ending_balance,
        user,
        beginning_balance=None,
        period_start_date=None,
        period_end_date=None,
    ):
        """
        Start a new bank reconciliation session.

        Args:
            bank_account: BankAccount instance to reconcile
            statement_date: Date of the bank statement
            ending_balance: Ending balance from bank statement
            user: User starting the reconciliation
            beginning_balance: Optional beginning balance from statement
            period_start_date: Optional start date of reconciliation period
            period_end_date: Optional end date of reconciliation period

        Returns:
            BankReconciliation instance
        """
        from apps.core.audit_models import AuditLog

        from .bank_models import BankReconciliation

        try:
            with transaction.atomic():
                # Calculate book balances
                book_ending_balance = BankReconciliationService.calculate_book_balance(
                    bank_account, statement_date
                )

                # Use last reconciled balance as beginning balance if not provided
                if beginning_balance is None:
                    beginning_balance = bank_account.reconciled_balance

                book_beginning_balance = bank_account.reconciled_balance

                # Create reconciliation
                reconciliation = BankReconciliation.objects.create(
                    tenant=bank_account.tenant,
                    bank_account=bank_account,
                    reconciliation_date=statement_date,
                    period_start_date=period_start_date,
                    period_end_date=period_end_date or statement_date,
                    statement_beginning_balance=beginning_balance,
                    statement_ending_balance=ending_balance,
                    book_beginning_balance=book_beginning_balance,
                    book_ending_balance=book_ending_balance,
                    status="IN_PROGRESS",
                    created_by=user,
                )

                # Calculate initial variance
                reconciliation.calculate_variance()

                # Log audit trail
                AuditLog.objects.create(
                    tenant=bank_account.tenant,
                    user=user,
                    category=AuditLog.CATEGORY_DATA,
                    action=AuditLog.ACTION_CREATE,
                    severity=AuditLog.SEVERITY_INFO,
                    description=f"Started bank reconciliation for {bank_account.account_name} as of {statement_date}",
                    object_id=str(reconciliation.id),
                    new_values={
                        "reconciliation_id": str(reconciliation.id),
                        "bank_account": bank_account.account_name,
                        "statement_date": str(statement_date),
                        "ending_balance": str(ending_balance),
                    },
                )

                logger.info(
                    f"Reconciliation started for {bank_account.account_name} by {user.username}"
                )

                return reconciliation

        except Exception as e:
            logger.error(f"Failed to start reconciliation: {str(e)}")
            raise

    @staticmethod
    def mark_reconciled(transaction_ids, reconciliation, user):
        """
        Mark transactions as reconciled.

        Args:
            transaction_ids: List of transaction IDs to mark as reconciled
            reconciliation: BankReconciliation instance
            user: User performing the reconciliation

        Returns:
            int: Number of transactions marked as reconciled
        """
        from apps.core.audit_models import AuditLog

        from .bank_models import BankTransaction

        try:
            with transaction.atomic():
                # Get transactions with tenant filtering
                transactions = BankTransaction.objects.filter(
                    id__in=transaction_ids,
                    tenant=reconciliation.tenant,
                    bank_account=reconciliation.bank_account,
                    is_reconciled=False,
                )

                count = 0
                for txn in transactions:
                    txn.mark_reconciled(reconciliation, user)
                    count += 1

                    # Log audit trail for each transaction
                    AuditLog.objects.create(
                        tenant=reconciliation.tenant,
                        user=user,
                        category="ACCOUNTING",
                        action="TRANSACTION_RECONCILED",
                        severity="INFO",
                        description=f"Marked transaction {txn.description} as reconciled",
                        before_value="unreconciled",
                        after_value="reconciled",
                    )

                # Update reconciliation totals
                reconciliation.calculate_totals()
                reconciliation.calculate_variance()

                logger.info(
                    f"Marked {count} transactions as reconciled for reconciliation {reconciliation.id}"
                )

                return count

        except Exception as e:
            logger.error(f"Failed to mark transactions as reconciled: {str(e)}")
            raise

    @staticmethod
    def unreconcile_transaction(transaction, reason, user):
        """
        Unreconcile a transaction.

        Args:
            transaction: BankTransaction instance to unreconcile
            reason: Reason for unreconciling
            user: User performing the unreconciliation
        """
        from apps.core.audit_models import AuditLog

        try:
            with transaction.atomic():
                reconciliation = transaction.reconciliation

                # Unreconcile the transaction
                transaction.unreconcile(reason, user)

                # Update reconciliation totals if still in progress
                if reconciliation and reconciliation.status == "IN_PROGRESS":
                    reconciliation.calculate_totals()
                    reconciliation.calculate_variance()

                # Log audit trail
                AuditLog.objects.create(
                    tenant=transaction.tenant,
                    user=user,
                    category="ACCOUNTING",
                    action="TRANSACTION_UNRECONCILED",
                    severity="WARNING",
                    description=f"Unreconciled transaction: {transaction.description}. Reason: {reason}",
                    before_value="reconciled",
                    after_value="unreconciled",
                )

                logger.info(
                    f"Transaction {transaction.id} unreconciled by {user.username}. Reason: {reason}"
                )

        except Exception as e:
            logger.error(f"Failed to unreconcile transaction: {str(e)}")
            raise

    @staticmethod
    def complete_reconciliation(reconciliation, user):
        """
        Complete a bank reconciliation.

        Args:
            reconciliation: BankReconciliation instance to complete
            user: User completing the reconciliation

        Returns:
            BankReconciliation: Completed reconciliation
        """
        from apps.core.audit_models import AuditLog

        try:
            with transaction.atomic():
                # Complete the reconciliation
                reconciliation.complete(user)

                # Log audit trail
                AuditLog.objects.create(
                    tenant=reconciliation.tenant,
                    user=user,
                    category="ACCOUNTING",
                    action="RECONCILIATION_COMPLETED",
                    severity="INFO",
                    description=f"Completed bank reconciliation for {reconciliation.bank_account.account_name} as of {reconciliation.reconciliation_date}. Variance: {reconciliation.variance}",
                    after_value=f"balanced={reconciliation.is_balanced}, variance={reconciliation.variance}",
                )

                logger.info(
                    f"Reconciliation {reconciliation.id} completed by {user.username}. Balanced: {reconciliation.is_balanced}"
                )

                return reconciliation

        except Exception as e:
            logger.error(f"Failed to complete reconciliation: {str(e)}")
            raise

    @staticmethod
    def cancel_reconciliation(reconciliation, reason, user):
        """
        Cancel a bank reconciliation.

        Args:
            reconciliation: BankReconciliation instance to cancel
            reason: Reason for cancellation
            user: User cancelling the reconciliation
        """
        from apps.core.audit_models import AuditLog

        try:
            with transaction.atomic():
                # Cancel the reconciliation
                reconciliation.cancel(reason)

                # Log audit trail
                AuditLog.objects.create(
                    tenant=reconciliation.tenant,
                    user=user,
                    category="ACCOUNTING",
                    action="RECONCILIATION_CANCELLED",
                    severity="WARNING",
                    description=f"Cancelled bank reconciliation for {reconciliation.bank_account.account_name}. Reason: {reason}",
                    before_value=f"status={reconciliation.status}",
                    after_value="status=CANCELLED",
                )

                logger.info(
                    f"Reconciliation {reconciliation.id} cancelled by {user.username}. Reason: {reason}"
                )

        except Exception as e:
            logger.error(f"Failed to cancel reconciliation: {str(e)}")
            raise

    @staticmethod
    def auto_match_transactions(reconciliation):
        """
        Automatically match bank transactions with journal entries.

        Args:
            reconciliation: BankReconciliation instance

        Returns:
            dict: Statistics about matching (matched_count, suggestions)
        """
        from .bank_models import BankTransaction

        try:
            matched_count = 0
            suggestions = []

            # Get unreconciled transactions for this reconciliation
            unreconciled = BankTransaction.objects.filter(
                bank_account=reconciliation.bank_account,
                is_reconciled=False,
                transaction_date__lte=reconciliation.reconciliation_date,
            )

            for txn in unreconciled:
                # Try to find matching journal entries
                matches = BankReconciliationService.suggest_matches(txn)

                if matches:
                    # If we have a high-confidence match (score > 0.9), auto-match
                    best_match = matches[0]
                    if best_match["confidence"] > 0.9:
                        txn.match_journal_entry(best_match["journal_entry"])
                        matched_count += 1
                    else:
                        # Add to suggestions for manual review
                        suggestions.append({"transaction": txn, "matches": matches})

            logger.info(
                f"Auto-matched {matched_count} transactions for reconciliation {reconciliation.id}"
            )

            return {"matched_count": matched_count, "suggestions": suggestions}

        except Exception as e:
            logger.error(f"Failed to auto-match transactions: {str(e)}")
            return {"matched_count": 0, "suggestions": []}

    @staticmethod
    def suggest_matches(transaction):
        """
        Suggest potential journal entry matches for a bank transaction.

        Args:
            transaction: BankTransaction instance

        Returns:
            list: List of potential matches with confidence scores
        """
        from datetime import timedelta

        try:
            matches = []

            # Search for journal entries within ±3 days of transaction date
            date_range_start = transaction.transaction_date - timedelta(days=3)
            date_range_end = transaction.transaction_date + timedelta(days=3)

            # Get journal entries from django-ledger
            # Note: This is a simplified version - in production, you'd want to
            # search through TransactionModel entries that match the amount
            from django_ledger.models import TransactionModel

            # Find transactions with matching amount
            potential_matches = TransactionModel.objects.filter(
                journal_entry__ledger__entity__jewelryentity__tenant=transaction.tenant,
                journal_entry__timestamp__date__gte=date_range_start,
                journal_entry__timestamp__date__lte=date_range_end,
                amount=transaction.amount,
            ).select_related("journal_entry", "account")

            for je_txn in potential_matches:
                confidence = 0.5  # Base confidence for amount match

                # Increase confidence for exact date match
                if je_txn.journal_entry.timestamp.date() == transaction.transaction_date:
                    confidence += 0.3

                # Increase confidence for description similarity
                if (
                    transaction.description.lower() in je_txn.description.lower()
                    or je_txn.description.lower() in transaction.description.lower()
                ):
                    confidence += 0.2

                matches.append(
                    {
                        "journal_entry": je_txn.journal_entry,
                        "transaction": je_txn,
                        "confidence": min(confidence, 1.0),
                        "reason": f"Amount match: {transaction.amount}, Date: {je_txn.journal_entry.timestamp.date()}",
                    }
                )

            # Sort by confidence (highest first)
            matches.sort(key=lambda x: x["confidence"], reverse=True)

            return matches[:5]  # Return top 5 matches

        except Exception as e:
            logger.error(f"Failed to suggest matches: {str(e)}")
            return []

    @staticmethod
    def create_adjusting_entry(reconciliation, description, amount, account_code, is_debit, user):
        """
        Create an adjusting journal entry during reconciliation.

        Args:
            reconciliation: BankReconciliation instance
            description: Description of the adjustment
            amount: Amount of the adjustment
            account_code: Account code for the offsetting entry
            is_debit: True if bank account should be debited, False if credited
            user: User creating the adjustment

        Returns:
            JournalEntryModel: Created journal entry
        """
        from apps.core.audit_models import AuditLog

        try:
            with transaction.atomic():
                # Get tenant's accounting entity
                jewelry_entity = JewelryEntity.objects.get(tenant=reconciliation.tenant)
                entity = jewelry_entity.ledger_entity
                ledger = entity.ledgermodel_set.first()

                if not ledger:
                    raise ValueError(f"No ledger found for entity {entity}")

                # Create journal entry
                journal_entry = JournalEntryModel.objects.create(
                    ledger=ledger,
                    description=f"Bank Reconciliation Adjustment: {description}",
                    posted=False,
                )

                # Get bank account from chart of accounts
                # Note: This assumes the bank account has a corresponding GL account
                bank_gl_account = AccountModel.objects.filter(
                    coa_model__entity=entity,
                    name__icontains=reconciliation.bank_account.account_name,
                ).first()

                if not bank_gl_account:
                    # Fallback to default cash account
                    config = AccountingConfiguration.objects.get(tenant=reconciliation.tenant)
                    bank_gl_account = AccountModel.objects.get(
                        coa_model__entity=entity, code=config.default_cash_account
                    )

                # Get offsetting account
                offset_account = AccountModel.objects.get(
                    coa_model__entity=entity, code=account_code
                )

                # Create transactions
                if is_debit:
                    # Debit bank account
                    TransactionModel.objects.create(
                        journal_entry=journal_entry,
                        account=bank_gl_account,
                        amount=amount,
                        tx_type="debit",
                        description=description,
                    )
                    # Credit offset account
                    TransactionModel.objects.create(
                        journal_entry=journal_entry,
                        account=offset_account,
                        amount=amount,
                        tx_type="credit",
                        description=description,
                    )
                else:
                    # Credit bank account
                    TransactionModel.objects.create(
                        journal_entry=journal_entry,
                        account=bank_gl_account,
                        amount=amount,
                        tx_type="credit",
                        description=description,
                    )
                    # Debit offset account
                    TransactionModel.objects.create(
                        journal_entry=journal_entry,
                        account=offset_account,
                        amount=amount,
                        tx_type="debit",
                        description=description,
                    )

                # Post the journal entry
                journal_entry.posted = True
                journal_entry.save()

                # Update reconciliation adjustment total
                reconciliation.total_adjustments += amount
                reconciliation.save(update_fields=["total_adjustments", "updated_at"])

                # Log audit trail
                AuditLog.objects.create(
                    tenant=reconciliation.tenant,
                    user=user,
                    category="ACCOUNTING",
                    action="ADJUSTING_ENTRY_CREATED",
                    severity="INFO",
                    description=f"Created adjusting entry during reconciliation: {description}",
                    after_value=f"amount={amount}, journal_entry={journal_entry.id}",
                )

                logger.info(
                    f"Adjusting entry created for reconciliation {reconciliation.id} by {user.username}"
                )

                return journal_entry

        except Exception as e:
            logger.error(f"Failed to create adjusting entry: {str(e)}")
            raise

    @staticmethod
    def export_reconciliation_report_to_pdf(reconciliation, deposits, withdrawals):
        """
        Export bank reconciliation report to PDF format.

        Args:
            reconciliation: BankReconciliation instance
            deposits: List of deposit transactions (CREDIT)
            withdrawals: List of withdrawal transactions (DEBIT)

        Returns:
            bytes: PDF file content

        Implements Requirements: 4.4, 4.6
        """
        try:
            from io import BytesIO

            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

            # Create PDF buffer
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                topMargin=0.5 * inch,
                bottomMargin=0.5 * inch,
                leftMargin=0.75 * inch,
                rightMargin=0.75 * inch,
            )

            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                spaceAfter=12,
                alignment=1,  # Center alignment
            )
            heading_style = ParagraphStyle(
                "CustomHeading", parent=styles["Heading2"], fontSize=14, spaceAfter=10
            )
            normal_style = styles["Normal"]

            # Build PDF content
            story = []

            # Title
            story.append(Paragraph("Bank Reconciliation Report", title_style))
            story.append(Spacer(1, 12))

            # Bank Account Information
            bank_info_data = [
                ["Bank Account:", reconciliation.bank_account.account_name],
                ["Bank Name:", reconciliation.bank_account.bank_name],
                ["Account Number:", reconciliation.bank_account.masked_account_number],
                ["Reconciliation Date:", reconciliation.reconciliation_date.strftime("%B %d, %Y")],
                [
                    "Status:",
                    (
                        "Balanced"
                        if reconciliation.is_balanced
                        else f"Variance: ${reconciliation.variance:,.2f}"
                    ),
                ],
            ]

            if reconciliation.completed_by:
                bank_info_data.append(
                    [
                        "Completed By:",
                        reconciliation.completed_by.get_full_name()
                        or reconciliation.completed_by.username,
                    ]
                )
                bank_info_data.append(
                    [
                        "Completed Date:",
                        reconciliation.completed_date.strftime("%B %d, %Y %I:%M %p"),
                    ]
                )

            bank_info_table = Table(bank_info_data, colWidths=[2 * inch, 4.5 * inch])
            bank_info_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                        ("ALIGN", (1, 0), (1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            story.append(bank_info_table)
            story.append(Spacer(1, 20))

            # Balance Summary
            story.append(Paragraph("Balance Summary", heading_style))

            balance_data = [
                ["Description", "Amount"],
                [
                    "Statement Beginning Balance",
                    f"${reconciliation.statement_beginning_balance:,.2f}",
                ],
                ["Add: Deposits", f"${reconciliation.total_deposits:,.2f}"],
                ["Less: Withdrawals", f"${reconciliation.total_withdrawals:,.2f}"],
                ["Add/Less: Adjustments", f"${reconciliation.total_adjustments:,.2f}"],
                ["Statement Ending Balance", f"${reconciliation.statement_ending_balance:,.2f}"],
                ["", ""],
                ["Book Beginning Balance", f"${reconciliation.book_beginning_balance:,.2f}"],
                ["Book Ending Balance", f"${reconciliation.book_ending_balance:,.2f}"],
                ["", ""],
                [
                    "Variance",
                    f"${reconciliation.variance:,.2f}" if reconciliation.variance != 0 else "$0.00",
                ],
            ]

            balance_table = Table(balance_data, colWidths=[4.5 * inch, 2 * inch])
            balance_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 11),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                        ("TOPPADDING", (0, 0), (-1, 0), 10),
                        ("FONTNAME", (0, 1), (0, 5), "Helvetica"),
                        ("FONTNAME", (0, 7), (0, 8), "Helvetica"),
                        ("FONTNAME", (0, 10), (0, 10), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 10), (-1, 10), 12),
                        ("BACKGROUND", (0, 10), (-1, 10), colors.lightgrey),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            story.append(balance_table)
            story.append(Spacer(1, 20))

            # Deposits Section
            if deposits:
                story.append(Paragraph(f"Deposits ({len(deposits)} transactions)", heading_style))

                deposit_data = [["Date", "Description", "Reference", "Amount"]]
                for deposit in deposits:
                    deposit_data.append(
                        [
                            deposit.transaction_date.strftime("%m/%d/%Y"),
                            deposit.description[:40],
                            deposit.reference_number or "",
                            f"${deposit.amount:,.2f}",
                        ]
                    )

                # Add total row
                deposit_data.append(
                    ["", "", "Total Deposits:", f"${reconciliation.total_deposits:,.2f}"]
                )

                deposit_table = Table(
                    deposit_data, colWidths=[1 * inch, 3 * inch, 1.5 * inch, 1 * inch]
                )
                deposit_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                            ("TOPPADDING", (0, 0), (-1, 0), 8),
                            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                            ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )
                story.append(deposit_table)
                story.append(Spacer(1, 20))

            # Withdrawals Section
            if withdrawals:
                story.append(
                    Paragraph(f"Withdrawals ({len(withdrawals)} transactions)", heading_style)
                )

                withdrawal_data = [["Date", "Description", "Reference", "Amount"]]
                for withdrawal in withdrawals:
                    withdrawal_data.append(
                        [
                            withdrawal.transaction_date.strftime("%m/%d/%Y"),
                            withdrawal.description[:40],
                            withdrawal.reference_number or "",
                            f"${withdrawal.amount:,.2f}",
                        ]
                    )

                # Add total row
                withdrawal_data.append(
                    ["", "", "Total Withdrawals:", f"${reconciliation.total_withdrawals:,.2f}"]
                )

                withdrawal_table = Table(
                    withdrawal_data, colWidths=[1 * inch, 3 * inch, 1.5 * inch, 1 * inch]
                )
                withdrawal_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                            ("TOPPADDING", (0, 0), (-1, 0), 8),
                            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                            ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )
                story.append(withdrawal_table)
                story.append(Spacer(1, 20))

            # Notes section if present
            if reconciliation.notes:
                story.append(Paragraph("Notes", heading_style))
                story.append(Paragraph(reconciliation.notes, normal_style))

            # Build PDF
            doc.build(story)

            # Get PDF data
            pdf_data = buffer.getvalue()
            buffer.close()

            return pdf_data

        except Exception as e:
            logger.error(f"Failed to generate reconciliation PDF: {str(e)}")
            raise


class BankStatementImportService:
    """
    Service class for handling bank statement imports.

    Supports parsing CSV, OFX, and QFX file formats and automatically
    matching imported transactions with existing journal entries.

    Requirements: 4.3, 6.4
    """

    @staticmethod
    def import_statement(statement_import, user):
        """
        Import bank statement from file.

        Args:
            statement_import: BankStatementImport instance
            user: User performing the import

        Returns:
            dict: Import results with statistics
        """
        from apps.core.audit_models import AuditLog

        from .bank_models import BankTransaction

        logger.info(
            f"Starting bank statement import {statement_import.id} for {statement_import.bank_account.account_name}"
        )

        try:
            # Mark as processing
            statement_import.start_processing()

            # Parse file based on format
            if statement_import.file_format == "CSV":
                transactions = BankStatementImportService._parse_csv(statement_import)
            elif statement_import.file_format in ["OFX", "QFX", "QBO"]:
                transactions = BankStatementImportService._parse_ofx(statement_import)
            else:
                raise ValueError(f"Unsupported file format: {statement_import.file_format}")

            # Import transactions
            imported_count = 0
            matched_count = 0
            duplicate_count = 0
            error_count = 0

            with transaction.atomic():
                for txn_data in transactions:
                    try:
                        # Check for duplicates
                        existing = BankTransaction.objects.filter(
                            tenant=statement_import.tenant,
                            bank_account=statement_import.bank_account,
                            transaction_date=txn_data["date"],
                            amount=txn_data["amount"],
                            description=txn_data["description"],
                        ).first()

                        if existing:
                            duplicate_count += 1
                            continue

                        # Create bank transaction
                        bank_txn = BankTransaction.objects.create(
                            tenant=statement_import.tenant,
                            bank_account=statement_import.bank_account,
                            transaction_date=txn_data["date"],
                            description=txn_data["description"],
                            amount=txn_data["amount"],
                            transaction_type=txn_data["type"],
                            reference_number=txn_data.get("reference", ""),
                            statement_import=statement_import,
                            created_by=user,
                        )

                        imported_count += 1

                        # Try to auto-match with journal entries
                        if BankStatementImportService._auto_match_transaction(bank_txn):
                            matched_count += 1

                    except Exception as e:
                        logger.error(f"Error importing transaction: {str(e)}")
                        error_count += 1
                        continue

                # Update import statistics
                statement_import.transactions_imported = imported_count
                statement_import.transactions_matched = matched_count
                statement_import.transactions_duplicates = duplicate_count
                statement_import.transactions_errors = error_count
                statement_import.complete_processing(success=True)

                # Log audit trail
                AuditLog.objects.create(
                    tenant=statement_import.tenant,
                    user=user,
                    category="ACCOUNTING",
                    action="BANK_STATEMENT_IMPORTED",
                    severity="INFO",
                    description=f"Imported bank statement for {statement_import.bank_account.account_name}",
                    after_value=f"imported={imported_count}, matched={matched_count}, duplicates={duplicate_count}, errors={error_count}",
                )

                logger.info(
                    f"Bank statement import {statement_import.id} completed: "
                    f"{imported_count} imported, {matched_count} matched, "
                    f"{duplicate_count} duplicates, {error_count} errors"
                )

                return {
                    "success": True,
                    "imported": imported_count,
                    "matched": matched_count,
                    "duplicates": duplicate_count,
                    "errors": error_count,
                }

        except Exception as e:
            logger.error(f"Failed to import bank statement: {str(e)}")
            statement_import.complete_processing(success=False, error_message=str(e))

            return {
                "success": False,
                "error": str(e),
                "imported": 0,
                "matched": 0,
                "duplicates": 0,
                "errors": 0,
            }

    @staticmethod
    def _parse_csv(statement_import):  # noqa: C901
        """
        Parse CSV bank statement file.

        Expected CSV format:
        Date, Description, Amount, Type (or Debit/Credit columns)

        Args:
            statement_import: BankStatementImport instance

        Returns:
            list: List of transaction dictionaries
        """
        import csv
        import io
        from datetime import datetime
        from decimal import Decimal, InvalidOperation

        transactions = []

        try:
            # Read file content
            statement_import.file.seek(0)
            content = statement_import.file.read()

            # Try to decode as UTF-8, fallback to latin-1
            try:
                text_content = content.decode("utf-8")
            except UnicodeDecodeError:
                text_content = content.decode("latin-1")

            # Parse CSV
            csv_file = io.StringIO(text_content)
            reader = csv.DictReader(csv_file)

            for row in reader:
                # Try to parse date (support multiple formats)
                date_str = row.get("Date") or row.get("date") or row.get("Transaction Date")
                if not date_str:
                    continue

                # Parse date
                txn_date = None
                for date_format in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                    try:
                        txn_date = datetime.strptime(date_str.strip(), date_format).date()
                        break
                    except ValueError:
                        continue

                if not txn_date:
                    logger.warning(f"Could not parse date: {date_str}")
                    continue

                # Get description
                description = (
                    row.get("Description")
                    or row.get("description")
                    or row.get("Memo")
                    or row.get("memo")
                    or ""
                )

                # Get amount and type
                amount = None
                txn_type = None

                # Check if there's a single Amount column with type
                if "Amount" in row or "amount" in row:
                    amount_str = row.get("Amount") or row.get("amount")
                    try:
                        amount = abs(Decimal(amount_str.replace(",", "").replace("$", "")))
                        # Determine type from sign or Type column
                        if "Type" in row or "type" in row:
                            type_str = (row.get("Type") or row.get("type")).upper()
                            txn_type = (
                                "DEBIT"
                                if "DEBIT" in type_str or "WITHDRAWAL" in type_str
                                else "CREDIT"
                            )
                        else:
                            # Use sign of amount
                            txn_type = (
                                "DEBIT"
                                if float(amount_str.replace(",", "").replace("$", "")) < 0
                                else "CREDIT"
                            )
                    except (ValueError, InvalidOperation):
                        logger.warning(f"Could not parse amount: {amount_str}")
                        continue

                # Check for separate Debit/Credit columns
                elif "Debit" in row or "debit" in row or "Credit" in row or "credit" in row:
                    debit_str = row.get("Debit") or row.get("debit") or "0"
                    credit_str = row.get("Credit") or row.get("credit") or "0"

                    try:
                        debit = Decimal(debit_str.replace(",", "").replace("$", "") or "0")
                        credit = Decimal(credit_str.replace(",", "").replace("$", "") or "0")

                        if debit > 0:
                            amount = debit
                            txn_type = "DEBIT"
                        elif credit > 0:
                            amount = credit
                            txn_type = "CREDIT"
                        else:
                            continue
                    except (ValueError, InvalidOperation):
                        logger.warning(f"Could not parse debit/credit: {debit_str}/{credit_str}")
                        continue

                if not amount or not txn_type:
                    continue

                # Get reference number
                reference = (
                    row.get("Reference")
                    or row.get("reference")
                    or row.get("Check Number")
                    or row.get("check_number")
                    or ""
                )

                transactions.append(
                    {
                        "date": txn_date,
                        "description": description.strip(),
                        "amount": amount,
                        "type": txn_type,
                        "reference": reference.strip(),
                    }
                )

        except Exception as e:
            logger.error(f"Error parsing CSV file: {str(e)}")
            raise

        return transactions

    @staticmethod
    def _parse_ofx(statement_import):  # noqa: C901
        """
        Parse OFX/QFX bank statement file.

        Args:
            statement_import: BankStatementImport instance

        Returns:
            list: List of transaction dictionaries
        """
        from decimal import Decimal

        transactions = []

        try:
            # Try to import ofxparse library
            try:
                from ofxparse import OfxParser
            except ImportError:
                raise ImportError(
                    "ofxparse library is required for OFX/QFX file parsing. "
                    "Please install it: pip install ofxparse"
                )

            # Parse OFX file
            statement_import.file.seek(0)
            ofx = OfxParser.parse(statement_import.file)

            # Get account
            if not ofx.accounts:
                raise ValueError("No accounts found in OFX file")

            account = ofx.accounts[0]

            # Parse transactions
            for txn in account.statement.transactions:
                # Determine transaction type
                txn_type = "CREDIT" if txn.amount > 0 else "DEBIT"
                amount = abs(Decimal(str(txn.amount)))

                transactions.append(
                    {
                        "date": txn.date.date() if hasattr(txn.date, "date") else txn.date,
                        "description": txn.memo or txn.payee or "",
                        "amount": amount,
                        "type": txn_type,
                        "reference": txn.checknum or txn.id or "",
                    }
                )

            # Update statement dates and balance if available
            if account.statement.start_date:
                statement_import.statement_start_date = (
                    account.statement.start_date.date()
                    if hasattr(account.statement.start_date, "date")
                    else account.statement.start_date
                )

            if account.statement.end_date:
                statement_import.statement_end_date = (
                    account.statement.end_date.date()
                    if hasattr(account.statement.end_date, "date")
                    else account.statement.end_date
                )

            if account.statement.balance:
                statement_import.statement_balance = Decimal(str(account.statement.balance))

            statement_import.save(
                update_fields=[
                    "statement_start_date",
                    "statement_end_date",
                    "statement_balance",
                    "updated_at",
                ]
            )

        except ImportError as e:
            logger.error(f"OFX parsing library not available: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Error parsing OFX file: {str(e)}")
            raise

        return transactions

    @staticmethod
    def _auto_match_transaction(bank_transaction):
        """
        Attempt to automatically match a bank transaction with a journal entry.

        Matching criteria:
        1. Date within +/- 3 days
        2. Amount matches exactly
        3. Description contains keywords

        Args:
            bank_transaction: BankTransaction instance

        Returns:
            bool: True if matched, False otherwise
        """
        from datetime import timedelta

        from django_ledger.models import JournalEntryModel, TransactionModel

        try:
            # Get tenant's entity
            jewelry_entity = JewelryEntity.objects.get(tenant=bank_transaction.tenant)
            entity = jewelry_entity.ledger_entity

            # Search for matching journal entries
            start_date = bank_transaction.transaction_date - timedelta(days=3)
            end_date = bank_transaction.transaction_date + timedelta(days=3)

            # Get journal entries in date range
            journal_entries = JournalEntryModel.objects.filter(
                ledger__entity=entity,
                timestamp__date__gte=start_date,
                timestamp__date__lte=end_date,
                posted=True,
            )

            # Try to find matching entry
            for je in journal_entries:
                # Get transactions for this journal entry
                je_transactions = TransactionModel.objects.filter(journal_entry=je)

                # Calculate total amount
                total_amount = sum(abs(txn.amount) for txn in je_transactions)

                # Check if amount matches
                if abs(total_amount - bank_transaction.amount) < Decimal("0.01"):
                    # Check if description has any matching keywords
                    description_lower = bank_transaction.description.lower()
                    je_description_lower = je.description.lower()

                    # Extract keywords (words longer than 3 characters)
                    bank_keywords = set(word for word in description_lower.split() if len(word) > 3)
                    je_keywords = set(
                        word for word in je_description_lower.split() if len(word) > 3
                    )

                    # Check for keyword overlap
                    if bank_keywords & je_keywords:
                        # Match found!
                        bank_transaction.match_journal_entry(je)
                        logger.info(
                            f"Auto-matched bank transaction {bank_transaction.id} "
                            f"with journal entry {je.id}"
                        )
                        return True

            return False

        except Exception as e:
            logger.error(f"Error auto-matching transaction: {str(e)}")
            return False
