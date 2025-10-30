"""
Reporting services for the jewelry shop SaaS platform.

Implements Requirement 15: Advanced Reporting and Analytics
- Report execution engine
- Parameter validation and processing
- Data export services
- Email delivery services
"""

import csv
import json
import logging
import os
import tempfile
from datetime import timedelta
from typing import Any, Dict, List, Tuple

from django.conf import settings
from django.core.mail import EmailMessage
from django.db import connection
from django.template.loader import render_to_string
from django.utils import timezone

import openpyxl
from import_export import resources
from import_export.formats.base_formats import CSV, JSON, XLSX
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.core.models import Tenant
from apps.reporting.models import Report, ReportCategory, ReportExecution

logger = logging.getLogger(__name__)

# Optional WeasyPrint import
try:
    import weasyprint

    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    logger.warning("WeasyPrint not available. PDF export will use ReportLab only.")


class ReportParameterProcessor:
    """
    Process and validate report parameters.
    """

    def __init__(self, report: Report):
        self.report = report
        self.parameter_definitions = {
            param.name: param for param in report.parameter_definitions.all()
        }

    def validate_parameters(self, parameters: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """
        Validate report parameters against their definitions.

        Args:
            parameters: Dictionary of parameter values

        Returns:
            Tuple of (is_valid, error_messages)
        """
        errors = []

        # Check required parameters
        for param_name, param_def in self.parameter_definitions.items():
            if param_def.is_required and param_name not in parameters:
                errors.append(f"Required parameter '{param_def.label}' is missing")
                continue

            if param_name in parameters:
                value = parameters[param_name]
                if not param_def.validate_value(value):
                    errors.append(f"Invalid value for parameter '{param_def.label}'")

        return len(errors) == 0, errors

    def process_parameters(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process parameters and apply defaults.

        Args:
            parameters: Raw parameter values

        Returns:
            Processed parameters with defaults applied
        """
        processed = parameters.copy()

        # Apply default values for missing parameters
        for param_name, param_def in self.parameter_definitions.items():
            if param_name not in processed and param_def.default_value is not None:
                processed[param_name] = param_def.default_value

        # Process date ranges
        for param_name, param_def in self.parameter_definitions.items():
            if param_def.parameter_type == "DATERANGE" and param_name in processed:
                date_range = processed[param_name]
                if isinstance(date_range, dict):
                    # Convert to proper date objects
                    if "start_date" in date_range:
                        processed[f"{param_name}_start"] = date_range["start_date"]
                    if "end_date" in date_range:
                        processed[f"{param_name}_end"] = date_range["end_date"]

        return processed


class ReportQueryEngine:
    """
    Execute report queries and return data.
    """

    def __init__(self, tenant: Tenant):
        self.tenant = tenant

    def execute_query(self, report: Report, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Execute a report query with parameters.

        Args:
            report: Report instance
            parameters: Processed parameters

        Returns:
            List of result rows as dictionaries
        """
        query_config = report.query_config

        if report.report_type == "PREDEFINED":
            return self._execute_predefined_report(report, parameters)
        elif report.report_type == "CUSTOM":
            return self._execute_custom_query(query_config, parameters)
        else:
            raise ValueError(f"Unsupported report type: {report.report_type}")

    def _execute_predefined_report(
        self, report: Report, parameters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Execute a predefined report."""
        report_name = report.query_config.get("report_name")

        # Map report names to methods to reduce complexity
        report_methods = {
            # Sales reports
            "sales_summary": self._get_sales_summary,
            "sales_by_product": self._get_sales_by_product,
            "sales_by_employee": self._get_sales_by_employee,
            "sales_by_branch": self._get_sales_by_branch,
            # Inventory reports
            "inventory_valuation": self._get_inventory_valuation,
            "inventory_turnover": self._get_inventory_turnover,
            "dead_stock": self._get_dead_stock,
            # Financial reports
            "financial_summary": self._get_financial_summary,
            "revenue_trends": self._get_revenue_trends,
            "expense_breakdown": self._get_expense_breakdown,
            # Customer reports
            "top_customers": self._get_top_customers,
            "customer_acquisition": self._get_customer_acquisition,
            "loyalty_analytics": self._get_loyalty_analytics,
            # Legacy support
            "customer_analysis": self._get_customer_analysis,
        }

        if report_name not in report_methods:
            raise ValueError(f"Unknown predefined report: {report_name}")

        return report_methods[report_name](parameters)

    def _execute_custom_query(
        self, query_config: Dict[str, Any], parameters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Execute a custom SQL query."""
        sql_query = query_config.get("sql")
        if not sql_query:
            raise ValueError("Custom report must have SQL query")

        # Set tenant context for RLS
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )

            # Execute the main query
            cursor.execute(sql_query, parameters)
            columns = [col[0] for col in cursor.description]
            results = []

            for row in cursor.fetchall():
                results.append(dict(zip(columns, row)))

            return results

    # Sales Reports
    def _get_sales_summary(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get daily sales summary report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=30))
        end_date = parameters.get("date_range_end", timezone.now())
        branch_id = parameters.get("branch_id")

        sql = """
        SELECT
            DATE(s.created_at) as sale_date,
            b.name as branch_name,
            COUNT(s.id) as total_sales,
            SUM(s.total) as total_amount,
            AVG(s.total) as average_sale,
            SUM(s.tax) as total_tax,
            SUM(s.discount) as total_discount
        FROM sales s
        JOIN branches b ON s.branch_id = b.id
        WHERE s.created_at >= %s AND s.created_at <= %s
        """

        params = [start_date, end_date]

        if branch_id:
            sql += " AND s.branch_id = %s"
            params.append(branch_id)

        sql += " GROUP BY DATE(s.created_at), b.name ORDER BY sale_date DESC"

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_sales_by_product(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get sales by product report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=30))
        end_date = parameters.get("date_range_end", timezone.now())
        branch_id = parameters.get("branch_id")
        category_id = parameters.get("category_id")

        sql = """
        SELECT
            i.sku,
            i.name as product_name,
            pc.name as category_name,
            SUM(si.quantity) as total_quantity_sold,
            SUM(si.subtotal) as total_revenue,
            AVG(si.unit_price) as average_price,
            COUNT(DISTINCT s.id) as number_of_sales,
            i.karat,
            i.weight_grams
        FROM sale_items si
        JOIN sales s ON si.sale_id = s.id
        JOIN inventory_items i ON si.inventory_item_id = i.id
        JOIN inventory_categories pc ON i.category_id = pc.id
        WHERE s.created_at >= %s AND s.created_at <= %s
        """

        params = [start_date, end_date]

        if branch_id:
            sql += " AND s.branch_id = %s"
            params.append(branch_id)

        if category_id:
            sql += " AND i.category_id = %s"
            params.append(category_id)

        sql += """
        GROUP BY i.id, i.sku, i.name, pc.name, i.karat, i.weight_grams
        ORDER BY total_revenue DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_sales_by_employee(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get sales by employee report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=30))
        end_date = parameters.get("date_range_end", timezone.now())
        branch_id = parameters.get("branch_id")

        sql = """
        SELECT
            u.first_name,
            u.last_name,
            u.email,
            b.name as branch_name,
            COUNT(s.id) as total_sales,
            SUM(s.total) as total_revenue,
            AVG(s.total) as average_sale_value,
            SUM(s.tax) as total_tax_collected,
            MIN(s.created_at) as first_sale_date,
            MAX(s.created_at) as last_sale_date
        FROM sales s
        JOIN users u ON s.employee_id = u.id
        JOIN branches b ON s.branch_id = b.id
        WHERE s.created_at >= %s AND s.created_at <= %s
        """

        params = [start_date, end_date]

        if branch_id:
            sql += " AND s.branch_id = %s"
            params.append(branch_id)

        sql += """
        GROUP BY u.id, u.first_name, u.last_name, u.email, b.name
        ORDER BY total_revenue DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_sales_by_branch(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get sales by branch report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=30))
        end_date = parameters.get("date_range_end", timezone.now())

        sql = """
        SELECT
            b.name as branch_name,
            b.address as branch_address,
            COUNT(s.id) as total_sales,
            SUM(s.total) as total_revenue,
            AVG(s.total) as average_sale_value,
            SUM(s.tax) as total_tax_collected,
            COUNT(DISTINCT s.employee_id) as active_employees,
            COUNT(DISTINCT s.customer_id) as unique_customers,
            MIN(s.created_at) as first_sale_date,
            MAX(s.created_at) as last_sale_date
        FROM sales s
        JOIN branches b ON s.branch_id = b.id
        WHERE s.created_at >= %s AND s.created_at <= %s
        GROUP BY b.id, b.name, b.address
        ORDER BY total_revenue DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, [start_date, end_date])
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Inventory Reports
    def _get_inventory_valuation(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get inventory valuation report data."""
        branch_id = parameters.get("branch_id")
        category_id = parameters.get("category_id")

        sql = """
        SELECT
            i.sku,
            i.name,
            pc.name as category_name,
            b.name as branch_name,
            i.quantity,
            i.cost_price,
            i.selling_price,
            (i.quantity * i.cost_price) as total_cost_value,
            (i.quantity * i.selling_price) as total_selling_value,
            ((i.selling_price - i.cost_price) / i.cost_price * 100) as markup_percentage,
            i.karat,
            i.weight_grams,
            i.created_at as date_added
        FROM inventory_items i
        JOIN inventory_categories pc ON i.category_id = pc.id
        JOIN branches b ON i.branch_id = b.id
        WHERE i.is_active = true
        """

        params = []

        if branch_id:
            sql += " AND i.branch_id = %s"
            params.append(branch_id)

        if category_id:
            sql += " AND i.category_id = %s"
            params.append(category_id)

        sql += " ORDER BY total_selling_value DESC"

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_inventory_turnover(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get inventory turnover report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=90))
        end_date = parameters.get("date_range_end", timezone.now())
        branch_id = parameters.get("branch_id")

        sql = """
        SELECT
            i.sku,
            i.name as product_name,
            pc.name as category_name,
            b.name as branch_name,
            i.quantity as current_stock,
            COALESCE(SUM(si.quantity), 0) as quantity_sold,
            i.cost_price,
            (i.quantity * i.cost_price) as current_value,
            CASE
                WHEN i.quantity > 0 THEN COALESCE(SUM(si.quantity), 0) / i.quantity
                ELSE 0
            END as turnover_ratio,
            CASE
                WHEN COALESCE(SUM(si.quantity), 0) > 0 THEN
                    (EXTRACT(DAYS FROM (%s - %s)) * i.quantity) / COALESCE(SUM(si.quantity), 1)
                ELSE NULL
            END as days_to_sell_current_stock
        FROM inventory_items i
        JOIN inventory_categories pc ON i.category_id = pc.id
        JOIN branches b ON i.branch_id = b.id
        LEFT JOIN sale_items si ON i.id = si.inventory_item_id
        LEFT JOIN sales s ON si.sale_id = s.id AND s.created_at >= %s AND s.created_at <= %s
        WHERE i.is_active = true
        """

        params = [end_date, start_date, start_date, end_date]

        if branch_id:
            sql += " AND i.branch_id = %s"
            params.append(branch_id)

        sql += """
        GROUP BY i.id, i.sku, i.name, pc.name, b.name, i.quantity, i.cost_price
        ORDER BY turnover_ratio DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_dead_stock(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get dead stock analysis report data."""
        days_threshold = parameters.get("days_threshold", 90)
        branch_id = parameters.get("branch_id")
        cutoff_date = timezone.now() - timedelta(days=days_threshold)

        sql = """
        SELECT
            i.sku,
            i.name as product_name,
            pc.name as category_name,
            b.name as branch_name,
            i.quantity as current_stock,
            i.cost_price,
            (i.quantity * i.cost_price) as tied_up_value,
            i.created_at as date_added,
            EXTRACT(DAYS FROM (NOW() - i.created_at)) as days_in_inventory,
            COALESCE(MAX(s.created_at), i.created_at) as last_sale_date,
            EXTRACT(DAYS FROM (NOW() - COALESCE(MAX(s.created_at), i.created_at))) as days_since_last_sale,
            COUNT(si.id) as total_sales_count
        FROM inventory_items i
        JOIN inventory_categories pc ON i.category_id = pc.id
        JOIN branches b ON i.branch_id = b.id
        LEFT JOIN sale_items si ON i.id = si.inventory_item_id
        LEFT JOIN sales s ON si.sale_id = s.id
        WHERE i.is_active = true
        AND i.quantity > 0
        AND (
            NOT EXISTS (
                SELECT 1 FROM sale_items si2
                JOIN sales s2 ON si2.sale_id = s2.id
                WHERE si2.inventory_item_id = i.id
                AND s2.created_at >= %s
            )
            OR i.created_at < %s
        )
        """

        params = [cutoff_date, cutoff_date]

        if branch_id:
            sql += " AND i.branch_id = %s"
            params.append(branch_id)

        sql += """
        GROUP BY i.id, i.sku, i.name, pc.name, b.name, i.quantity, i.cost_price, i.created_at
        ORDER BY tied_up_value DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Financial Reports
    def _get_financial_summary(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get financial summary report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=30))
        end_date = parameters.get("date_range_end", timezone.now())

        sql = """
        SELECT
            'Sales Revenue' as category,
            SUM(total - tax) as amount,
            'REVENUE' as type,
            COUNT(*) as transaction_count
        FROM sales
        WHERE created_at >= %s AND created_at <= %s
        UNION ALL
        SELECT
            'Tax Collected' as category,
            SUM(tax) as amount,
            'REVENUE' as type,
            COUNT(*) as transaction_count
        FROM sales
        WHERE created_at >= %s AND created_at <= %s
        UNION ALL
        SELECT
            'Inventory Purchases' as category,
            SUM(total_amount) as amount,
            'EXPENSE' as type,
            COUNT(*) as transaction_count
        FROM procurement_purchaseorders
        WHERE status = 'COMPLETED' AND created_at >= %s AND created_at <= %s
        ORDER BY type, amount DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, [start_date, end_date, start_date, end_date, start_date, end_date])
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_revenue_trends(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get revenue trends report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=90))
        end_date = parameters.get("date_range_end", timezone.now())
        group_by = parameters.get("group_by", "day")  # day, week, month

        if group_by == "week":
            date_trunc = "DATE_TRUNC('week', s.created_at)"
        elif group_by == "month":
            date_trunc = "DATE_TRUNC('month', s.created_at)"
        else:
            date_trunc = "DATE(s.created_at)"

        sql = f"""
        SELECT
            {date_trunc} as period,
            SUM(s.total) as total_revenue,
            SUM(s.total - s.tax) as net_revenue,
            SUM(s.tax) as tax_amount,
            COUNT(s.id) as transaction_count,
            AVG(s.total) as average_transaction_value,
            COUNT(DISTINCT s.customer_id) as unique_customers
        FROM sales s
        WHERE s.created_at >= %s AND s.created_at <= %s
        GROUP BY {date_trunc}
        ORDER BY period
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, [start_date, end_date])
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_expense_breakdown(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get expense breakdown report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=30))
        end_date = parameters.get("date_range_end", timezone.now())

        sql = """
        SELECT
            'Inventory Purchases' as expense_category,
            SUM(po.total_amount) as total_amount,
            COUNT(po.id) as transaction_count,
            AVG(po.total_amount) as average_amount,
            'Purchase Orders' as source_type
        FROM procurement_purchaseorders po
        WHERE po.status = 'COMPLETED'
        AND po.created_at >= %s AND po.created_at <= %s

        UNION ALL

        SELECT
            'Repair Services' as expense_category,
            SUM(ro.cost_estimate) as total_amount,
            COUNT(ro.id) as transaction_count,
            AVG(ro.cost_estimate) as average_amount,
            'Repair Orders' as source_type
        FROM repair_repairorders ro
        WHERE ro.status = 'COMPLETED'
        AND ro.created_at >= %s AND ro.created_at <= %s

        ORDER BY total_amount DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, [start_date, end_date, start_date, end_date])
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Customer Reports
    def _get_top_customers(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get top customers report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=90))
        end_date = parameters.get("date_range_end", timezone.now())
        limit = parameters.get("limit", 50)

        sql = """
        SELECT
            c.customer_number,
            c.first_name,
            c.last_name,
            c.email,
            c.phone,
            lt.name as loyalty_tier,
            c.loyalty_points,
            c.store_credit,
            c.total_purchases as lifetime_value,
            COUNT(s.id) as period_orders,
            SUM(s.total) as period_spending,
            AVG(s.total) as average_order_value,
            MAX(s.created_at) as last_purchase_date,
            MIN(s.created_at) as first_purchase_date,
            EXTRACT(DAYS FROM (MAX(s.created_at) - MIN(s.created_at))) as customer_lifespan_days
        FROM crm_customers c
        LEFT JOIN crm_loyalty_tiers lt ON c.loyalty_tier_id = lt.id
        LEFT JOIN sales s ON c.id = s.customer_id
            AND s.created_at >= %s AND s.created_at <= %s
        GROUP BY c.id, c.customer_number, c.first_name, c.last_name,
                 c.email, c.phone, lt.name, c.loyalty_points,
                 c.store_credit, c.total_purchases
        HAVING COUNT(s.id) > 0
        ORDER BY period_spending DESC
        LIMIT %s
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, [start_date, end_date, limit])
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_customer_acquisition(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get customer acquisition report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=90))
        end_date = parameters.get("date_range_end", timezone.now())
        group_by = parameters.get("group_by", "month")  # day, week, month

        if group_by == "week":
            date_trunc = "DATE_TRUNC('week', c.created_at)"
        elif group_by == "day":
            date_trunc = "DATE(c.created_at)"
        else:
            date_trunc = "DATE_TRUNC('month', c.created_at)"

        sql = f"""
        SELECT
            {date_trunc} as period,
            COUNT(c.id) as new_customers,
            COUNT(first_purchase.customer_id) as customers_with_first_purchase,
            COALESCE(SUM(first_purchase.first_purchase_amount), 0) as first_purchase_revenue,
            COALESCE(AVG(first_purchase.first_purchase_amount), 0) as avg_first_purchase_value,
            COUNT(CASE WHEN lt.name != 'BRONZE' THEN 1 END) as premium_signups
        FROM crm_customers c
        LEFT JOIN crm_loyalty_tiers lt ON c.loyalty_tier_id = lt.id
        LEFT JOIN (
            SELECT
                s.customer_id,
                MIN(s.total) as first_purchase_amount,
                MIN(s.created_at) as first_purchase_date
            FROM sales s
            GROUP BY s.customer_id
        ) first_purchase ON c.id = first_purchase.customer_id
            AND first_purchase.first_purchase_date >= %s
            AND first_purchase.first_purchase_date <= %s
        WHERE c.created_at >= %s AND c.created_at <= %s
        GROUP BY {date_trunc}
        ORDER BY period
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, [start_date, end_date, start_date, end_date])
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_loyalty_analytics(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get loyalty program analytics report data."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=90))
        end_date = parameters.get("date_range_end", timezone.now())

        sql = """
        SELECT
            lt.name as loyalty_tier,
            COUNT(c.id) as customer_count,
            AVG(c.loyalty_points) as avg_points_balance,
            SUM(c.loyalty_points) as total_points_outstanding,
            AVG(c.total_purchases) as avg_lifetime_value,
            COUNT(period_sales.customer_id) as active_customers_in_period,
            COALESCE(AVG(period_sales.period_spending), 0) as avg_period_spending,
            COALESCE(SUM(period_sales.period_spending), 0) as total_period_spending,
            COUNT(CASE WHEN c.store_credit > 0 THEN 1 END) as customers_with_store_credit,
            COALESCE(AVG(CASE WHEN c.store_credit > 0 THEN c.store_credit END), 0) as avg_store_credit
        FROM crm_customers c
        LEFT JOIN crm_loyalty_tiers lt ON c.loyalty_tier_id = lt.id
        LEFT JOIN (
            SELECT
                s.customer_id,
                SUM(s.total) as period_spending,
                COUNT(s.id) as period_orders
            FROM sales s
            WHERE s.created_at >= %s AND s.created_at <= %s
            GROUP BY s.customer_id
        ) period_sales ON c.id = period_sales.customer_id
        GROUP BY lt.name
        ORDER BY
            CASE lt.name
                WHEN 'PLATINUM' THEN 1
                WHEN 'GOLD' THEN 2
                WHEN 'SILVER' THEN 3
                WHEN 'BRONZE' THEN 4
                ELSE 5
            END
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, [start_date, end_date])
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def _get_customer_analysis(self, parameters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Get customer analysis report data (legacy method)."""
        start_date = parameters.get("date_range_start", timezone.now() - timedelta(days=90))
        end_date = parameters.get("date_range_end", timezone.now())

        sql = """
        SELECT
            c.customer_number,
            c.first_name,
            c.last_name,
            c.email,
            c.phone,
            lt.name as loyalty_tier,
            c.loyalty_points,
            c.store_credit,
            c.total_purchases,
            COUNT(s.id) as total_orders,
            COALESCE(SUM(s.total), 0) as period_spending,
            MAX(s.created_at) as last_purchase_date
        FROM crm_customers c
        LEFT JOIN crm_loyalty_tiers lt ON c.loyalty_tier_id = lt.id
        LEFT JOIN sales s ON c.id = s.customer_id
            AND s.created_at >= %s AND s.created_at <= %s
        GROUP BY c.id, c.customer_number, c.first_name, c.last_name,
                 c.email, c.phone, lt.name, c.loyalty_points,
                 c.store_credit, c.total_purchases
        ORDER BY period_spending DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('app.current_tenant', %s, false)", [str(self.tenant.id)]
            )
            cursor.execute(sql, [start_date, end_date])
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]


class ReportDataResource(resources.Resource):
    """
    Django-import-export resource for report data.
    """

    def __init__(self, data: List[Dict[str, Any]], **kwargs):
        super().__init__(**kwargs)
        self.data = data

        # Dynamically create fields based on data keys
        if data:
            for key in data[0].keys():
                self.fields[key] = resources.Field(column_name=key)

    def get_queryset(self):
        # Return empty queryset since we're working with raw data
        return []

    def export(self, queryset=None, *args, **kwargs):
        """Export the data as a Dataset."""
        from tablib import Dataset

        if not self.data:
            return Dataset()

        # Create dataset with headers
        headers = list(self.data[0].keys())
        dataset = Dataset(headers=headers)

        # Add data rows
        for row in self.data:
            dataset.append([row.get(header, "") for header in headers])

        return dataset


class ReportExportService:
    """
    Export report data to various formats using multiple libraries.

    Implements Requirement 15: Advanced Reporting and Analytics
    - PDF export using ReportLab and WeasyPrint
    - Excel export using openpyxl and django-import-export
    - CSV export using django-import-export
    """

    def __init__(self, tenant: Tenant):
        self.tenant = tenant

    def export_to_csv(self, data: List[Dict[str, Any]], filename: str) -> str:
        """
        Export data to CSV format using django-import-export.

        Args:
            data: Report data
            filename: Output filename

        Returns:
            Path to the generated file
        """
        if not data:
            raise ValueError("No data to export")

        filepath = os.path.join(tempfile.gettempdir(), filename)

        # Use django-import-export for CSV export
        resource = ReportDataResource(data)
        dataset = resource.export()
        csv_format = CSV()
        export_data = csv_format.export_data(dataset)

        with open(filepath, "w", encoding="utf-8") as csvfile:
            csvfile.write(export_data)

        return filepath

    def export_to_csv_basic(self, data: List[Dict[str, Any]], filename: str) -> str:
        """
        Export data to CSV format using basic CSV writer (fallback method).

        Args:
            data: Report data
            filename: Output filename

        Returns:
            Path to the generated file
        """
        if not data:
            raise ValueError("No data to export")

        filepath = os.path.join(tempfile.gettempdir(), filename)

        with open(filepath, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = data[0].keys()
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

        return filepath

    def export_to_excel(
        self, data: List[Dict[str, Any]], filename: str, report_name: str = ""
    ) -> str:
        """
        Export data to Excel format using django-import-export with enhanced formatting.

        Args:
            data: Report data
            filename: Output filename
            report_name: Report name for the title

        Returns:
            Path to the generated file
        """
        if not data:
            raise ValueError("No data to export")

        filepath = os.path.join(tempfile.gettempdir(), filename)

        # Use django-import-export for Excel export
        resource = ReportDataResource(data)
        dataset = resource.export()
        xlsx_format = XLSX()
        export_data = xlsx_format.export_data(dataset)

        # Save the basic export first
        with open(filepath, "wb") as excelfile:
            excelfile.write(export_data)

        # Now enhance with formatting using openpyxl
        if report_name:
            self._enhance_excel_formatting(filepath, report_name)

        return filepath

    def export_to_excel_advanced(
        self, data: List[Dict[str, Any]], filename: str, report_name: str = ""
    ) -> str:
        """
        Export data to Excel format with advanced formatting using openpyxl.

        Args:
            data: Report data
            filename: Output filename
            report_name: Report name for the title

        Returns:
            Path to the generated file
        """
        if not data:
            raise ValueError("No data to export")

        filepath = os.path.join(tempfile.gettempdir(), filename)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Report Data"

        # Add title and headers
        start_row = self._add_excel_title(worksheet, report_name)
        headers = self._add_excel_headers(worksheet, data, start_row)

        # Add data and format
        self._add_excel_data(worksheet, data, headers, start_row)
        self._adjust_excel_columns(worksheet)

        workbook.save(filepath)
        return filepath

    def _add_excel_title(self, worksheet, report_name: str) -> int:
        """Add title section to Excel worksheet."""
        if report_name:
            worksheet["A1"] = report_name
            worksheet["A1"].font = Font(size=16, bold=True)
            worksheet["A2"] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
            worksheet["A3"] = f"Tenant: {self.tenant.company_name if self.tenant else 'Unknown'}"
            return 5
        return 1

    def _add_excel_headers(self, worksheet, data: List[Dict], start_row: int) -> List[str]:
        """Add headers to Excel worksheet."""
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        return headers

    def _add_excel_data(self, worksheet, data: List[Dict], headers: List[str], start_row: int):
        """Add data rows to Excel worksheet."""
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                worksheet.cell(row=row_idx, column=col_idx, value=value)

    def _adjust_excel_columns(self, worksheet):
        """Auto-adjust column widths in Excel worksheet."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _enhance_excel_formatting(self, filepath: str, report_name: str) -> None:
        """
        Enhance Excel file with additional formatting.

        Args:
            filepath: Path to the Excel file
            report_name: Report name for the title
        """
        try:
            workbook = openpyxl.load_workbook(filepath)
            worksheet = workbook.active

            # Insert title rows at the top
            worksheet.insert_rows(1, 3)

            # Add title
            worksheet["A1"] = report_name
            worksheet["A1"].font = Font(size=16, bold=True)
            worksheet["A2"] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
            worksheet["A3"] = f"Tenant: {self.tenant.company_name if self.tenant else 'Unknown'}"

            # Format header row (now row 4)
            for cell in worksheet[4]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )

            workbook.save(filepath)
        except Exception as e:
            logger.warning(f"Failed to enhance Excel formatting: {e}")
            # Continue without enhanced formatting

    def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        report_name: str = "",
        use_weasyprint: bool = False,
    ) -> str:
        """
        Export data to PDF format using ReportLab or WeasyPrint.

        Args:
            data: Report data
            filename: Output filename
            report_name: Report name for the title
            use_weasyprint: Whether to use WeasyPrint instead of ReportLab

        Returns:
            Path to the generated file
        """
        if not data:
            raise ValueError("No data to export")

        filepath = os.path.join(tempfile.gettempdir(), filename)

        if use_weasyprint and WEASYPRINT_AVAILABLE:
            return self._export_to_pdf_weasyprint(data, filepath, report_name)
        else:
            if use_weasyprint and not WEASYPRINT_AVAILABLE:
                logger.warning("WeasyPrint requested but not available. Falling back to ReportLab.")
            return self._export_to_pdf_reportlab(data, filepath, report_name)

    def _export_to_pdf_reportlab(
        self, data: List[Dict[str, Any]], filepath: str, report_name: str = ""
    ) -> str:
        """
        Export data to PDF format using ReportLab.
        """
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Add title and metadata
        if report_name:
            title = Paragraph(report_name, styles["Title"])
            story.append(title)
            story.append(Spacer(1, 12))

            subtitle = Paragraph(
                f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]
            )
            story.append(subtitle)

            tenant_info = Paragraph(
                f"Tenant: {self.tenant.company_name if self.tenant else 'Unknown'}",
                styles["Normal"],
            )
            story.append(tenant_info)
            story.append(Spacer(1, 12))

        # Prepare table data
        headers = list(data[0].keys())
        table_data = [headers]

        for row in data:
            table_data.append([str(row.get(header, "")) for header in headers])

        # Create table with improved styling
        table = Table(table_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )

        story.append(table)

        # Add footer with row count
        story.append(Spacer(1, 12))
        footer = Paragraph(f"Total rows: {len(data)}", styles["Normal"])
        story.append(footer)

        doc.build(story)
        return filepath

    def _export_to_pdf_weasyprint(
        self, data: List[Dict[str, Any]], filepath: str, report_name: str = ""
    ) -> str:
        """
        Export data to PDF format using WeasyPrint.
        """
        if not WEASYPRINT_AVAILABLE:
            raise ImportError("WeasyPrint is not available. Please install system dependencies.")

        # Create HTML content
        html_content = self._generate_html_for_pdf(data, report_name)

        # Generate PDF using WeasyPrint
        weasyprint.HTML(string=html_content).write_pdf(filepath)

        return filepath

    def _generate_html_for_pdf(self, data: List[Dict[str, Any]], report_name: str = "") -> str:
        """
        Generate HTML content for PDF export.
        """
        headers = list(data[0].keys())

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>{report_name or 'Report'}</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 20px;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 30px;
                }}
                .title {{
                    font-size: 24px;
                    font-weight: bold;
                    margin-bottom: 10px;
                }}
                .subtitle {{
                    font-size: 14px;
                    color: #666;
                    margin-bottom: 5px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }}
                th {{
                    background-color: #f2f2f2;
                    font-weight: bold;
                }}
                tr:nth-child(even) {{
                    background-color: #f9f9f9;
                }}
                .footer {{
                    margin-top: 20px;
                    text-align: center;
                    font-size: 12px;
                    color: #666;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="title">{report_name or 'Report'}</div>
                <div class="subtitle">Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
                <div class="subtitle">Tenant: {self.tenant.company_name if self.tenant else 'Unknown'}</div>
            </div>

            <table>
                <thead>
                    <tr>
                        {''.join(f'<th>{header}</th>' for header in headers)}
                    </tr>
                </thead>
                <tbody>
        """

        for row in data:
            html += "<tr>"
            for header in headers:
                value = str(row.get(header, ""))
                html += f"<td>{value}</td>"
            html += "</tr>"

        html += f"""
                </tbody>
            </table>

            <div class="footer">
                Total rows: {len(data)}
            </div>
        </body>
        </html>
        """

        return html

    def export_to_json(
        self, data: List[Dict[str, Any]], filename: str, use_import_export: bool = False
    ) -> str:
        """
        Export data to JSON format.

        Args:
            data: Report data
            filename: Output filename
            use_import_export: Whether to use django-import-export JSON format

        Returns:
            Path to the generated file
        """
        if not data:
            raise ValueError("No data to export")

        filepath = os.path.join(tempfile.gettempdir(), filename)

        if use_import_export:
            # Use django-import-export for JSON export
            resource = ReportDataResource(data)
            dataset = resource.export()
            json_format = JSON()
            export_data = json_format.export_data(dataset)

            with open(filepath, "w", encoding="utf-8") as jsonfile:
                jsonfile.write(export_data)
        else:
            # Use custom JSON format with metadata (compatible with existing tests)
            export_data = {
                "generated_at": timezone.now().isoformat(),
                "tenant": self.tenant.company_name if self.tenant else "Unknown",
                "row_count": len(data),
                "data": data,
            }

            with open(filepath, "w", encoding="utf-8") as jsonfile:
                json.dump(export_data, jsonfile, indent=2, default=str)

        return filepath

    def export_data(
        self, data: List[Dict[str, Any]], format_type: str, filename: str, report_name: str = ""
    ) -> str:
        """
        Export data to the specified format.

        Args:
            data: Report data
            format_type: Export format (CSV, EXCEL, PDF, JSON)
            filename: Output filename
            report_name: Report name for the title

        Returns:
            Path to the generated file
        """
        if not data:
            raise ValueError("No data to export")

        format_type = format_type.upper()

        try:
            if format_type == "CSV":
                return self.export_to_csv(data, filename)
            elif format_type == "EXCEL":
                return self.export_to_excel(data, filename, report_name)
            elif format_type == "PDF":
                return self.export_to_pdf(data, filename, report_name)
            elif format_type == "JSON":
                return self.export_to_json(data, filename)
            else:
                raise ValueError(f"Unsupported export format: {format_type}")

        except Exception as e:
            logger.error(f"Failed to export data to {format_type}: {e}")
            raise

    def get_supported_formats(self) -> List[str]:
        """
        Get list of supported export formats.

        Returns:
            List of supported format names
        """
        return ["CSV", "EXCEL", "PDF", "JSON"]

    def get_format_mime_type(self, format_type: str) -> str:
        """
        Get MIME type for the specified format.

        Args:
            format_type: Export format

        Returns:
            MIME type string
        """
        mime_types = {
            "CSV": "text/csv",
            "EXCEL": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "PDF": "application/pdf",
            "JSON": "application/json",
        }

        return mime_types.get(format_type.upper(), "application/octet-stream")

    def get_format_extension(self, format_type: str) -> str:
        """
        Get file extension for the specified format.

        Args:
            format_type: Export format

        Returns:
            File extension string
        """
        extensions = {
            "CSV": ".csv",
            "EXCEL": ".xlsx",
            "PDF": ".pdf",
            "JSON": ".json",
        }

        return extensions.get(format_type.upper(), ".txt")


class ReportEmailService:
    """
    Handle email delivery of reports.
    """

    def __init__(self, tenant: Tenant):
        self.tenant = tenant

    def send_report_email(
        self,
        execution: ReportExecution,
        file_path: str,
        recipients: List[str],
        subject: str = "",
        body: str = "",
    ) -> bool:
        """
        Send report via email.

        Args:
            execution: Report execution instance
            file_path: Path to the report file
            recipients: List of email addresses
            subject: Email subject (optional)
            body: Email body (optional)

        Returns:
            True if email sent successfully
        """
        try:
            # Generate default subject and body if not provided
            if not subject:
                subject = f"Report: {execution.report.name} - {timezone.now().strftime('%Y-%m-%d')}"

            if not body:
                body = self._generate_email_body(execution)

            # Create email
            email = EmailMessage(
                subject=subject, body=body, from_email=settings.DEFAULT_FROM_EMAIL, to=recipients
            )

            # Attach report file
            if file_path and os.path.exists(file_path):
                filename = os.path.basename(file_path)
                with open(file_path, "rb") as f:
                    email.attach(filename, f.read(), self._get_content_type(filename))

            # Send email
            email.send()

            # Update execution record
            execution.email_sent = True
            execution.email_recipients = recipients
            execution.save(update_fields=["email_sent", "email_recipients"])

            logger.info(f"Report email sent successfully to {len(recipients)} recipients")
            return True

        except Exception as e:
            logger.error(f"Failed to send report email: {e}")
            return False

    def _generate_email_body(self, execution: ReportExecution) -> str:
        """Generate default email body."""
        context = {
            "tenant_name": self.tenant.company_name if self.tenant else "Unknown",
            "report_name": execution.report.name,
            "execution_date": execution.started_at.strftime("%Y-%m-%d %H:%M:%S"),
            "row_count": execution.row_count,
            "duration": execution.duration_display,
        }

        return render_to_string("reporting/email/report_delivery.txt", context)

    def _get_content_type(self, filename: str) -> str:
        """Get MIME content type for file."""
        extension = filename.lower().split(".")[-1]

        content_types = {
            "pdf": "application/pdf",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "csv": "text/csv",
            "json": "application/json",
        }

        return content_types.get(extension, "application/octet-stream")


class PrebuiltReportService:
    """
    Service for managing pre-built report definitions and metadata.
    """

    @staticmethod
    def get_prebuilt_reports() -> List[Dict[str, Any]]:
        """Get list of all available pre-built reports."""
        return [
            # Sales Reports
            {
                "id": "sales_summary",
                "name": "Daily Sales Summary",
                "description": "Daily breakdown of sales by branch with totals and averages",
                "category": "SALES",
                "icon": "fas fa-chart-line",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 30,
                    },
                    {"name": "branch_id", "type": "BRANCH", "required": False},
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "sales_by_product",
                "name": "Sales by Product",
                "description": "Product performance analysis with quantities and revenue",
                "category": "SALES",
                "icon": "fas fa-box",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 30,
                    },
                    {"name": "branch_id", "type": "BRANCH", "required": False},
                    {"name": "category_id", "type": "SELECT", "required": False},
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "sales_by_employee",
                "name": "Sales by Employee",
                "description": "Employee performance analysis with sales metrics",
                "category": "SALES",
                "icon": "fas fa-user-tie",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 30,
                    },
                    {"name": "branch_id", "type": "BRANCH", "required": False},
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "sales_by_branch",
                "name": "Sales by Branch",
                "description": "Branch performance comparison with key metrics",
                "category": "SALES",
                "icon": "fas fa-store",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 30,
                    },
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            # Inventory Reports
            {
                "id": "inventory_valuation",
                "name": "Inventory Valuation",
                "description": "Current inventory value by product and category",
                "category": "INVENTORY",
                "icon": "fas fa-warehouse",
                "parameters": [
                    {"name": "branch_id", "type": "BRANCH", "required": False},
                    {"name": "category_id", "type": "SELECT", "required": False},
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "inventory_turnover",
                "name": "Inventory Turnover Analysis",
                "description": "Product turnover rates and stock movement analysis",
                "category": "INVENTORY",
                "icon": "fas fa-sync-alt",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 90,
                    },
                    {"name": "branch_id", "type": "BRANCH", "required": False},
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "dead_stock",
                "name": "Dead Stock Analysis",
                "description": "Identify slow-moving and dead stock items",
                "category": "INVENTORY",
                "icon": "fas fa-exclamation-triangle",
                "parameters": [
                    {"name": "days_threshold", "type": "NUMBER", "required": False, "default": 90},
                    {"name": "branch_id", "type": "BRANCH", "required": False},
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            # Financial Reports
            {
                "id": "financial_summary",
                "name": "Financial Summary",
                "description": "Revenue and expense summary with key financial metrics",
                "category": "FINANCIAL",
                "icon": "fas fa-dollar-sign",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 30,
                    },
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "revenue_trends",
                "name": "Revenue Trends",
                "description": "Revenue trends over time with period comparisons",
                "category": "FINANCIAL",
                "icon": "fas fa-chart-area",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 90,
                    },
                    {
                        "name": "group_by",
                        "type": "SELECT",
                        "required": False,
                        "default": "day",
                        "options": [
                            {"value": "day", "label": "Daily"},
                            {"value": "week", "label": "Weekly"},
                            {"value": "month", "label": "Monthly"},
                        ],
                    },
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "expense_breakdown",
                "name": "Expense Breakdown",
                "description": "Detailed breakdown of business expenses by category",
                "category": "FINANCIAL",
                "icon": "fas fa-receipt",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 30,
                    },
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            # Customer Reports
            {
                "id": "top_customers",
                "name": "Top Customers",
                "description": "Best customers by spending and purchase frequency",
                "category": "CUSTOMER",
                "icon": "fas fa-crown",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 90,
                    },
                    {"name": "limit", "type": "NUMBER", "required": False, "default": 50},
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "customer_acquisition",
                "name": "Customer Acquisition",
                "description": "New customer acquisition trends and first purchase analysis",
                "category": "CUSTOMER",
                "icon": "fas fa-user-plus",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 90,
                    },
                    {
                        "name": "group_by",
                        "type": "SELECT",
                        "required": False,
                        "default": "month",
                        "options": [
                            {"value": "day", "label": "Daily"},
                            {"value": "week", "label": "Weekly"},
                            {"value": "month", "label": "Monthly"},
                        ],
                    },
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
            {
                "id": "loyalty_analytics",
                "name": "Loyalty Program Analytics",
                "description": "Loyalty tier analysis and program performance metrics",
                "category": "CUSTOMER",
                "icon": "fas fa-medal",
                "parameters": [
                    {
                        "name": "date_range",
                        "type": "DATERANGE",
                        "required": True,
                        "default_days": 90,
                    },
                ],
                "output_formats": ["PDF", "EXCEL", "CSV"],
            },
        ]

    @staticmethod
    def get_prebuilt_report(report_id: str) -> Dict[str, Any]:
        """Get a specific pre-built report definition."""
        reports = PrebuiltReportService.get_prebuilt_reports()
        for report in reports:
            if report["id"] == report_id:
                return report
        raise ValueError(f"Pre-built report '{report_id}' not found")

    @staticmethod
    def get_reports_by_category(category: str) -> List[Dict[str, Any]]:
        """Get pre-built reports filtered by category."""
        reports = PrebuiltReportService.get_prebuilt_reports()
        return [report for report in reports if report["category"] == category]

    @staticmethod
    def create_prebuilt_report_instance(tenant: Tenant, report_id: str, user) -> Report:
        """Create a Report instance for a pre-built report."""
        report_def = PrebuiltReportService.get_prebuilt_report(report_id)

        # Get or create the category
        category, _ = ReportCategory.objects.get_or_create(
            category_type=report_def["category"],
            defaults={
                "name": f"{report_def['category'].title()} Reports",
                "description": f"Pre-built {report_def['category'].lower()} reports",
                "icon": report_def["icon"],
            },
        )

        # Create the report instance
        report = Report.objects.create(
            tenant=tenant,
            name=report_def["name"],
            description=report_def["description"],
            category=category,
            report_type="PREDEFINED",
            query_config={"report_name": report_id},
            parameters={"parameters": report_def["parameters"]},
            output_formats=report_def["output_formats"],
            created_by=user,
            is_public=True,  # Pre-built reports are public within tenant
        )

        return report


class ReportExecutionService:
    """
    Main service for executing reports.
    """

    def __init__(self, tenant: Tenant):
        self.tenant = tenant
        self.parameter_processor = None
        self.query_engine = ReportQueryEngine(tenant)
        self.export_service = ReportExportService(tenant)
        self.email_service = ReportEmailService(tenant)

    def execute_report(
        self,
        report: Report,
        parameters: Dict[str, Any],
        output_format: str,
        user,
        email_recipients: List[str] = None,
        trigger_type: str = "MANUAL",
    ) -> ReportExecution:
        """
        Execute a report with given parameters.

        Args:
            report: Report to execute
            parameters: Report parameters
            output_format: Output format (PDF, EXCEL, CSV, JSON)
            user: User executing the report
            email_recipients: Optional email recipients
            trigger_type: How the report was triggered

        Returns:
            ReportExecution instance
        """
        # Create execution record
        execution = ReportExecution.objects.create(
            report=report,
            trigger_type=trigger_type,
            parameters=parameters,
            output_format=output_format,
            executed_by=user,
            email_recipients=email_recipients or [],
        )

        try:
            # Update status to running
            execution.status = "RUNNING"
            execution.save(update_fields=["status"])

            # Validate parameters
            self.parameter_processor = ReportParameterProcessor(report)
            is_valid, errors = self.parameter_processor.validate_parameters(parameters)

            if not is_valid:
                raise ValueError(f"Parameter validation failed: {', '.join(errors)}")

            # Process parameters
            processed_params = self.parameter_processor.process_parameters(parameters)

            # Execute query
            data = self.query_engine.execute_query(report, processed_params)

            if not data:
                logger.warning(f"Report {report.name} returned no data")

            # Generate filename
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{report.name.replace(' ', '_')}_{timestamp}.{output_format.lower()}"

            # Export data
            file_path = self._export_data(data, output_format, filename, report.name)

            # Get file size
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0

            # Mark as completed
            execution.mark_completed(file_path, len(data))
            execution.result_file_size = file_size
            execution.save(update_fields=["result_file_size"])

            # Send email if recipients provided
            if email_recipients:
                self.email_service.send_report_email(execution, file_path, email_recipients)

            # Update report statistics
            report.increment_run_count()

            logger.info(f"Report {report.name} executed successfully: {len(data)} rows")

            return execution

        except Exception as e:
            logger.error(f"Report execution failed: {e}")
            execution.mark_failed(str(e))
            raise

    def _export_data(
        self, data: List[Dict[str, Any]], format_type: str, filename: str, report_name: str
    ) -> str:
        """Export data to the specified format."""
        if format_type == "CSV":
            return self.export_service.export_to_csv(data, filename)
        elif format_type == "EXCEL":
            return self.export_service.export_to_excel(data, filename, report_name)
        elif format_type == "PDF":
            return self.export_service.export_to_pdf(data, filename, report_name)
        elif format_type == "JSON":
            return self.export_service.export_to_json(data, filename)
        else:
            raise ValueError(f"Unsupported export format: {format_type}")
