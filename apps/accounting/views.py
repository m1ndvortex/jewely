"""
Views for the accounting module.
"""

import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import models, transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from django_ledger.models import AccountModel

from apps.core.decorators import tenant_access_required

from .models import AccountingConfiguration, JewelryEntity
from .services import AccountingService

logger = logging.getLogger(__name__)


@login_required
@tenant_access_required
def accounting_dashboard(request):
    """
    Main accounting dashboard landing page.
    """
    # Check if accounting is set up
    try:
        JewelryEntity.objects.get(tenant=request.user.tenant)
        is_setup = True
    except JewelryEntity.DoesNotExist:
        is_setup = False

    # Get quick stats
    context = {
        "is_setup": is_setup,
        "page_title": "Accounting Dashboard",
    }

    if is_setup:
        # Get current month financial summary
        end_date = date.today()
        start_date = date(end_date.year, end_date.month, 1)
        reports = AccountingService.get_financial_reports(request.user.tenant, start_date, end_date)

        context.update(
            {
                "reports": reports,
                "start_date": start_date,
                "end_date": end_date,
            }
        )

    return render(request, "accounting/dashboard.html", context)


@login_required
@tenant_access_required
def financial_reports(request):
    """
    Display financial reports dashboard.
    """
    # Get date range from request or default to current month
    end_date = date.today()
    start_date = date(end_date.year, end_date.month, 1)

    if request.GET.get("start_date"):
        start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
    if request.GET.get("end_date"):
        end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

    # Get financial reports
    reports = AccountingService.get_financial_reports(request.user.tenant, start_date, end_date)

    context = {
        "reports": reports,
        "start_date": start_date,
        "end_date": end_date,
        "page_title": "Financial Reports",
    }

    return render(request, "accounting/financial_reports.html", context)


@login_required
@tenant_access_required
def accounting_configuration(request):
    """
    Display and update accounting configuration.
    """
    config, created = AccountingConfiguration.objects.get_or_create(tenant=request.user.tenant)

    if request.method == "POST":
        # Update configuration
        config.use_automatic_journal_entries = (
            request.POST.get("use_automatic_journal_entries") == "on"
        )
        config.inventory_valuation_method = request.POST.get("inventory_valuation_method", "FIFO")
        config.default_cash_account = request.POST.get("default_cash_account", "1001")
        config.default_card_account = request.POST.get("default_card_account", "1002")
        config.default_inventory_account = request.POST.get("default_inventory_account", "1200")
        config.default_cogs_account = request.POST.get("default_cogs_account", "5001")
        config.default_sales_account = request.POST.get("default_sales_account", "4001")
        config.default_tax_account = request.POST.get("default_tax_account", "2003")

        config.save()
        messages.success(request, "Accounting configuration updated successfully.")

    context = {"config": config, "page_title": "Accounting Configuration"}

    return render(request, "accounting/configuration.html", context)


@login_required
@tenant_access_required
def chart_of_accounts(request):
    """
    Display chart of accounts.
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get chart of accounts
        coa = entity.chartofaccountmodel_set.first()
        accounts = coa.accountmodel_set.all().order_by("code") if coa else []

        context = {"accounts": accounts, "entity": entity, "page_title": "Chart of Accounts"}

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        context = {"accounts": [], "entity": None, "page_title": "Chart of Accounts"}

    return render(request, "accounting/chart_of_accounts.html", context)


@login_required
@tenant_access_required
@require_http_methods(["GET"])
def account_balance_api(request, account_code):
    """
    API endpoint to get account balance.
    """
    as_of_date = request.GET.get("as_of_date")
    if as_of_date:
        as_of_date = datetime.strptime(as_of_date, "%Y-%m-%d").date()
    else:
        as_of_date = date.today()

    balance = AccountingService.get_account_balance(request.user.tenant, account_code, as_of_date)

    return JsonResponse(
        {
            "account_code": account_code,
            "balance": str(balance),
            "as_of_date": as_of_date.isoformat(),
        }
    )


@csrf_exempt
@login_required
@tenant_access_required
@require_http_methods(["POST"])
def setup_accounting_api(request):
    """
    API endpoint to set up accounting for tenant.
    """
    try:
        with transaction.atomic():
            jewelry_entity = AccountingService.setup_tenant_accounting(
                request.user.tenant, request.user
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Accounting set up successfully",
                    "entity_id": str(jewelry_entity.ledger_entity.uuid),
                }
            )

    except Exception as e:
        import traceback

        logger.error(f"Setup accounting API error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse(
            {"success": False, "message": f"Failed to set up accounting: {str(e)}"}, status=500
        )


@login_required
@tenant_access_required
@require_http_methods(["GET"])
def export_financial_reports_pdf(request):
    """
    Export financial reports to PDF.
    """
    try:
        # Get date range from request or default to current month
        end_date = date.today()
        start_date = date(end_date.year, end_date.month, 1)

        if request.GET.get("start_date"):
            start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
        if request.GET.get("end_date"):
            end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

        # Generate PDF
        pdf_data = AccountingService.export_financial_reports_to_pdf(
            request.user.tenant, start_date, end_date
        )

        # Create response
        from django.http import HttpResponse

        response = HttpResponse(pdf_data, content_type="application/pdf")
        filename = f"{request.user.tenant.slug}_financial_reports_{start_date}_{end_date}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        logger.error(f"Failed to export PDF: {str(e)}")
        messages.error(request, f"Failed to export PDF: {str(e)}")
        return redirect("accounting:financial_reports")


@login_required
@tenant_access_required
@require_http_methods(["GET"])
def export_financial_reports_excel(request):
    """
    Export financial reports to Excel.
    """
    try:
        # Get date range from request or default to current month
        end_date = date.today()
        start_date = date(end_date.year, end_date.month, 1)

        if request.GET.get("start_date"):
            start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
        if request.GET.get("end_date"):
            end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

        # Generate Excel
        excel_data = AccountingService.export_financial_reports_to_excel(
            request.user.tenant, start_date, end_date
        )

        # Create response
        from django.http import HttpResponse

        response = HttpResponse(
            excel_data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"{request.user.tenant.slug}_financial_reports_{start_date}_{end_date}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        logger.error(f"Failed to export Excel: {str(e)}")
        messages.error(request, f"Failed to export Excel: {str(e)}")
        return redirect("accounting:financial_reports")


@login_required
@tenant_access_required
def journal_entries(request):
    """
    Display journal entries list (legacy view - redirects to journal_entry_list).
    """
    return redirect("accounting:journal_entry_list")


@login_required
@tenant_access_required
def general_ledger(request):
    """
    Display general ledger with all transactions.
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get date range and account filter
        end_date = date.today()
        start_date = date(end_date.year, end_date.month, 1)
        account_code = request.GET.get("account")

        if request.GET.get("start_date"):
            start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
        if request.GET.get("end_date"):
            end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

        # Get chart of accounts
        coa = entity.chartofaccountmodel_set.first()
        accounts = coa.accountmodel_set.all().order_by("code") if coa else []

        # Get transactions
        from django_ledger.models import TransactionModel

        transactions = TransactionModel.objects.filter(
            journal_entry__ledger__entity=entity,
            journal_entry__timestamp__date__gte=start_date,
            journal_entry__timestamp__date__lte=end_date,
        ).select_related("journal_entry", "account")

        if account_code:
            transactions = transactions.filter(account__code=account_code)

        transactions = transactions.order_by("-journal_entry__timestamp")

        context = {
            "transactions": transactions,
            "accounts": accounts,
            "selected_account": account_code,
            "start_date": start_date,
            "end_date": end_date,
            "page_title": "General Ledger",
        }

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        context = {"transactions": [], "accounts": [], "page_title": "General Ledger"}

    return render(request, "accounting/general_ledger.html", context)


@login_required
@tenant_access_required
def accounts_payable(request):
    """
    Display accounts payable dashboard.
    """
    from apps.procurement.models import PurchaseOrder

    # Get outstanding purchase orders
    outstanding_pos = PurchaseOrder.objects.filter(
        tenant=request.user.tenant, status__in=["APPROVED", "SENT", "PARTIALLY_RECEIVED"]
    ).select_related("supplier")

    # Calculate totals
    total_payable = sum(po.total for po in outstanding_pos)
    overdue_pos = [
        po
        for po in outstanding_pos
        if po.expected_delivery_date and po.expected_delivery_date < date.today()
    ]
    total_overdue = sum(po.total for po in overdue_pos)

    context = {
        "outstanding_pos": outstanding_pos,
        "total_payable": total_payable,
        "overdue_count": len(overdue_pos),
        "total_overdue": total_overdue,
        "page_title": "Accounts Payable",
    }

    return render(request, "accounting/accounts_payable.html", context)


@login_required
@tenant_access_required
def accounts_receivable(request):
    """
    Display accounts receivable dashboard.
    """
    from apps.sales.models import Sale

    # Get sales with outstanding balances (if you have credit sales)
    # For now, showing recent sales
    recent_sales = (
        Sale.objects.filter(tenant=request.user.tenant)
        .select_related("customer", "branch")
        .order_by("-created_at")[:50]
    )

    # Calculate totals
    total_sales = sum(sale.total for sale in recent_sales)

    context = {
        "recent_sales": recent_sales,
        "total_sales": total_sales,
        "page_title": "Accounts Receivable",
    }

    return render(request, "accounting/accounts_receivable.html", context)


@login_required
@tenant_access_required
def bank_reconciliation(request):
    """
    Display bank reconciliation interface.

    Shows bank accounts, unreconciled transactions, and reconciliation summary.
    Requirement: 4.1, 4.2, 4.3, 4.4
    """
    from .bank_models import BankAccount, BankReconciliation, BankTransaction

    # Get all bank accounts for this tenant
    bank_accounts = BankAccount.objects.filter(tenant=request.user.tenant, is_active=True).order_by(
        "-is_default", "account_name"
    )

    # Get selected account
    selected_account_id = request.GET.get("account")
    selected_account = None
    transactions = []
    reconciliation = None
    past_reconciliations = []

    if selected_account_id:
        try:
            selected_account = BankAccount.objects.get(
                id=selected_account_id, tenant=request.user.tenant
            )

            # Get ONLY unreconciled transactions for this account
            transactions = BankTransaction.objects.filter(
                bank_account=selected_account,
                tenant=request.user.tenant,
                is_reconciled=False,  # Only show unreconciled transactions
            ).order_by("-transaction_date", "-created_at")

            # Get current in-progress reconciliation if any
            reconciliation = BankReconciliation.objects.filter(
                bank_account=selected_account, tenant=request.user.tenant, status="IN_PROGRESS"
            ).first()

            # Get past completed reconciliations
            past_reconciliations = (
                BankReconciliation.objects.filter(
                    bank_account=selected_account, tenant=request.user.tenant, status="COMPLETED"
                )
                .select_related("completed_by")
                .order_by("-reconciliation_date")
            )

        except BankAccount.DoesNotExist:
            messages.error(request, "Bank account not found.")
            selected_account_id = None

    context = {
        "bank_accounts": bank_accounts,
        "selected_account": selected_account,
        "selected_account_id": selected_account_id,
        "transactions": transactions,
        "reconciliation": reconciliation,
        "past_reconciliations": past_reconciliations,
        "today": date.today(),
        "page_title": "Bank Reconciliation",
    }

    return render(request, "accounting/bank_reconciliation.html", context)


@login_required
@tenant_access_required
def add_account(request):
    """
    Create a new account in the chart of accounts.
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity
        coa = entity.chartofaccountmodel_set.first()

        if not coa:
            messages.error(request, "Chart of Accounts not found. Please set up accounting first.")
            return redirect("accounting:dashboard")

        if request.method == "POST":
            code = request.POST.get("code")
            name = request.POST.get("name")
            role = request.POST.get("role")
            balance_type = request.POST.get("balance_type")
            active = request.POST.get("active") == "1"

            # Check if account code already exists
            from django_ledger.models import AccountModel

            if AccountModel.objects.filter(coa_model=coa, code=code).exists():
                messages.error(request, f"Account with code {code} already exists.")
            else:
                # Create the account as a root node in the MPTT tree
                # AccountModel uses MPTT (Modified Preorder Tree Traversal) for
                # hierarchical structure
                AccountModel.add_root(
                    coa_model=coa,
                    code=code,
                    name=name,
                    role=role,
                    balance_type=balance_type,
                    active=active,
                )

                messages.success(request, f"Account {code} - {name} created successfully.")
                return redirect("accounting:chart_of_accounts")

        context = {"page_title": "Add Account"}
        return render(request, "accounting/account_form.html", context)

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")


@login_required
@tenant_access_required
def edit_account(request, account_code):
    """
    Edit an existing account in the chart of accounts.
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity
        coa = entity.chartofaccountmodel_set.first()

        if not coa:
            messages.error(request, "Chart of Accounts not found.")
            return redirect("accounting:dashboard")

        from django_ledger.models import AccountModel

        account = AccountModel.objects.get(coa_model=coa, code=account_code)

        if request.method == "POST":
            account.name = request.POST.get("name")
            account.role = request.POST.get("role")
            account.balance_type = request.POST.get("balance_type")
            account.active = request.POST.get("active") == "1"
            account.save()

            messages.success(
                request, f"Account {account.code} - {account.name} updated successfully."
            )
            return redirect("accounting:chart_of_accounts")

        context = {"account": account, "page_title": "Edit Account"}
        return render(request, "accounting/account_form.html", context)

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")
    except AccountModel.DoesNotExist:
        messages.error(request, f"Account {account_code} not found.")
        return redirect("accounting:chart_of_accounts")


# ============================================================================
# Manual Journal Entry Views (Task 1.2)
# ============================================================================


@login_required
@tenant_access_required
def journal_entry_list(request):
    """
    Display journal entries list with filtering by date, status, account.

    Implements filtering and pagination for manual journal entries.
    Requirement: 1.1, 1.2, 1.7, 1.8
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get date range from request or default to current month
        end_date = date.today()
        start_date = date(end_date.year, end_date.month, 1)

        if request.GET.get("start_date"):
            start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
        if request.GET.get("end_date"):
            end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

        # Get filter parameters
        status_filter = request.GET.get("status", "")
        account_filter = request.GET.get("account", "")

        # Get journal entries from django-ledger with tenant filtering
        from django_ledger.models import JournalEntryModel

        entries = (
            JournalEntryModel.objects.filter(
                ledger__entity=entity,
                timestamp__date__gte=start_date,
                timestamp__date__lte=end_date,
            )
            .select_related("ledger")
            .prefetch_related("transactionmodel_set__account")
        )

        # Apply status filter
        if status_filter:
            if status_filter == "posted":
                entries = entries.filter(posted=True)
            elif status_filter == "unposted":
                entries = entries.filter(posted=False)

        # Apply account filter
        if account_filter:
            entries = entries.filter(transactionmodel__account__code=account_filter).distinct()

        entries = entries.order_by("-timestamp")

        # Get chart of accounts for filter dropdown
        coa = entity.chartofaccountmodel_set.first()
        accounts = []
        if coa:
            from django_ledger.models import AccountModel

            accounts = AccountModel.objects.filter(coa_model=coa, active=True).order_by("code")

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description=f"Viewed journal entries list (filters: status={status_filter}, account={account_filter})",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        context = {
            "entries": entries,
            "start_date": start_date,
            "end_date": end_date,
            "status_filter": status_filter,
            "account_filter": account_filter,
            "accounts": accounts,
            "page_title": "Journal Entries",
        }

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        context = {
            "entries": [],
            "accounts": [],
            "page_title": "Journal Entries",
        }

    return render(request, "accounting/journal_entries/list.html", context)


@login_required
@tenant_access_required
def journal_entry_create(request):  # noqa: C901
    """
    Create a new manual journal entry with formset handling.

    Handles dynamic line items and validates that debits equal credits.
    Requirement: 1.1, 1.2, 1.3, 1.7, 1.8
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get chart of accounts
        coa = entity.chartofaccountmodel_set.first()
        if not coa:
            messages.error(request, "Chart of Accounts not found. Please set up accounting first.")
            return redirect("accounting:dashboard")

        # Get the ledger
        ledger = entity.ledgermodel_set.first()
        if not ledger:
            messages.error(request, "Ledger not found. Please set up accounting first.")
            return redirect("accounting:dashboard")

        from .forms import JournalEntryForm, JournalEntryLineInlineFormSet

        if request.method == "POST":
            form = JournalEntryForm(request.POST, tenant=request.user.tenant, user=request.user)

            # Create a temporary journal entry instance for the formset
            if form.is_valid():
                with transaction.atomic():
                    # Create journal entry (unposted)
                    journal_entry = form.save(commit=False)
                    journal_entry.ledger = ledger
                    journal_entry.posted = False
                    journal_entry.save()

                    # Handle formset
                    formset = JournalEntryLineInlineFormSet(
                        request.POST,
                        instance=journal_entry,
                        tenant=request.user.tenant,
                        coa=coa,
                    )

                    if formset.is_valid():
                        # Save the lines
                        lines = formset.save(commit=False)

                        for line in lines:
                            # Get debit/credit from cleaned_data
                            for form_instance in formset.forms:
                                if form_instance.instance == line:
                                    debit = form_instance.cleaned_data.get("debit") or 0
                                    credit = form_instance.cleaned_data.get("credit") or 0

                                    # Set transaction type and amount
                                    if debit > 0:
                                        line.tx_type = "debit"
                                        line.amount = debit
                                    else:
                                        line.tx_type = "credit"
                                        line.amount = credit

                                    line.journal_entry = journal_entry
                                    line.save()
                                    break

                        # Delete removed lines
                        for obj in formset.deleted_objects:
                            obj.delete()

                        # Audit logging
                        from apps.core.audit_models import AuditLog

                        AuditLog.objects.create(
                            tenant=request.user.tenant,
                            user=request.user,
                            category=AuditLog.CATEGORY_DATA,
                            action=AuditLog.ACTION_CREATE,
                            severity=AuditLog.SEVERITY_INFO,
                            description=f"Created journal entry: {journal_entry.description}",
                            ip_address=request.META.get("REMOTE_ADDR"),
                            user_agent=request.META.get("HTTP_USER_AGENT", ""),
                            request_method=request.method,
                            request_path=request.path,
                        )

                        messages.success(
                            request,
                            "Journal entry created successfully. Entry is unposted and can be edited.",
                        )
                        return redirect("accounting:journal_entry_detail", pk=journal_entry.uuid)
                    else:
                        # Formset validation failed
                        messages.error(
                            request, "Please correct the errors in the journal entry lines."
                        )
                        # Delete the journal entry since formset failed
                        journal_entry.delete()
            else:
                # Form validation failed
                formset = JournalEntryLineInlineFormSet(
                    request.POST,
                    tenant=request.user.tenant,
                    coa=coa,
                )
        else:
            form = JournalEntryForm(tenant=request.user.tenant, user=request.user)
            formset = JournalEntryLineInlineFormSet(tenant=request.user.tenant, coa=coa)

        context = {
            "form": form,
            "formset": formset,
            "page_title": "Create Journal Entry",
        }

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")

    return render(request, "accounting/journal_entries/form.html", context)


@login_required
@tenant_access_required
def journal_entry_detail(request, pk):
    """
    Display journal entry details (read-only).

    Shows all line items and audit trail information.
    Requirement: 1.1, 1.2, 1.6, 1.8
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get journal entry with tenant filtering
        from django_ledger.models import JournalEntryModel

        entry = (
            JournalEntryModel.objects.filter(ledger__entity=entity, uuid=pk)
            .select_related("ledger")
            .prefetch_related("transactionmodel_set__account")
            .first()
        )

        if not entry:
            messages.error(request, "Journal entry not found.")
            return redirect("accounting:journal_entry_list")

        # Calculate totals
        from decimal import Decimal

        total_debits = Decimal("0.00")
        total_credits = Decimal("0.00")

        for txn in entry.transactionmodel_set.all():
            if txn.tx_type == "debit":
                total_debits += txn.amount
            else:
                total_credits += txn.amount

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description=f"Viewed journal entry detail: {entry.description}",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        context = {
            "entry": entry,
            "total_debits": total_debits,
            "total_credits": total_credits,
            "is_balanced": total_debits == total_credits,
            "page_title": f"Journal Entry: {entry.description}",
        }

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")

    return render(request, "accounting/journal_entries/detail.html", context)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def journal_entry_post(request, pk):
    """
    Post a journal entry to the general ledger.

    Makes the entry immutable and updates account balances.
    Requirement: 1.4, 1.7, 1.8
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get journal entry with tenant filtering
        from django_ledger.models import JournalEntryModel

        entry = JournalEntryModel.objects.filter(ledger__entity=entity, uuid=pk).first()

        if not entry:
            messages.error(request, "Journal entry not found.")
            return redirect("accounting:journal_entry_list")

        # Check if already posted
        if entry.posted:
            messages.warning(request, "Journal entry is already posted.")
            return redirect("accounting:journal_entry_detail", pk=pk)

        # Validate that debits equal credits
        from decimal import Decimal

        total_debits = Decimal("0.00")
        total_credits = Decimal("0.00")

        for txn in entry.transactionmodel_set.all():
            if txn.tx_type == "debit":
                total_debits += txn.amount
            else:
                total_credits += txn.amount

        if total_debits != total_credits:
            messages.error(
                request,
                f"Cannot post journal entry: debits (${total_debits:,.2f}) do not equal credits (${total_credits:,.2f}).",
            )
            return redirect("accounting:journal_entry_detail", pk=pk)

        # Post the entry
        with transaction.atomic():
            entry.posted = True
            entry.save()

            # Audit logging
            from apps.core.audit_models import AuditLog

            AuditLog.objects.create(
                tenant=request.user.tenant,
                user=request.user,
                category=AuditLog.CATEGORY_DATA,
                action=AuditLog.ACTION_UPDATE,
                severity=AuditLog.SEVERITY_INFO,
                description=f"Posted journal entry: {entry.description}",
                old_values={"posted": False},
                new_values={"posted": True},
                ip_address=request.META.get("REMOTE_ADDR"),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
                request_method=request.method,
                request_path=request.path,
            )

            messages.success(
                request,
                "Journal entry posted successfully. Entry is now immutable and affects account balances.",
            )

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")

    return redirect("accounting:journal_entry_detail", pk=pk)


@login_required
@tenant_access_required
def journal_entry_reverse(request, pk):
    """
    Create a reversing journal entry.

    Creates a new entry with debits and credits swapped.
    Requirement: 1.5, 1.7, 1.8
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get original journal entry with tenant filtering
        from django_ledger.models import JournalEntryModel

        original_entry = JournalEntryModel.objects.filter(ledger__entity=entity, uuid=pk).first()

        if not original_entry:
            messages.error(request, "Journal entry not found.")
            return redirect("accounting:journal_entry_list")

        # Check if original entry is posted
        if not original_entry.posted:
            messages.error(request, "Cannot reverse an unposted journal entry.")
            return redirect("accounting:journal_entry_detail", pk=pk)

        if request.method == "POST":
            # Create reversing entry
            with transaction.atomic():
                # Get the ledger
                ledger = entity.ledgermodel_set.first()

                # Create new journal entry
                reversing_entry = JournalEntryModel.objects.create(
                    ledger=ledger,
                    description=f"REVERSAL: {original_entry.description}",
                    posted=False,  # Create as unposted so user can review
                )

                # Create reversed transactions
                from django_ledger.models import TransactionModel

                for original_txn in original_entry.transactionmodel_set.all():
                    # Swap debit/credit
                    reversed_tx_type = "credit" if original_txn.tx_type == "debit" else "debit"

                    TransactionModel.objects.create(
                        journal_entry=reversing_entry,
                        account=original_txn.account,
                        amount=original_txn.amount,
                        tx_type=reversed_tx_type,
                        description=f"Reversal of: {original_txn.description or ''}",
                    )

                # Audit logging
                from apps.core.audit_models import AuditLog

                AuditLog.objects.create(
                    tenant=request.user.tenant,
                    user=request.user,
                    category=AuditLog.CATEGORY_DATA,
                    action=AuditLog.ACTION_CREATE,
                    severity=AuditLog.SEVERITY_INFO,
                    description=f"Created reversing entry for: {original_entry.description}",
                    metadata={
                        "original_entry_id": str(original_entry.uuid),
                        "reversing_entry_id": str(reversing_entry.uuid),
                    },
                    ip_address=request.META.get("REMOTE_ADDR"),
                    user_agent=request.META.get("HTTP_USER_AGENT", ""),
                    request_method=request.method,
                    request_path=request.path,
                )

                messages.success(
                    request,
                    "Reversing entry created successfully. Please review and post the entry.",
                )
                return redirect("accounting:journal_entry_detail", pk=reversing_entry.uuid)

        context = {
            "original_entry": original_entry,
            "page_title": f"Reverse Journal Entry: {original_entry.description}",
        }

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")

    return render(request, "accounting/journal_entries/confirm_reverse.html", context)


@login_required
@tenant_access_required
def journal_entry_edit(request, pk):  # noqa: C901
    """
    Edit an unposted journal entry.

    Only allows editing of unposted entries. Posted entries are immutable.
    Requirement: 1.5, 1.7, 1.8
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get journal entry with tenant filtering
        from django_ledger.models import JournalEntryModel

        entry = JournalEntryModel.objects.filter(ledger__entity=entity, uuid=pk).first()

        if not entry:
            messages.error(request, "Journal entry not found.")
            return redirect("accounting:journal_entry_list")

        # Check if entry is posted
        if entry.posted:
            messages.error(
                request,
                "Cannot edit a posted journal entry. Posted entries are immutable. Please create a reversing entry instead.",
            )
            return redirect("accounting:journal_entry_detail", pk=pk)

        # Get chart of accounts
        coa = entity.chartofaccountmodel_set.first()
        if not coa:
            messages.error(request, "Chart of Accounts not found.")
            return redirect("accounting:dashboard")

        from .forms import JournalEntryForm, JournalEntryLineInlineFormSet

        if request.method == "POST":
            form = JournalEntryForm(
                request.POST, instance=entry, tenant=request.user.tenant, user=request.user
            )

            if form.is_valid():
                with transaction.atomic():
                    journal_entry = form.save(commit=False)
                    journal_entry.save()

                    # Handle formset
                    formset = JournalEntryLineInlineFormSet(
                        request.POST,
                        instance=journal_entry,
                        tenant=request.user.tenant,
                        coa=coa,
                    )

                    if formset.is_valid():
                        # Delete all existing lines
                        journal_entry.transactionmodel_set.all().delete()

                        # Save new lines
                        lines = formset.save(commit=False)

                        for line in lines:
                            # Get debit/credit from cleaned_data
                            for form_instance in formset.forms:
                                if form_instance.instance == line:
                                    debit = form_instance.cleaned_data.get("debit") or 0
                                    credit = form_instance.cleaned_data.get("credit") or 0

                                    # Set transaction type and amount
                                    if debit > 0:
                                        line.tx_type = "debit"
                                        line.amount = debit
                                    else:
                                        line.tx_type = "credit"
                                        line.amount = credit

                                    line.journal_entry = journal_entry
                                    line.save()
                                    break

                        # Audit logging
                        from apps.core.audit_models import AuditLog

                        AuditLog.objects.create(
                            tenant=request.user.tenant,
                            user=request.user,
                            category=AuditLog.CATEGORY_DATA,
                            action=AuditLog.ACTION_UPDATE,
                            severity=AuditLog.SEVERITY_INFO,
                            description=f"Updated journal entry: {journal_entry.description}",
                            ip_address=request.META.get("REMOTE_ADDR"),
                            user_agent=request.META.get("HTTP_USER_AGENT", ""),
                            request_method=request.method,
                            request_path=request.path,
                        )

                        messages.success(request, "Journal entry updated successfully.")
                        return redirect("accounting:journal_entry_detail", pk=journal_entry.uuid)
                    else:
                        messages.error(
                            request, "Please correct the errors in the journal entry lines."
                        )
            else:
                formset = JournalEntryLineInlineFormSet(
                    request.POST,
                    instance=entry,
                    tenant=request.user.tenant,
                    coa=coa,
                )
        else:
            form = JournalEntryForm(instance=entry, tenant=request.user.tenant, user=request.user)
            formset = JournalEntryLineInlineFormSet(
                instance=entry, tenant=request.user.tenant, coa=coa
            )

        context = {
            "form": form,
            "formset": formset,
            "entry": entry,
            "page_title": f"Edit Journal Entry: {entry.description}",
        }

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")

    return render(request, "accounting/journal_entries/form.html", context)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def journal_entry_delete(request, pk):
    """
    Delete an unposted journal entry.

    Only allows deletion of unposted entries. Posted entries cannot be deleted.
    Requirement: 1.6, 1.7, 1.8
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get journal entry with tenant filtering
        from django_ledger.models import JournalEntryModel

        entry = JournalEntryModel.objects.filter(ledger__entity=entity, uuid=pk).first()

        if not entry:
            messages.error(request, "Journal entry not found.")
            return redirect("accounting:journal_entry_list")

        # Check if entry is posted
        if entry.posted:
            messages.error(
                request,
                "Cannot delete a posted journal entry. Posted entries are immutable. Please create a reversing entry instead.",
            )
            return redirect("accounting:journal_entry_detail", pk=pk)

        # Store description for audit log
        description = entry.description

        # Delete the entry
        with transaction.atomic():
            entry.delete()

            # Audit logging
            from apps.core.audit_models import AuditLog

            AuditLog.objects.create(
                tenant=request.user.tenant,
                user=request.user,
                category=AuditLog.CATEGORY_DATA,
                action=AuditLog.ACTION_DELETE,
                severity=AuditLog.SEVERITY_INFO,
                description=f"Deleted unposted journal entry: {description}",
                ip_address=request.META.get("REMOTE_ADDR"),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
                request_method=request.method,
                request_path=request.path,
            )

            messages.success(request, f"Journal entry '{description}' deleted successfully.")

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")

    return redirect("accounting:journal_entry_list")


# ============================================================================
# Supplier Accounting Views (Task 2.3)
# ============================================================================


@login_required
@tenant_access_required
def supplier_accounting_detail(request, supplier_id):
    """
    Display supplier accounting details with bills, payments, and balance.

    Extends existing supplier detail with accounting information including
    total purchases, outstanding balance, and payment history.

    Requirements: 2.7, 14.1, 14.2, 14.3, 14.7, 14.8
    """
    from decimal import Decimal

    from apps.procurement.models import Supplier

    from .bill_models import Bill, BillPayment

    try:
        # Get supplier with tenant filtering
        supplier = (
            Supplier.objects.filter(tenant=request.user.tenant, id=supplier_id)
            .select_related("tenant")
            .first()
        )

        if not supplier:
            messages.error(request, "Supplier not found.")
            return redirect("procurement:supplier_list")

        # Get all bills for this supplier with tenant filtering
        bills = (
            Bill.objects.filter(tenant=request.user.tenant, supplier=supplier)
            .select_related("supplier", "created_by", "approved_by")
            .prefetch_related("lines", "payments")
            .order_by("-bill_date")
        )

        # Calculate totals
        total_purchases = Decimal("0.00")
        outstanding_balance = Decimal("0.00")
        total_paid = Decimal("0.00")

        for bill in bills:
            total_purchases += bill.total
            outstanding_balance += bill.amount_due
            total_paid += bill.amount_paid

        # Get payment history (all payments for this supplier's bills)
        payment_history = (
            BillPayment.objects.filter(tenant=request.user.tenant, bill__supplier=supplier)
            .select_related("bill", "created_by")
            .order_by("-payment_date")[:20]  # Last 20 payments
        )

        # Get aging breakdown
        aging_buckets = {
            "Current": Decimal("0.00"),
            "1-30 days": Decimal("0.00"),
            "31-60 days": Decimal("0.00"),
            "61-90 days": Decimal("0.00"),
            "90+ days": Decimal("0.00"),
        }

        for bill in bills:
            if bill.amount_due > 0:
                aging_buckets[bill.aging_bucket] += bill.amount_due

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description=f"Viewed supplier accounting detail: {supplier.name}",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        context = {
            "supplier": supplier,
            "bills": bills,
            "payment_history": payment_history,
            "total_purchases": total_purchases,
            "outstanding_balance": outstanding_balance,
            "total_paid": total_paid,
            "aging_buckets": aging_buckets,
            "page_title": f"Supplier Accounting: {supplier.name}",
        }

        return render(request, "accounting/suppliers/accounting_detail.html", context)

    except Exception as e:
        logger.error(f"Error in supplier_accounting_detail: {str(e)}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("procurement:supplier_list")


@login_required
@tenant_access_required
def supplier_list(request):
    """
    Display list of suppliers with accounting summary information.

    Provides quick access to supplier accounting details and statements
    from within the accounting module.

    Requirements: 14.1, 14.2, 14.7, 14.8
    """
    from apps.procurement.models import Supplier

    from .bill_models import Bill

    try:
        # Get all suppliers for the tenant
        suppliers = Supplier.objects.filter(tenant=request.user.tenant).select_related("tenant")

        # Get search and filter parameters
        search_query = request.GET.get("search", "")
        status_filter = request.GET.get("status", "")

        # Apply search filter
        if search_query:
            suppliers = suppliers.filter(
                models.Q(name__icontains=search_query)
                | models.Q(contact_person__icontains=search_query)
                | models.Q(email__icontains=search_query)
            )

        # Apply status filter
        if status_filter:
            suppliers = suppliers.filter(is_active=(status_filter == "active"))

        # Calculate accounting summary for each supplier
        supplier_data = []
        for supplier in suppliers:
            # Get total bills
            bills = Bill.objects.filter(tenant=request.user.tenant, supplier=supplier)
            total_purchases = bills.aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")

            # Get outstanding balance (unpaid bills)
            outstanding_bills = bills.exclude(status="paid")
            outstanding_balance = outstanding_bills.aggregate(total=models.Sum("total"))[
                "total"
            ] or Decimal("0.00")

            # Get paid amount
            paid_bills = bills.filter(status="paid")
            total_paid = paid_bills.aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")

            # Count bills
            bill_count = bills.count()

            supplier_data.append(
                {
                    "supplier": supplier,
                    "total_purchases": total_purchases,
                    "outstanding_balance": outstanding_balance,
                    "total_paid": total_paid,
                    "bill_count": bill_count,
                }
            )

        # Sort by outstanding balance (highest first)
        supplier_data.sort(key=lambda x: x["outstanding_balance"], reverse=True)

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description="Viewed supplier list from accounting module",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        context = {
            "supplier_data": supplier_data,
            "search_query": search_query,
            "status_filter": status_filter,
            "page_title": "Suppliers - Accounting",
        }

        return render(request, "accounting/suppliers/list.html", context)

    except Exception as e:
        logger.error(f"Error in supplier_list: {str(e)}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("accounting:dashboard")


@login_required
@tenant_access_required
def supplier_statement(request, supplier_id):
    """
    Generate supplier statement showing all transactions and current balance.

    Displays all bills and payments for a supplier within a date range,
    with opening and closing balances.

    Requirements: 2.8, 14.3, 14.7, 14.8
    """
    from decimal import Decimal

    from apps.procurement.models import Supplier

    from .bill_models import Bill, BillPayment

    try:
        # Get supplier with tenant filtering
        supplier = (
            Supplier.objects.filter(tenant=request.user.tenant, id=supplier_id)
            .select_related("tenant")
            .first()
        )

        if not supplier:
            messages.error(request, "Supplier not found.")
            return redirect("procurement:supplier_list")

        # Get date range from request or default to current month
        end_date = date.today()
        start_date = date(end_date.year, end_date.month, 1)

        if request.GET.get("start_date"):
            start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
        if request.GET.get("end_date"):
            end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

        # Get bills within date range with tenant filtering
        bills = (
            Bill.objects.filter(
                tenant=request.user.tenant,
                supplier=supplier,
                bill_date__range=[start_date, end_date],
            )
            .select_related("supplier", "created_by")
            .prefetch_related("lines", "payments")
            .order_by("bill_date")
        )

        # Get payments within date range with tenant filtering
        payments = (
            BillPayment.objects.filter(
                tenant=request.user.tenant,
                bill__supplier=supplier,
                payment_date__range=[start_date, end_date],
            )
            .select_related("bill", "created_by")
            .order_by("payment_date")
        )

        # Calculate opening balance (bills before start_date minus payments before start_date)
        opening_bills = Bill.objects.filter(
            tenant=request.user.tenant, supplier=supplier, bill_date__lt=start_date
        ).aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")

        opening_payments = BillPayment.objects.filter(
            tenant=request.user.tenant, bill__supplier=supplier, payment_date__lt=start_date
        ).aggregate(total=models.Sum("amount"))["total"] or Decimal("0.00")

        opening_balance = opening_bills - opening_payments

        # Calculate period totals
        period_bills_total = bills.aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")
        period_payments_total = payments.aggregate(total=models.Sum("amount"))["total"] or Decimal(
            "0.00"
        )

        # Calculate closing balance
        closing_balance = opening_balance + period_bills_total - period_payments_total

        # Combine bills and payments into a single transaction list
        transactions = []

        for bill in bills:
            transactions.append(
                {
                    "date": bill.bill_date,
                    "type": "Bill",
                    "reference": bill.bill_number,
                    "description": f"Bill #{bill.bill_number}",
                    "debit": bill.total,
                    "credit": Decimal("0.00"),
                    "balance": None,  # Will calculate running balance below
                }
            )

        for payment in payments:
            transactions.append(
                {
                    "date": payment.payment_date,
                    "type": "Payment",
                    "reference": payment.reference_number or f"Payment #{payment.id}",
                    "description": f"Payment for Bill #{payment.bill.bill_number}",
                    "debit": Decimal("0.00"),
                    "credit": payment.amount,
                    "balance": None,
                }
            )

        # Sort by date
        transactions.sort(key=lambda x: x["date"])

        # Calculate running balance
        running_balance = opening_balance
        for txn in transactions:
            running_balance += txn["debit"] - txn["credit"]
            txn["balance"] = running_balance

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description=f"Generated supplier statement: {supplier.name} ({start_date} to {end_date})",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        context = {
            "supplier": supplier,
            "start_date": start_date,
            "end_date": end_date,
            "opening_balance": opening_balance,
            "closing_balance": closing_balance,
            "period_bills_total": period_bills_total,
            "period_payments_total": period_payments_total,
            "transactions": transactions,
            "page_title": f"Supplier Statement: {supplier.name}",
        }

        return render(request, "accounting/suppliers/statement.html", context)

    except Exception as e:
        logger.error(f"Error in supplier_statement: {str(e)}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("procurement:supplier_list")


@login_required
@tenant_access_required
def supplier_statement_pdf(request, supplier_id):  # noqa: C901
    """
    Generate supplier statement as PDF export.

    Creates a PDF document showing all bills and payments for a supplier
    within a date range, with opening and closing balances.

    Requirements: 2.8, 14.3
    """
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from apps.procurement.models import Supplier

    from .bill_models import Bill, BillPayment

    try:
        # Get supplier with tenant filtering
        supplier = (
            Supplier.objects.filter(tenant=request.user.tenant, id=supplier_id)
            .select_related("tenant")
            .first()
        )

        if not supplier:
            messages.error(request, "Supplier not found.")
            return redirect("procurement:supplier_list")

        # Get date range from request or default to current month
        end_date = date.today()
        start_date = date(end_date.year, end_date.month, 1)

        if request.GET.get("start_date"):
            start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
        if request.GET.get("end_date"):
            end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

        # Get bills within date range with tenant filtering
        bills = (
            Bill.objects.filter(
                tenant=request.user.tenant,
                supplier=supplier,
                bill_date__range=[start_date, end_date],
            )
            .select_related("supplier", "created_by")
            .prefetch_related("lines", "payments")
            .order_by("bill_date")
        )

        # Get payments within date range with tenant filtering
        payments = (
            BillPayment.objects.filter(
                tenant=request.user.tenant,
                bill__supplier=supplier,
                payment_date__range=[start_date, end_date],
            )
            .select_related("bill", "created_by")
            .order_by("payment_date")
        )

        # Calculate opening balance
        opening_bills = Bill.objects.filter(
            tenant=request.user.tenant, supplier=supplier, bill_date__lt=start_date
        ).aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")

        opening_payments = BillPayment.objects.filter(
            tenant=request.user.tenant, bill__supplier=supplier, payment_date__lt=start_date
        ).aggregate(total=models.Sum("amount"))["total"] or Decimal("0.00")

        opening_balance = opening_bills - opening_payments

        # Calculate period totals
        period_bills_total = bills.aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")
        period_payments_total = payments.aggregate(total=models.Sum("amount"))["total"] or Decimal(
            "0.00"
        )

        # Calculate closing balance
        closing_balance = opening_balance + period_bills_total - period_payments_total

        # Combine bills and payments into a single transaction list
        transactions = []

        for bill in bills:
            transactions.append(
                {
                    "date": bill.bill_date,
                    "type": "Bill",
                    "reference": bill.bill_number,
                    "description": f"Bill #{bill.bill_number}",
                    "debit": bill.total,
                    "credit": Decimal("0.00"),
                    "balance": None,
                }
            )

        for payment in payments:
            transactions.append(
                {
                    "date": payment.payment_date,
                    "type": "Payment",
                    "reference": payment.reference_number or f"Payment #{payment.id}",
                    "description": f"Payment for Bill #{payment.bill.bill_number}",
                    "debit": Decimal("0.00"),
                    "credit": payment.amount,
                    "balance": None,
                }
            )

        # Sort by date
        transactions.sort(key=lambda x: x["date"])

        # Calculate running balance
        running_balance = opening_balance
        for txn in transactions:
            running_balance += txn["debit"] - txn["credit"]
            txn["balance"] = running_balance

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch
        )
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1f2937"),
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#374151"),
        )
        normal_style = styles["Normal"]

        # Title
        elements.append(Paragraph("Supplier Statement", title_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Supplier Information and Period
        supplier_info = f"""
        <b>{supplier.name}</b><br/>
        {supplier.contact_person or ''}<br/>
        {supplier.address or ''}<br/>
        {supplier.email or ''}<br/>
        {supplier.phone or ''}
        """

        period_info = f"""
        <b>{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}</b><br/>
        Generated on {date.today().strftime('%B %d, %Y')}
        """

        info_data = [
            ["Supplier Information", "Statement Period"],
            [
                Paragraph(supplier_info, normal_style),
                Paragraph(period_info, normal_style),
            ],
        ]

        info_table = Table(info_data, colWidths=[3.5 * inch, 3.5 * inch])
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("ALIGN", (0, 0), (-1, 0), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("TOPPADDING", (0, 0), (-1, 0), 12),
                    ("VALIGN", (0, 1), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        elements.append(info_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Balance Summary
        balance_data = [
            ["Opening Balance", "Period Bills", "Period Payments", "Closing Balance"],
            [
                f"${opening_balance:,.2f}",
                f"${period_bills_total:,.2f}",
                f"${period_payments_total:,.2f}",
                f"${closing_balance:,.2f}",
            ],
        ]

        balance_table = Table(balance_data, colWidths=[1.75 * inch] * 4)
        balance_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("TOPPADDING", (0, 0), (-1, -1), 12),
                    ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 1), (-1, 1), 12),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        elements.append(balance_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Transaction Details
        elements.append(Paragraph("Transaction Details", heading_style))
        elements.append(Spacer(1, 0.2 * inch))

        if transactions:
            # Transaction table header
            txn_data = [
                ["Date", "Type", "Reference", "Description", "Charges", "Payments", "Balance"]
            ]

            # Opening balance row
            txn_data.append(
                [
                    start_date.strftime("%Y-%m-%d"),
                    "",
                    "",
                    "Opening Balance",
                    "",
                    "",
                    f"${opening_balance:,.2f}",
                ]
            )

            # Transaction rows
            for txn in transactions:
                txn_data.append(
                    [
                        txn["date"].strftime("%Y-%m-%d"),
                        txn["type"],
                        txn["reference"],
                        txn["description"],
                        f"${txn['debit']:,.2f}" if txn["debit"] > 0 else "-",
                        f"${txn['credit']:,.2f}" if txn["credit"] > 0 else "-",
                        f"${txn['balance']:,.2f}",
                    ]
                )

            # Closing balance row
            txn_data.append(
                [
                    end_date.strftime("%Y-%m-%d"),
                    "",
                    "",
                    "Closing Balance",
                    f"${period_bills_total:,.2f}",
                    f"${period_payments_total:,.2f}",
                    f"${closing_balance:,.2f}",
                ]
            )

            txn_table = Table(
                txn_data,
                colWidths=[
                    0.8 * inch,
                    0.6 * inch,
                    0.9 * inch,
                    1.8 * inch,
                    0.9 * inch,
                    0.9 * inch,
                    0.9 * inch,
                ],
            )
            txn_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#dbeafe")),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f3f4f6")),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            elements.append(txn_table)
        else:
            elements.append(
                Paragraph("No transactions found for the selected period.", normal_style)
            )

        # Footer
        elements.append(Spacer(1, 0.5 * inch))
        footer_text = (
            "This is a computer-generated statement and does not require a signature.<br/>"
            "For questions about this statement, please contact your accounting department."
        )
        elements.append(Paragraph(footer_text, normal_style))

        # Build PDF
        doc.build(elements)

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description=f"Exported supplier statement PDF: {supplier.name} ({start_date} to {end_date})",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        # Return PDF response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="supplier_statement_{supplier.name.replace(" ", "_")}_'
            f'{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.pdf"'
        )
        return response

    except Exception as e:
        logger.error(f"Error in supplier_statement_pdf: {str(e)}")
        messages.error(request, f"An error occurred generating PDF: {str(e)}")
        return redirect("accounting:supplier_statement", supplier_id=supplier_id)


# ============================================================================
# Bill Management Views (Task 2.5)
# ============================================================================


@login_required
@tenant_access_required
def bill_list(request):
    """
    Display bills list with filtering by supplier, date range, status, and amount.

    Shows all bills for the tenant with aging information and status badges.
    Supports search and filtering.

    Requirements: 2.3, 2.7, 2.8
    """
    from decimal import Decimal

    from apps.procurement.models import Supplier

    from .bill_models import Bill

    try:
        # Get filter parameters
        supplier_filter = request.GET.get("supplier", "")
        status_filter = request.GET.get("status", "")
        start_date = request.GET.get("start_date", "")
        end_date = request.GET.get("end_date", "")
        min_amount = request.GET.get("min_amount", "")
        max_amount = request.GET.get("max_amount", "")
        search_query = request.GET.get("search", "")

        # Get all bills for this tenant
        bills = (
            Bill.objects.filter(tenant=request.user.tenant)
            .select_related("supplier", "created_by", "approved_by")
            .prefetch_related("lines", "payments")
        )

        # Apply filters
        if supplier_filter:
            bills = bills.filter(supplier_id=supplier_filter)

        if status_filter:
            bills = bills.filter(status=status_filter)

        if start_date:
            bills = bills.filter(bill_date__gte=datetime.strptime(start_date, "%Y-%m-%d").date())

        if end_date:
            bills = bills.filter(bill_date__lte=datetime.strptime(end_date, "%Y-%m-%d").date())

        if min_amount:
            bills = bills.filter(total__gte=Decimal(min_amount))

        if max_amount:
            bills = bills.filter(total__lte=Decimal(max_amount))

        if search_query:
            bills = bills.filter(
                models.Q(bill_number__icontains=search_query)
                | models.Q(supplier__name__icontains=search_query)
                | models.Q(notes__icontains=search_query)
            )

        bills = bills.order_by("-bill_date", "-created_at")

        # Calculate summary statistics
        total_bills = bills.count()
        total_amount = bills.aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")
        total_outstanding = bills.aggregate(
            outstanding=models.Sum(models.F("total") - models.F("amount_paid"))
        )["outstanding"] or Decimal("0.00")

        # Get unpaid bills count
        unpaid_bills_count = bills.filter(status__in=["APPROVED", "PARTIALLY_PAID"]).count()

        # Get overdue bills count
        overdue_bills_count = bills.filter(
            status__in=["APPROVED", "PARTIALLY_PAID"], due_date__lt=date.today()
        ).count()

        # Get suppliers for filter dropdown
        suppliers = Supplier.objects.filter(tenant=request.user.tenant, is_active=True).order_by(
            "name"
        )

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description=f"Viewed bills list (filters: supplier={supplier_filter}, status={status_filter})",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        context = {
            "bills": bills,
            "suppliers": suppliers,
            "supplier_filter": supplier_filter,
            "status_filter": status_filter,
            "start_date": start_date,
            "end_date": end_date,
            "min_amount": min_amount,
            "max_amount": max_amount,
            "search_query": search_query,
            "total_bills": total_bills,
            "total_amount": total_amount,
            "total_outstanding": total_outstanding,
            "unpaid_bills_count": unpaid_bills_count,
            "overdue_bills_count": overdue_bills_count,
            "status_choices": Bill.STATUS_CHOICES,
            "page_title": "Bills",
        }

        return render(request, "accounting/bills/list.html", context)

    except Exception as e:
        logger.error(f"Error in bill_list: {str(e)}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("accounting:dashboard")


@login_required
@tenant_access_required
def bill_create(request):  # noqa: C901
    """
    Create a new bill with automatic journal entry creation.

    Handles dynamic line items and creates journal entry debiting
    expense/asset accounts and crediting accounts payable.

    Requirements: 2.1, 2.2, 2.6, 2.7, 2.8
    """
    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get chart of accounts (optional - can create bills without COA)
        coa = entity.chartofaccountmodel_set.first()
        if not coa:
            messages.warning(
                request,
                "Chart of Accounts not set up. Bills can be created but account codes won't be available.",
            )

        from .forms import BillForm, BillLineInlineFormSet

        if request.method == "POST":
            form = BillForm(request.POST, tenant=request.user.tenant, user=request.user)

            if form.is_valid():
                with transaction.atomic():
                    # Create bill
                    bill = form.save(commit=False)
                    bill.tenant = request.user.tenant
                    bill.created_by = request.user
                    bill.status = "DRAFT"
                    bill.subtotal = Decimal("0.00")
                    bill.total = Decimal("0.00")
                    bill.amount_paid = Decimal("0.00")
                    bill.save()

                    # Handle formset
                    formset = BillLineInlineFormSet(
                        request.POST, instance=bill, tenant=request.user.tenant, coa=coa
                    )

                    if formset.is_valid():
                        # Save the lines
                        lines = formset.save(commit=False)

                        for line in lines:
                            line.bill = bill
                            line.save()

                        # Delete removed lines
                        for obj in formset.deleted_objects:
                            obj.delete()

                        # Calculate totals
                        bill.calculate_totals()

                        # Create journal entry only if COA is set up
                        if coa:
                            try:
                                journal_entry = _create_bill_journal_entry(
                                    bill, request.user.tenant, entity, coa
                                )
                                bill.journal_entry = journal_entry
                            except Exception as je_error:
                                logger.warning(f"Could not create journal entry: {str(je_error)}")
                                messages.warning(
                                    request,
                                    "Bill created but journal entry could not be created. Please check account setup.",
                                )

                        bill.status = "APPROVED"  # Auto-approve for now
                        bill.approved_by = request.user
                        bill.approved_at = timezone.now()
                        bill.save()

                        # Audit logging
                        from apps.core.audit_models import AuditLog

                        AuditLog.objects.create(
                            tenant=request.user.tenant,
                            user=request.user,
                            category=AuditLog.CATEGORY_DATA,
                            action=AuditLog.ACTION_CREATE,
                            severity=AuditLog.SEVERITY_INFO,
                            description=f"Created bill: {bill.bill_number} for supplier {bill.supplier.name}",
                            metadata={
                                "bill_id": str(bill.id),
                                "bill_number": bill.bill_number,
                                "supplier": bill.supplier.name,
                                "total": str(bill.total),
                            },
                            ip_address=request.META.get("REMOTE_ADDR"),
                            user_agent=request.META.get("HTTP_USER_AGENT", ""),
                            request_method=request.method,
                            request_path=request.path,
                        )

                        messages.success(
                            request,
                            f"Bill {bill.bill_number} created successfully with journal entry.",
                        )
                        return redirect("accounting:bill_detail", pk=bill.id)
                    else:
                        # Formset validation failed
                        messages.error(request, "Please correct the errors in the bill lines.")
                        # Delete the bill since formset failed
                        bill.delete()
            else:
                # Form validation failed
                from .bill_models import Bill as BillModel

                temp_bill = BillModel()
                formset = BillLineInlineFormSet(
                    request.POST, instance=temp_bill, tenant=request.user.tenant, coa=coa
                )
        else:
            form = BillForm(tenant=request.user.tenant, user=request.user)
            # Create a temporary unsaved Bill instance for the formset
            from .bill_models import Bill as BillModel

            temp_bill = BillModel()
            formset = BillLineInlineFormSet(instance=temp_bill, tenant=request.user.tenant, coa=coa)

        context = {
            "form": form,
            "formset": formset,
            "page_title": "Create Bill",
        }

        return render(request, "accounting/bills/form.html", context)

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")
    except Exception as e:
        logger.error(f"Error in bill_create: {str(e)}")
        import traceback

        logger.error(f"Traceback: {traceback.format_exc()}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("accounting:bill_list")


@login_required
@tenant_access_required
def bill_detail(request, pk):
    """
    Display bill details with payment history.

    Shows all line items, payments, and current balance.

    Requirements: 2.1, 2.2, 2.4, 2.7, 2.8
    """
    from .bill_models import Bill

    try:
        # Get bill with tenant filtering
        bill = (
            Bill.objects.filter(tenant=request.user.tenant, id=pk)
            .select_related("supplier", "created_by", "approved_by", "journal_entry")
            .prefetch_related("lines", "payments__created_by")
            .first()
        )

        if not bill:
            messages.error(request, "Bill not found.")
            return redirect("accounting:bill_list")

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description=f"Viewed bill detail: {bill.bill_number}",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        context = {
            "bill": bill,
            "page_title": f"Bill: {bill.bill_number}",
        }

        return render(request, "accounting/bills/detail.html", context)

    except Exception as e:
        logger.error(f"Error in bill_detail: {str(e)}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("accounting:bill_list")


@login_required
@tenant_access_required
def bill_pay(request, pk):
    """
    Record a payment against a bill with automatic journal entry creation.

    Creates journal entry debiting accounts payable and crediting cash/bank.

    Requirements: 2.4, 2.6, 2.7, 2.8
    """
    from .bill_models import Bill

    try:
        jewelry_entity = JewelryEntity.objects.get(tenant=request.user.tenant)
        entity = jewelry_entity.ledger_entity

        # Get chart of accounts
        coa = entity.chartofaccountmodel_set.first()
        if not coa:
            messages.error(request, "Chart of Accounts not found. Please set up accounting first.")
            return redirect("accounting:dashboard")

        # Get bill with tenant filtering
        bill = (
            Bill.objects.filter(tenant=request.user.tenant, id=pk)
            .select_related("supplier")
            .first()
        )

        if not bill:
            messages.error(request, "Bill not found.")
            return redirect("accounting:bill_list")

        # Check if bill is already paid
        if bill.status == "PAID":
            messages.warning(request, "This bill is already fully paid.")
            return redirect("accounting:bill_detail", pk=pk)

        # Check if bill is void
        if bill.status == "VOID":
            messages.error(request, "Cannot pay a void bill.")
            return redirect("accounting:bill_detail", pk=pk)

        from .forms import BillPaymentForm

        if request.method == "POST":
            form = BillPaymentForm(
                request.POST, tenant=request.user.tenant, user=request.user, bill=bill
            )

            if form.is_valid():
                with transaction.atomic():
                    # Create payment
                    payment = form.save(commit=False)
                    payment.tenant = request.user.tenant
                    payment.bill = bill
                    payment.created_by = request.user

                    # Create journal entry (debit AP, credit cash)
                    journal_entry = _create_payment_journal_entry(
                        payment, bill, request.user.tenant, entity, coa
                    )
                    payment.journal_entry = journal_entry
                    payment.save()

                    # Update bill's amount_paid and status
                    # This is handled automatically by BillPayment.save() which calls
                    # bill.add_payment()

                    # Audit logging
                    from apps.core.audit_models import AuditLog

                    AuditLog.objects.create(
                        tenant=request.user.tenant,
                        user=request.user,
                        category=AuditLog.CATEGORY_DATA,
                        action=AuditLog.ACTION_CREATE,
                        severity=AuditLog.SEVERITY_INFO,
                        description=f"Recorded payment of ${payment.amount:,.2f} for bill {bill.bill_number}",
                        metadata={
                            "payment_id": str(payment.id),
                            "bill_id": str(bill.id),
                            "bill_number": bill.bill_number,
                            "amount": str(payment.amount),
                            "payment_method": payment.payment_method,
                        },
                        ip_address=request.META.get("REMOTE_ADDR"),
                        user_agent=request.META.get("HTTP_USER_AGENT", ""),
                        request_method=request.method,
                        request_path=request.path,
                    )

                    messages.success(
                        request,
                        f"Payment of ${payment.amount:,.2f} recorded successfully with journal entry.",
                    )
                    return redirect("accounting:bill_detail", pk=pk)
        else:
            form = BillPaymentForm(tenant=request.user.tenant, user=request.user, bill=bill)

        context = {
            "form": form,
            "bill": bill,
            "page_title": f"Record Payment: {bill.bill_number}",
        }

        return render(request, "accounting/bills/payment_form.html", context)

    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting not set up for this tenant.")
        return redirect("accounting:dashboard")
    except Exception as e:
        logger.error(f"Error in bill_pay: {str(e)}")
        import traceback

        logger.error(f"Traceback: {traceback.format_exc()}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("accounting:bill_detail", pk=pk)


# ============================================================================
# Helper Functions for Journal Entry Creation
# ============================================================================


def _create_bill_journal_entry(bill, tenant, entity, coa):
    """
    Create journal entry for a bill.

    Debits: Expense/Asset accounts (from line items)
    Credits: Accounts Payable

    Requirements: 2.2, 2.7
    """
    from django_ledger.models import JournalEntryModel, TransactionModel

    # Get the ledger
    ledger = entity.ledgermodel_set.first()

    # Get accounts payable account
    ap_account = AccountModel.objects.filter(
        coa_model=coa, role__in=["LIABILITY_CL_ACC_PAYABLE", "LIABILITY_CL"], active=True
    ).first()

    if not ap_account:
        raise ValueError("Accounts Payable account not found in chart of accounts")

    # Create journal entry
    journal_entry = JournalEntryModel.objects.create(
        ledger=ledger,
        description=f"Bill {bill.bill_number} from {bill.supplier.name}",
        posted=True,  # Auto-post bill entries
    )

    # Create debit transactions for each line item (expense/asset)
    for line in bill.lines.all():
        # Get the account from the line's account code
        account = AccountModel.objects.filter(coa_model=coa, code=line.account, active=True).first()

        if not account:
            raise ValueError(f"Account {line.account} not found in chart of accounts")

        TransactionModel.objects.create(
            journal_entry=journal_entry,
            account=account,
            amount=line.amount,
            tx_type="debit",
            description=line.description,
        )

    # Create credit transaction for accounts payable
    TransactionModel.objects.create(
        journal_entry=journal_entry,
        account=ap_account,
        amount=bill.total,
        tx_type="credit",
        description=f"Bill {bill.bill_number} - {bill.supplier.name}",
    )

    return journal_entry


def _create_payment_journal_entry(payment, bill, tenant, entity, coa):
    """
    Create journal entry for a bill payment.

    Debits: Accounts Payable
    Credits: Cash/Bank account

    Requirements: 2.4, 2.6, 2.7
    """
    from django_ledger.models import JournalEntryModel, TransactionModel

    # Get the ledger
    ledger = entity.ledgermodel_set.first()

    # Get accounts payable account
    ap_account = AccountModel.objects.filter(
        coa_model=coa, role__in=["LIABILITY_CL_ACC_PAYABLE", "LIABILITY_CL"], active=True
    ).first()

    if not ap_account:
        raise ValueError("Accounts Payable account not found in chart of accounts")

    # Get cash/bank account based on payment method
    if payment.payment_method in ["CASH"]:
        cash_account = AccountModel.objects.filter(
            coa_model=coa, role="ASSET_CA_CASH", active=True
        ).first()
    else:
        # For checks, cards, transfers, use checking account
        cash_account = AccountModel.objects.filter(
            coa_model=coa, role__in=["ASSET_CA_CHECKING", "ASSET_CA"], active=True
        ).first()

    if not cash_account:
        raise ValueError("Cash/Bank account not found in chart of accounts")

    # Create journal entry
    journal_entry = JournalEntryModel.objects.create(
        ledger=ledger,
        description=f"Payment for Bill {bill.bill_number} - {bill.supplier.name}",
        posted=True,  # Auto-post payment entries
    )

    # Create debit transaction for accounts payable
    TransactionModel.objects.create(
        journal_entry=journal_entry,
        account=ap_account,
        amount=payment.amount,
        tx_type="debit",
        description=f"Payment for Bill {bill.bill_number}",
    )

    # Create credit transaction for cash/bank
    TransactionModel.objects.create(
        journal_entry=journal_entry,
        account=cash_account,
        amount=payment.amount,
        tx_type="credit",
        description=f"Payment to {bill.supplier.name} - {payment.payment_method}",
    )

    return journal_entry


@login_required
@tenant_access_required
def aged_payables_report(request):  # noqa: C901
    """
    Display aged payables report with 30/60/90/90+ day buckets.

    Shows amounts owed to suppliers grouped by aging buckets.
    Supports PDF and Excel export.

    Requirements: 2.5
    """
    from collections import defaultdict
    from datetime import date

    from .bill_models import Bill

    # Get as_of_date from request or default to today
    as_of_date_str = request.GET.get("as_of_date")
    if as_of_date_str:
        try:
            as_of_date = datetime.strptime(as_of_date_str, "%Y-%m-%d").date()
        except ValueError:
            as_of_date = date.today()
    else:
        as_of_date = date.today()

    # Get all unpaid bills for the tenant
    bills = (
        Bill.objects.filter(
            tenant=request.user.tenant,
            status__in=["APPROVED", "PARTIALLY_PAID"],
        )
        .select_related("supplier")
        .order_by("supplier__name", "due_date")
    )

    # Group bills by supplier and calculate aging buckets
    supplier_data = defaultdict(
        lambda: {
            "supplier": None,
            "current": 0,
            "days_1_30": 0,
            "days_31_60": 0,
            "days_61_90": 0,
            "days_90_plus": 0,
            "total": 0,
            "bills": [],
        }
    )

    for bill in bills:
        supplier_id = bill.supplier.id
        amount_due = bill.amount_due

        # Store supplier reference
        if supplier_data[supplier_id]["supplier"] is None:
            supplier_data[supplier_id]["supplier"] = bill.supplier

        # Calculate days overdue based on as_of_date
        days_overdue = (as_of_date - bill.due_date).days if as_of_date > bill.due_date else 0

        # Categorize into aging buckets
        if days_overdue <= 0:
            supplier_data[supplier_id]["current"] += amount_due
        elif days_overdue <= 30:
            supplier_data[supplier_id]["days_1_30"] += amount_due
        elif days_overdue <= 60:
            supplier_data[supplier_id]["days_31_60"] += amount_due
        elif days_overdue <= 90:
            supplier_data[supplier_id]["days_61_90"] += amount_due
        else:
            supplier_data[supplier_id]["days_90_plus"] += amount_due

        supplier_data[supplier_id]["total"] += amount_due
        supplier_data[supplier_id]["bills"].append(
            {
                "bill_number": bill.bill_number,
                "bill_date": bill.bill_date,
                "due_date": bill.due_date,
                "amount_due": amount_due,
                "days_overdue": days_overdue,
            }
        )

    # Convert to list and sort by supplier name
    supplier_list = sorted(
        supplier_data.values(), key=lambda x: x["supplier"].name if x["supplier"] else ""
    )

    # Calculate grand totals
    grand_totals = {
        "current": sum(s["current"] for s in supplier_list),
        "days_1_30": sum(s["days_1_30"] for s in supplier_list),
        "days_31_60": sum(s["days_31_60"] for s in supplier_list),
        "days_61_90": sum(s["days_61_90"] for s in supplier_list),
        "days_90_plus": sum(s["days_90_plus"] for s in supplier_list),
        "total": sum(s["total"] for s in supplier_list),
    }

    # Check for export format
    export_format = request.GET.get("export")

    if export_format == "pdf":
        return _export_aged_payables_pdf(
            request.user.tenant, supplier_list, grand_totals, as_of_date
        )
    elif export_format == "excel":
        return _export_aged_payables_excel(
            request.user.tenant, supplier_list, grand_totals, as_of_date
        )

    context = {
        "supplier_list": supplier_list,
        "grand_totals": grand_totals,
        "as_of_date": as_of_date,
        "page_title": "Aged Payables Report",
    }

    return render(request, "accounting/reports/aged_payables.html", context)


def _export_aged_payables_pdf(tenant, supplier_list, grand_totals, as_of_date):
    """
    Export aged payables report as PDF.
    """
    from io import BytesIO

    from django.http import HttpResponse

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    # Create response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="aged_payables_{as_of_date.strftime("%Y%m%d")}.pdf"'
    )

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5 * inch)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(
        f"<b>{tenant.company_name}</b><br/>Aged Payables Report<br/>As of {as_of_date.strftime('%B %d, %Y')}",
        styles["Title"],
    )
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))

    # Table data
    table_data = [
        ["Supplier", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total"]
    ]

    for supplier_data in supplier_list:
        table_data.append(
            [
                supplier_data["supplier"].name,
                f"${supplier_data['current']:,.2f}",
                f"${supplier_data['days_1_30']:,.2f}",
                f"${supplier_data['days_31_60']:,.2f}",
                f"${supplier_data['days_61_90']:,.2f}",
                f"${supplier_data['days_90_plus']:,.2f}",
                f"${supplier_data['total']:,.2f}",
            ]
        )

    # Add grand totals row
    table_data.append(
        [
            "TOTAL",
            f"${grand_totals['current']:,.2f}",
            f"${grand_totals['days_1_30']:,.2f}",
            f"${grand_totals['days_31_60']:,.2f}",
            f"${grand_totals['days_61_90']:,.2f}",
            f"${grand_totals['days_90_plus']:,.2f}",
            f"${grand_totals['total']:,.2f}",
        ]
    )

    # Create table
    table = Table(
        table_data,
        colWidths=[
            2.5 * inch,
            1.2 * inch,
            1.2 * inch,
            1.2 * inch,
            1.2 * inch,
            1.2 * inch,
            1.2 * inch,
        ],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)

    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


def _export_aged_payables_excel(tenant, supplier_list, grand_totals, as_of_date):
    """
    Export aged payables report as Excel.
    """
    from django.http import HttpResponse

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Aged Payables"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{tenant.company_name} - Aged Payables Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"As of {as_of_date.strftime('%B %d, %Y')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Supplier", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total"]
    ws.append([])  # Empty row
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for supplier_data in supplier_list:
        ws.append(
            [
                supplier_data["supplier"].name,
                supplier_data["current"],
                supplier_data["days_1_30"],
                supplier_data["days_31_60"],
                supplier_data["days_61_90"],
                supplier_data["days_90_plus"],
                supplier_data["total"],
            ]
        )

    # Grand totals row
    total_row = ws.max_row + 1
    ws.append(
        [
            "TOTAL",
            grand_totals["current"],
            grand_totals["days_1_30"],
            grand_totals["days_31_60"],
            grand_totals["days_61_90"],
            grand_totals["days_90_plus"],
            grand_totals["total"],
        ]
    )

    # Style totals row
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_font = Font(bold=True)
    for col_num in range(1, 8):
        cell = ws.cell(row=total_row, column=col_num)
        cell.fill = total_fill
        cell.font = total_font

    # Format currency columns
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="aged_payables_{as_of_date.strftime("%Y%m%d")}.xlsx"'
    )

    wb.save(response)
    return response


# ============================================================================
# Customer Accounting Views (Task 3.3)
# ============================================================================


@login_required
@tenant_access_required
def customer_accounting_detail(request, customer_id):  # noqa: C901
    """
    Customer accounting detail view with invoices, payments, and balance.

    Extends existing customer detail with accounting information:
    - Outstanding invoices
    - Payment history
    - Current balance
    - Credit limit status
    - Total sales

    Implements Requirements: 3.7, 15.1, 15.2, 15.3, 15.4, 15.7, 15.8
    """
    from apps.crm.models import Customer

    from .invoice_models import Invoice, InvoicePayment

    # Get customer with tenant filtering
    customer = get_object_or_404(
        Customer.objects.select_related("loyalty_tier"),
        id=customer_id,
        tenant=request.user.tenant,
    )

    # Get all invoices for this customer
    invoices = (
        Invoice.objects.filter(customer=customer, tenant=request.user.tenant)
        .select_related("journal_entry")
        .order_by("-invoice_date")
    )

    # Calculate outstanding balance
    outstanding_invoices = invoices.exclude(status__in=["PAID", "VOID"])
    outstanding_balance = sum(
        invoice.total - invoice.amount_paid for invoice in outstanding_invoices
    )

    # Get recent payments
    recent_payments = (
        InvoicePayment.objects.filter(invoice__customer=customer, tenant=request.user.tenant)
        .select_related("invoice", "created_by")
        .order_by("-payment_date")[:20]
    )

    # Calculate total sales
    total_sales = sum(
        invoice.total for invoice in invoices.filter(status__in=["SENT", "PARTIALLY_PAID", "PAID"])
    )

    # Calculate credit utilization
    credit_utilization_pct = 0
    if customer.credit_limit > 0:
        credit_utilization_pct = (outstanding_balance / customer.credit_limit) * 100

    # Check if customer is over credit limit
    over_credit_limit = (
        outstanding_balance > customer.credit_limit if customer.credit_limit > 0 else False
    )

    # Calculate payment statistics
    paid_invoices = invoices.filter(status="PAID")
    if paid_invoices.exists():
        # Calculate average days to pay
        total_days = 0
        count = 0
        for invoice in paid_invoices:
            if invoice.amount_paid >= invoice.total:
                # Find the last payment that completed the invoice
                last_payment = (
                    InvoicePayment.objects.filter(invoice=invoice).order_by("-payment_date").first()
                )
                if last_payment:
                    days_to_pay = (last_payment.payment_date - invoice.invoice_date).days
                    total_days += days_to_pay
                    count += 1

        avg_days_to_pay = total_days / count if count > 0 else 0
    else:
        avg_days_to_pay = 0

    # Calculate payment reliability score (0-100)
    # Based on: on-time payments, average days to pay, credit utilization
    payment_reliability_score = 100
    if paid_invoices.exists():
        # Deduct points for late payments
        overdue_count = invoices.filter(status="OVERDUE").count()
        if overdue_count > 0:
            payment_reliability_score -= min(overdue_count * 10, 40)

        # Deduct points for slow payment
        if avg_days_to_pay > 30:
            payment_reliability_score -= min((avg_days_to_pay - 30) / 2, 30)

        # Deduct points for high credit utilization
        if credit_utilization_pct > 80:
            payment_reliability_score -= 20
        elif credit_utilization_pct > 50:
            payment_reliability_score -= 10

    payment_reliability_score = max(0, payment_reliability_score)

    # Audit logging
    from apps.core.audit_models import AuditLog

    AuditLog.objects.create(
        tenant=request.user.tenant,
        user=request.user,
        category=AuditLog.CATEGORY_DATA,
        action=AuditLog.ACTION_API_GET,
        severity=AuditLog.SEVERITY_INFO,
        description=f"Viewed customer accounting detail for {customer.get_full_name()} (ID: {customer_id})",
        ip_address=request.META.get("REMOTE_ADDR"),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        request_method=request.method,
        request_path=request.path,
    )

    context = {
        "customer": customer,
        "invoices": invoices[:50],  # Limit to 50 most recent
        "outstanding_invoices": outstanding_invoices,
        "outstanding_balance": outstanding_balance,
        "recent_payments": recent_payments,
        "total_sales": total_sales,
        "credit_utilization_pct": credit_utilization_pct,
        "over_credit_limit": over_credit_limit,
        "avg_days_to_pay": avg_days_to_pay,
        "payment_reliability_score": payment_reliability_score,
        "page_title": f"Customer Accounting: {customer.get_full_name()}",
    }

    return render(request, "accounting/customers/accounting_detail.html", context)


@login_required
@tenant_access_required
def customer_statement(request, customer_id):
    """
    Generate customer statement showing all transactions and current balance.

    Shows:
    - Customer information
    - All invoices with dates and amounts
    - All payments with dates and amounts
    - Running balance
    - Current outstanding balance

    Implements Requirements: 15.3, 15.8
    """
    from apps.crm.models import Customer

    from .invoice_models import Invoice, InvoicePayment

    # Get customer with tenant filtering
    customer = get_object_or_404(
        Customer.objects.select_related("loyalty_tier"),
        id=customer_id,
        tenant=request.user.tenant,
    )

    # Get date range from request or default to last 90 days
    from datetime import timedelta

    end_date = date.today()
    start_date = end_date - timedelta(days=90)

    if request.GET.get("start_date"):
        start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
    if request.GET.get("end_date"):
        end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

    # Get invoices in date range
    invoices = (
        Invoice.objects.filter(
            customer=customer,
            tenant=request.user.tenant,
            invoice_date__gte=start_date,
            invoice_date__lte=end_date,
        )
        .exclude(status="VOID")
        .order_by("invoice_date")
    )

    # Get payments in date range
    payments = (
        InvoicePayment.objects.filter(
            invoice__customer=customer,
            tenant=request.user.tenant,
            payment_date__gte=start_date,
            payment_date__lte=end_date,
        )
        .select_related("invoice")
        .order_by("payment_date")
    )

    # Combine invoices and payments into a single transaction list
    transactions = []

    for invoice in invoices:
        transactions.append(
            {
                "date": invoice.invoice_date,
                "type": "invoice",
                "reference": invoice.invoice_number,
                "description": f"Invoice #{invoice.invoice_number}",
                "debit": invoice.total,
                "credit": Decimal("0.00"),
                "invoice": invoice,
            }
        )

    for payment in payments:
        transactions.append(
            {
                "date": payment.payment_date,
                "type": "payment",
                "reference": payment.reference_number or f"Payment #{payment.id}",
                "description": f"Payment for Invoice #{payment.invoice.invoice_number}",
                "debit": Decimal("0.00"),
                "credit": payment.amount,
                "payment": payment,
            }
        )

    # Sort by date
    transactions.sort(key=lambda x: x["date"])

    # Calculate running balance
    balance = Decimal("0.00")
    for txn in transactions:
        balance += txn["debit"] - txn["credit"]
        txn["balance"] = balance

    # Calculate current outstanding balance (all unpaid invoices)
    current_outstanding = sum(
        invoice.total - invoice.amount_paid
        for invoice in Invoice.objects.filter(
            customer=customer, tenant=request.user.tenant
        ).exclude(status__in=["PAID", "VOID"])
    )

    # Audit logging
    from apps.core.audit_models import AuditLog

    AuditLog.objects.create(
        tenant=request.user.tenant,
        user=request.user,
        category=AuditLog.CATEGORY_DATA,
        action=AuditLog.ACTION_API_GET,
        severity=AuditLog.SEVERITY_INFO,
        description=f"Generated customer statement for {customer.get_full_name()} (ID: {customer_id})",
        ip_address=request.META.get("REMOTE_ADDR"),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        request_method=request.method,
        request_path=request.path,
    )

    context = {
        "customer": customer,
        "transactions": transactions,
        "start_date": start_date,
        "end_date": end_date,
        "current_outstanding": current_outstanding,
        "page_title": f"Statement: {customer.get_full_name()}",
    }

    return render(request, "accounting/customers/statement.html", context)


@login_required
@tenant_access_required
@require_http_methods(["GET"])
def customer_statement_pdf(request, customer_id):  # noqa: C901
    """
    Export customer statement as PDF.

    Generates a PDF version of the customer statement showing all transactions
    and current balance for the specified date range.

    Implements Requirements: 3.8, 15.3
    """
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from apps.crm.models import Customer

    from .invoice_models import Invoice, InvoicePayment

    try:
        # Get customer with tenant filtering
        customer = get_object_or_404(
            Customer.objects.select_related("loyalty_tier"),
            id=customer_id,
            tenant=request.user.tenant,
        )

        # Get date range from request or default to last 90 days
        from datetime import timedelta

        end_date = date.today()
        start_date = end_date - timedelta(days=90)

        if request.GET.get("start_date"):
            start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
        if request.GET.get("end_date"):
            end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

        # Get invoices in date range
        invoices = (
            Invoice.objects.filter(
                customer=customer,
                tenant=request.user.tenant,
                invoice_date__gte=start_date,
                invoice_date__lte=end_date,
            )
            .exclude(status="VOID")
            .order_by("invoice_date")
        )

        # Get payments in date range
        payments = (
            InvoicePayment.objects.filter(
                invoice__customer=customer,
                tenant=request.user.tenant,
                payment_date__gte=start_date,
                payment_date__lte=end_date,
            )
            .select_related("invoice")
            .order_by("payment_date")
        )

        # Calculate opening balance (all invoices and payments before start_date)
        opening_invoices = Invoice.objects.filter(
            customer=customer, tenant=request.user.tenant, invoice_date__lt=start_date
        ).exclude(status="VOID").aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")

        opening_payments = InvoicePayment.objects.filter(
            invoice__customer=customer, tenant=request.user.tenant, payment_date__lt=start_date
        ).aggregate(total=models.Sum("amount"))["total"] or Decimal("0.00")

        opening_balance = opening_invoices - opening_payments

        # Calculate period totals
        period_invoices_total = invoices.aggregate(total=models.Sum("total"))["total"] or Decimal(
            "0.00"
        )
        period_payments_total = payments.aggregate(total=models.Sum("amount"))["total"] or Decimal(
            "0.00"
        )

        # Calculate closing balance
        closing_balance = opening_balance + period_invoices_total - period_payments_total

        # Combine invoices and payments into a single transaction list
        transactions = []

        for invoice in invoices:
            transactions.append(
                {
                    "date": invoice.invoice_date,
                    "type": "Invoice",
                    "reference": invoice.invoice_number,
                    "description": f"Invoice #{invoice.invoice_number}",
                    "debit": invoice.total,
                    "credit": Decimal("0.00"),
                    "balance": None,
                }
            )

        for payment in payments:
            transactions.append(
                {
                    "date": payment.payment_date,
                    "type": "Payment",
                    "reference": payment.reference_number or f"Payment #{payment.id}",
                    "description": f"Payment for Invoice #{payment.invoice.invoice_number}",
                    "debit": Decimal("0.00"),
                    "credit": payment.amount,
                    "balance": None,
                }
            )

        # Sort by date
        transactions.sort(key=lambda x: x["date"])

        # Calculate running balance
        running_balance = opening_balance
        for txn in transactions:
            running_balance += txn["debit"] - txn["credit"]
            txn["balance"] = running_balance

        # Calculate current outstanding balance (all unpaid invoices)
        current_outstanding = sum(
            invoice.total - invoice.amount_paid
            for invoice in Invoice.objects.filter(
                customer=customer, tenant=request.user.tenant
            ).exclude(status__in=["PAID", "VOID"])
        )

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch
        )
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1f2937"),
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#374151"),
        )
        normal_style = styles["Normal"]

        # Title
        elements.append(Paragraph("Customer Statement", title_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Customer Information and Period
        # Build address from customer fields
        address_parts = []
        if customer.address_line_1:
            address_parts.append(customer.address_line_1)
        if customer.address_line_2:
            address_parts.append(customer.address_line_2)
        if customer.city or customer.state or customer.postal_code:
            city_state_zip = ", ".join(
                filter(None, [customer.city, customer.state, customer.postal_code])
            )
            if city_state_zip:
                address_parts.append(city_state_zip)
        if customer.country:
            address_parts.append(customer.country)
        address_str = "<br/>".join(address_parts) if address_parts else ""

        customer_info = f"""
        <b>{customer.get_full_name()}</b><br/>
        {customer.customer_number}<br/>
        {address_str}<br/>
        {customer.email or ''}<br/>
        {customer.phone or ''}
        """

        period_info = f"""
        <b>{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}</b><br/>
        Generated on {date.today().strftime('%B %d, %Y')}
        """

        info_data = [
            ["Customer Information", "Statement Period"],
            [
                Paragraph(customer_info, normal_style),
                Paragraph(period_info, normal_style),
            ],
        ]

        info_table = Table(info_data, colWidths=[3.5 * inch, 3.5 * inch])
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("ALIGN", (0, 0), (-1, 0), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("TOPPADDING", (0, 0), (-1, 0), 12),
                    ("VALIGN", (0, 1), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        elements.append(info_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Balance Summary
        balance_data = [
            ["Opening Balance", "Period Charges", "Period Payments", "Closing Balance"],
            [
                f"${opening_balance:,.2f}",
                f"${period_invoices_total:,.2f}",
                f"${period_payments_total:,.2f}",
                f"${closing_balance:,.2f}",
            ],
        ]

        balance_table = Table(balance_data, colWidths=[1.75 * inch] * 4)
        balance_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("TOPPADDING", (0, 0), (-1, -1), 12),
                    ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 1), (-1, 1), 12),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        elements.append(balance_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Transaction Details
        elements.append(Paragraph("Transaction Details", heading_style))
        elements.append(Spacer(1, 0.2 * inch))

        if transactions:
            # Transaction table header
            txn_data = [
                ["Date", "Type", "Reference", "Description", "Charges", "Payments", "Balance"]
            ]

            # Opening balance row
            txn_data.append(
                [
                    start_date.strftime("%Y-%m-%d"),
                    "",
                    "",
                    "Opening Balance",
                    "",
                    "",
                    f"${opening_balance:,.2f}",
                ]
            )

            # Transaction rows
            for txn in transactions:
                txn_data.append(
                    [
                        txn["date"].strftime("%Y-%m-%d"),
                        txn["type"],
                        txn["reference"],
                        txn["description"],
                        f"${txn['debit']:,.2f}" if txn["debit"] > 0 else "-",
                        f"${txn['credit']:,.2f}" if txn["credit"] > 0 else "-",
                        f"${txn['balance']:,.2f}",
                    ]
                )

            # Closing balance row
            txn_data.append(
                [
                    end_date.strftime("%Y-%m-%d"),
                    "",
                    "",
                    "Closing Balance",
                    f"${period_invoices_total:,.2f}",
                    f"${period_payments_total:,.2f}",
                    f"${closing_balance:,.2f}",
                ]
            )

            txn_table = Table(
                txn_data,
                colWidths=[
                    0.8 * inch,
                    0.6 * inch,
                    0.9 * inch,
                    1.8 * inch,
                    0.9 * inch,
                    0.9 * inch,
                    0.9 * inch,
                ],
            )
            txn_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#dbeafe")),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f3f4f6")),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            elements.append(txn_table)
        else:
            elements.append(
                Paragraph("No transactions found for the selected period.", normal_style)
            )

        # Footer
        elements.append(Spacer(1, 0.5 * inch))
        if current_outstanding > 0:
            footer_text = (
                f"<b>Current Amount Due: ${current_outstanding:,.2f}</b><br/><br/>"
                "This is a computer-generated statement and does not require a signature.<br/>"
                "For questions about this statement, please contact your accounting department."
            )
        else:
            footer_text = (
                "This is a computer-generated statement and does not require a signature.<br/>"
                "For questions about this statement, please contact your accounting department."
            )
        elements.append(Paragraph(footer_text, normal_style))

        # Build PDF
        doc.build(elements)

        # Audit logging
        from apps.core.audit_models import AuditLog

        AuditLog.objects.create(
            tenant=request.user.tenant,
            user=request.user,
            category=AuditLog.CATEGORY_DATA,
            action=AuditLog.ACTION_API_GET,
            severity=AuditLog.SEVERITY_INFO,
            description=f"Exported customer statement PDF: {customer.get_full_name()} ({start_date} to {end_date})",
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
            request_method=request.method,
            request_path=request.path,
        )

        # Return PDF response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        customer_name = customer.get_full_name().replace(" ", "_")
        response["Content-Disposition"] = (
            f'attachment; filename="customer_statement_{customer_name}_'
            f'{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.pdf"'
        )
        return response

    except Exception as e:
        logger.error(f"Failed to generate customer statement PDF: {str(e)}")
        messages.error(request, f"Failed to generate PDF: {str(e)}")
        return redirect("accounting:customer_accounting_detail", customer_id=customer_id)


def validate_customer_credit_limit(customer, additional_amount):
    """
    Validate if a customer can take on additional credit.

    Checks if the customer's outstanding balance plus the additional amount
    would exceed their credit limit.

    Args:
        customer: Customer instance
        additional_amount: Decimal amount to add to outstanding balance

    Returns:
        tuple: (is_valid, message, current_balance, credit_available)

    Implements Requirement: 15.4
    """
    from .invoice_models import Invoice

    # Calculate current outstanding balance
    outstanding_invoices = Invoice.objects.filter(
        customer=customer, tenant=customer.tenant
    ).exclude(status__in=["PAID", "VOID"])

    current_outstanding = sum(
        invoice.total - invoice.amount_paid for invoice in outstanding_invoices
    )

    # Check credit limit
    if customer.credit_limit <= 0:
        # No credit limit set, allow any amount
        return (True, "No credit limit set", current_outstanding, None)

    new_balance = current_outstanding + additional_amount
    credit_available = customer.credit_limit - current_outstanding

    if new_balance > customer.credit_limit:
        over_limit = new_balance - customer.credit_limit
        message = (
            f"Credit limit exceeded. "
            f"Current balance: ${current_outstanding:.2f}, "
            f"Credit limit: ${customer.credit_limit:.2f}, "
            f"Available credit: ${credit_available:.2f}, "
            f"Would exceed by: ${over_limit:.2f}"
        )
        return (False, message, current_outstanding, credit_available)

    message = f"Credit check passed. Available credit: ${credit_available:.2f}"
    return (True, message, current_outstanding, credit_available)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def check_customer_credit_limit_api(request, customer_id):
    """
    API endpoint to check customer credit limit.

    Used by invoice creation forms to validate credit before creating invoice.

    Implements Requirement: 15.4
    """
    from apps.crm.models import Customer

    # Get customer with tenant filtering
    customer = get_object_or_404(
        Customer,
        id=customer_id,
        tenant=request.user.tenant,
    )

    # Get amount from request
    try:
        amount = Decimal(request.POST.get("amount", "0"))
    except (ValueError, TypeError):
        return JsonResponse({"error": "Invalid amount"}, status=400)

    # Validate credit limit
    is_valid, message, current_balance, credit_available = validate_customer_credit_limit(
        customer, amount
    )

    # Audit logging
    from apps.core.audit_models import AuditLog

    AuditLog.objects.create(
        tenant=request.user.tenant,
        user=request.user,
        category=AuditLog.CATEGORY_DATA,
        action=AuditLog.ACTION_API_GET,
        severity=AuditLog.SEVERITY_WARNING if not is_valid else AuditLog.SEVERITY_INFO,
        description=f"Credit limit check for {customer.get_full_name()}: {message}",
        ip_address=request.META.get("REMOTE_ADDR"),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        request_method=request.method,
        request_path=request.path,
    )

    return JsonResponse(
        {
            "is_valid": is_valid,
            "message": message,
            "current_balance": str(current_balance),
            "credit_limit": str(customer.credit_limit),
            "credit_available": str(credit_available) if credit_available is not None else None,
        }
    )


# ============================================================================
# Aged Receivables Report (Task 3.7)
# ============================================================================


@login_required
@tenant_access_required
def aged_receivables_report(request):  # noqa: C901
    """
    Display aged receivables report with 30/60/90/90+ day buckets.

    Shows amounts owed by customers grouped by aging buckets.
    Supports PDF and Excel export.

    Requirements: 3.7
    """
    from collections import defaultdict
    from datetime import date

    from .invoice_models import Invoice

    # Get as_of_date from request or default to today
    as_of_date_str = request.GET.get("as_of_date")
    if as_of_date_str:
        try:
            as_of_date = datetime.strptime(as_of_date_str, "%Y-%m-%d").date()
        except ValueError:
            as_of_date = date.today()
    else:
        as_of_date = date.today()

    # Get all unpaid invoices for the tenant
    invoices = (
        Invoice.objects.filter(
            tenant=request.user.tenant,
            status__in=["SENT", "PARTIALLY_PAID", "OVERDUE"],
        )
        .select_related("customer")
        .order_by("customer__first_name", "customer__last_name", "due_date")
    )

    # Group invoices by customer and calculate aging buckets
    customer_data = defaultdict(
        lambda: {
            "customer": None,
            "current": 0,
            "days_1_30": 0,
            "days_31_60": 0,
            "days_61_90": 0,
            "days_90_plus": 0,
            "total": 0,
            "invoices": [],
        }
    )

    for invoice in invoices:
        customer_id = invoice.customer.id
        amount_due = invoice.total - invoice.amount_paid

        # Store customer reference
        if customer_data[customer_id]["customer"] is None:
            customer_data[customer_id]["customer"] = invoice.customer

        # Calculate days overdue based on as_of_date
        days_overdue = (as_of_date - invoice.due_date).days if as_of_date > invoice.due_date else 0

        # Categorize into aging buckets
        if days_overdue <= 0:
            customer_data[customer_id]["current"] += amount_due
        elif days_overdue <= 30:
            customer_data[customer_id]["days_1_30"] += amount_due
        elif days_overdue <= 60:
            customer_data[customer_id]["days_31_60"] += amount_due
        elif days_overdue <= 90:
            customer_data[customer_id]["days_61_90"] += amount_due
        else:
            customer_data[customer_id]["days_90_plus"] += amount_due

        customer_data[customer_id]["total"] += amount_due
        customer_data[customer_id]["invoices"].append(
            {
                "invoice_number": invoice.invoice_number,
                "invoice_date": invoice.invoice_date,
                "due_date": invoice.due_date,
                "amount_due": amount_due,
                "days_overdue": days_overdue,
            }
        )

    # Convert to list and sort by customer name
    customer_list = sorted(
        customer_data.values(), key=lambda x: x["customer"].get_full_name() if x["customer"] else ""
    )

    # Calculate grand totals
    grand_totals = {
        "current": sum(c["current"] for c in customer_list),
        "days_1_30": sum(c["days_1_30"] for c in customer_list),
        "days_31_60": sum(c["days_31_60"] for c in customer_list),
        "days_61_90": sum(c["days_61_90"] for c in customer_list),
        "days_90_plus": sum(c["days_90_plus"] for c in customer_list),
        "total": sum(c["total"] for c in customer_list),
    }

    # Check for export format
    export_format = request.GET.get("export")

    if export_format == "pdf":
        return _export_aged_receivables_pdf(
            request.user.tenant, customer_list, grand_totals, as_of_date
        )
    elif export_format == "excel":
        return _export_aged_receivables_excel(
            request.user.tenant, customer_list, grand_totals, as_of_date
        )

    context = {
        "customer_list": customer_list,
        "grand_totals": grand_totals,
        "as_of_date": as_of_date,
        "page_title": "Aged Receivables Report",
    }

    return render(request, "accounting/reports/aged_receivables.html", context)


def _export_aged_receivables_pdf(tenant, customer_list, grand_totals, as_of_date):
    """
    Export aged receivables report as PDF.
    """
    from io import BytesIO

    from django.http import HttpResponse

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    # Create response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="aged_receivables_{as_of_date.strftime("%Y%m%d")}.pdf"'
    )

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5 * inch)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(
        f"<b>{tenant.company_name}</b><br/>Aged Receivables Report<br/>As of {as_of_date.strftime('%B %d, %Y')}",
        styles["Title"],
    )
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))

    # Table data
    table_data = [
        ["Customer", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total"]
    ]

    for customer_data in customer_list:
        table_data.append(
            [
                customer_data["customer"].get_full_name(),
                f"${customer_data['current']:,.2f}",
                f"${customer_data['days_1_30']:,.2f}",
                f"${customer_data['days_31_60']:,.2f}",
                f"${customer_data['days_61_90']:,.2f}",
                f"${customer_data['days_90_plus']:,.2f}",
                f"${customer_data['total']:,.2f}",
            ]
        )

    # Add grand totals row
    table_data.append(
        [
            "TOTAL",
            f"${grand_totals['current']:,.2f}",
            f"${grand_totals['days_1_30']:,.2f}",
            f"${grand_totals['days_31_60']:,.2f}",
            f"${grand_totals['days_61_90']:,.2f}",
            f"${grand_totals['days_90_plus']:,.2f}",
            f"${grand_totals['total']:,.2f}",
        ]
    )

    # Create table
    table = Table(
        table_data,
        colWidths=[
            2.5 * inch,
            1.2 * inch,
            1.2 * inch,
            1.2 * inch,
            1.2 * inch,
            1.2 * inch,
            1.2 * inch,
        ],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)

    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


def _export_aged_receivables_excel(tenant, customer_list, grand_totals, as_of_date):
    """
    Export aged receivables report as Excel.
    """
    from django.http import HttpResponse

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Aged Receivables"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{tenant.company_name} - Aged Receivables Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"As of {as_of_date.strftime('%B %d, %Y')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Customer", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total"]
    ws.append([])  # Empty row
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for customer_data in customer_list:
        ws.append(
            [
                customer_data["customer"].get_full_name(),
                customer_data["current"],
                customer_data["days_1_30"],
                customer_data["days_31_60"],
                customer_data["days_61_90"],
                customer_data["days_90_plus"],
                customer_data["total"],
            ]
        )

    # Grand totals row
    total_row = ws.max_row + 1
    ws.append(
        [
            "TOTAL",
            grand_totals["current"],
            grand_totals["days_1_30"],
            grand_totals["days_31_60"],
            grand_totals["days_61_90"],
            grand_totals["days_90_plus"],
            grand_totals["total"],
        ]
    )

    # Style totals row
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_font = Font(bold=True)
    for col_num in range(1, 8):
        cell = ws.cell(row=total_row, column=col_num)
        cell.fill = total_fill
        cell.font = total_font

    # Format currency columns
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="aged_receivables_{as_of_date.strftime("%Y%m%d")}.xlsx"'
    )

    wb.save(response)
    return response


# ============================================================================
# Invoice Management Views (Task 3.5)
# ============================================================================


@login_required
@tenant_access_required
def invoice_list(request):
    """
    List all invoices for the tenant with filtering and aging information.

    Implements Requirements: 3.3, 3.7
    """
    from .invoice_models import Invoice

    # Get filter parameters
    status_filter = request.GET.get("status", "")
    customer_filter = request.GET.get("customer", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    search = request.GET.get("search", "")

    # Base queryset with tenant filtering
    invoices = Invoice.objects.filter(tenant=request.user.tenant).select_related("customer")

    # Apply filters
    if status_filter:
        invoices = invoices.filter(status=status_filter)

    if customer_filter:
        invoices = invoices.filter(customer_id=customer_filter)

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            invoices = invoices.filter(invoice_date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            invoices = invoices.filter(invoice_date__lte=date_to_obj)
        except ValueError:
            pass

    if search:
        invoices = invoices.filter(
            models.Q(invoice_number__icontains=search)
            | models.Q(customer__first_name__icontains=search)
            | models.Q(customer__last_name__icontains=search)
            | models.Q(reference_number__icontains=search)
        )

    # Order by date (newest first)
    invoices = invoices.order_by("-invoice_date", "-created_at")

    # Calculate summary statistics
    total_invoices = invoices.count()
    total_amount = invoices.aggregate(total=models.Sum("total"))["total"] or Decimal("0.00")
    total_outstanding = invoices.aggregate(
        total=models.Sum(models.F("total") - models.F("amount_paid"))
    )["total"] or Decimal("0.00")

    # Get customers for filter dropdown
    from apps.crm.models import Customer

    customers = Customer.objects.filter(tenant=request.user.tenant, is_active=True).order_by(
        "first_name", "last_name"
    )

    # Audit logging
    from apps.core.audit_models import AuditLog

    AuditLog.objects.create(
        tenant=request.user.tenant,
        user=request.user,
        category=AuditLog.CATEGORY_DATA,
        action=AuditLog.ACTION_API_GET,
        severity=AuditLog.SEVERITY_INFO,
        description=f"Viewed invoice list ({total_invoices} invoices)",
        ip_address=request.META.get("REMOTE_ADDR"),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        request_method=request.method,
        request_path=request.path,
    )

    context = {
        "invoices": invoices,
        "total_invoices": total_invoices,
        "total_amount": total_amount,
        "total_outstanding": total_outstanding,
        "customers": customers,
        "status_filter": status_filter,
        "customer_filter": customer_filter,
        "date_from": date_from,
        "date_to": date_to,
        "search": search,
        "status_choices": Invoice.STATUS_CHOICES,
        "page_title": "Invoices",
    }

    return render(request, "accounting/invoices/list.html", context)


@login_required
@tenant_access_required
def invoice_create(request):  # noqa: C901
    """
    Create a new customer invoice with line items.

    Automatically creates journal entry: DR AR, CR Revenue, CR Tax

    Implements Requirements: 3.1, 3.2, 3.7
    """
    from .forms import InvoiceForm, InvoiceLineInlineFormSet
    from .services import InvoiceService

    # Check if accounting is set up
    try:
        JewelryEntity.objects.get(tenant=request.user.tenant)
    except JewelryEntity.DoesNotExist:
        messages.error(request, "Accounting is not set up for your tenant. Please contact support.")
        return redirect("accounting:dashboard")

    if request.method == "POST":
        form = InvoiceForm(request.POST, tenant=request.user.tenant, user=request.user)
        formset = InvoiceLineInlineFormSet(request.POST, instance=None)

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    # Create invoice
                    invoice = form.save(commit=False)
                    invoice.tenant = request.user.tenant
                    invoice.created_by = request.user

                    # Calculate payment terms and due date from customer if not set
                    customer = invoice.customer
                    if hasattr(customer, "payment_terms") and customer.payment_terms:
                        # Parse payment terms (e.g., "NET30" -> 30 days)
                        try:
                            days = int(customer.payment_terms.replace("NET", "").strip())
                            from datetime import timedelta

                            invoice.due_date = invoice.invoice_date + timedelta(days=days)
                        except (ValueError, AttributeError):
                            pass

                    invoice.save()

                    # Save line items
                    formset.instance = invoice
                    lines = formset.save(commit=False)
                    for line in lines:
                        line.save()

                    # Calculate totals from line items
                    invoice.calculate_totals()

                    # Check credit limit
                    credit_check = InvoiceService.check_customer_credit_limit(
                        customer, invoice.total
                    )
                    if not credit_check["within_limit"]:
                        messages.warning(request, credit_check["warning_message"])

                    # Create automatic journal entry
                    try:
                        je = InvoiceService.create_invoice_journal_entry(invoice, request.user)
                        if je:
                            messages.success(
                                request,
                                f"Invoice {invoice.invoice_number} created successfully with journal entry.",
                            )
                        else:
                            messages.warning(
                                request,
                                f"Invoice {invoice.invoice_number} created but journal entry failed. "
                                "Please create manually.",
                            )
                    except Exception as e:
                        logger.error(f"Failed to create journal entry for invoice: {str(e)}")
                        messages.warning(
                            request,
                            f"Invoice created but journal entry failed: {str(e)}. "
                            "Please create manually.",
                        )

                    # Audit logging
                    from apps.core.audit_models import AuditLog

                    AuditLog.objects.create(
                        tenant=request.user.tenant,
                        user=request.user,
                        category=AuditLog.CATEGORY_DATA,
                        action=AuditLog.ACTION_CREATE,
                        severity=AuditLog.SEVERITY_INFO,
                        description=f"Created invoice {invoice.invoice_number} for {customer}",
                        after_value=f"Total: ${invoice.total:,.2f}",
                        ip_address=request.META.get("REMOTE_ADDR"),
                        user_agent=request.META.get("HTTP_USER_AGENT", ""),
                        request_method=request.method,
                        request_path=request.path,
                    )

                    return redirect("accounting:invoice_detail", invoice_id=invoice.id)

            except Exception as e:
                logger.error(f"Failed to create invoice: {str(e)}")
                messages.error(request, f"Failed to create invoice: {str(e)}")
    else:
        form = InvoiceForm(tenant=request.user.tenant, user=request.user)
        formset = InvoiceLineInlineFormSet(instance=None)

    context = {
        "form": form,
        "formset": formset,
        "page_title": "Create Invoice",
        "submit_text": "Create Invoice",
    }

    return render(request, "accounting/invoices/form.html", context)


@login_required
@tenant_access_required
def invoice_detail(request, invoice_id):
    """
    View invoice details with payment history and line items.

    Implements Requirements: 3.3, 3.7
    """
    from .invoice_models import Invoice

    # Get invoice with tenant filtering
    invoice = get_object_or_404(
        Invoice.objects.select_related("customer", "created_by", "journal_entry"),
        id=invoice_id,
        tenant=request.user.tenant,
    )

    # Get line items
    lines = invoice.lines.all()

    # Get payments
    payments = invoice.payments.select_related("created_by").order_by("-payment_date")

    # Get applied credit memos
    applied_credits = invoice.applied_credits.select_related("customer").order_by("-credit_date")

    # Calculate summary
    total_payments = payments.aggregate(total=models.Sum("amount"))["total"] or Decimal("0.00")
    total_credits = applied_credits.aggregate(total=models.Sum("amount_used"))["total"] or Decimal(
        "0.00"
    )

    # Audit logging
    from apps.core.audit_models import AuditLog

    AuditLog.objects.create(
        tenant=request.user.tenant,
        user=request.user,
        category=AuditLog.CATEGORY_DATA,
        action=AuditLog.ACTION_API_GET,
        severity=AuditLog.SEVERITY_INFO,
        description=f"Viewed invoice {invoice.invoice_number}",
        ip_address=request.META.get("REMOTE_ADDR"),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        request_method=request.method,
        request_path=request.path,
    )

    context = {
        "invoice": invoice,
        "lines": lines,
        "payments": payments,
        "applied_credits": applied_credits,
        "total_payments": total_payments,
        "total_credits": total_credits,
        "page_title": f"Invoice {invoice.invoice_number}",
    }

    return render(request, "accounting/invoices/detail.html", context)


@login_required
@tenant_access_required
def invoice_receive_payment(request, invoice_id):
    """
    Record a payment against an invoice.

    Automatically creates journal entry: DR Cash, CR AR

    Implements Requirements: 3.4, 3.7
    """
    from .forms import InvoicePaymentForm
    from .invoice_models import Invoice
    from .services import InvoiceService

    # Get invoice with tenant filtering
    invoice = get_object_or_404(
        Invoice.objects.select_related("customer"),
        id=invoice_id,
        tenant=request.user.tenant,
    )

    # Check if invoice can receive payment
    if invoice.status == "VOID":
        messages.error(request, "Cannot record payment for a void invoice.")
        return redirect("accounting:invoice_detail", invoice_id=invoice.id)

    if invoice.amount_due <= Decimal("0.00"):
        messages.error(request, "Invoice is already fully paid.")
        return redirect("accounting:invoice_detail", invoice_id=invoice.id)

    if request.method == "POST":
        form = InvoicePaymentForm(
            request.POST, tenant=request.user.tenant, user=request.user, invoice=invoice
        )

        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create payment
                    payment = form.save(commit=False)
                    payment.tenant = request.user.tenant
                    payment.invoice = invoice
                    payment.created_by = request.user
                    payment.save()

                    # Invoice amount_paid is updated automatically by the model's save method

                    # Create automatic journal entry
                    try:
                        je = InvoiceService.create_payment_journal_entry(payment, request.user)
                        if je:
                            messages.success(
                                request,
                                f"Payment of ${payment.amount:,.2f} recorded successfully with journal entry.",
                            )
                        else:
                            messages.warning(
                                request,
                                "Payment recorded but journal entry failed. Please create manually.",
                            )
                    except Exception as e:
                        logger.error(f"Failed to create journal entry for payment: {str(e)}")
                        messages.warning(
                            request,
                            f"Payment recorded but journal entry failed: {str(e)}. "
                            "Please create manually.",
                        )

                    # Audit logging
                    from apps.core.audit_models import AuditLog

                    AuditLog.objects.create(
                        tenant=request.user.tenant,
                        user=request.user,
                        category=AuditLog.CATEGORY_DATA,
                        action=AuditLog.ACTION_CREATE,
                        severity=AuditLog.SEVERITY_INFO,
                        description=f"Recorded payment of ${payment.amount:,.2f} for invoice {invoice.invoice_number}",
                        after_value=f"Remaining balance: ${invoice.amount_due:,.2f}",
                        ip_address=request.META.get("REMOTE_ADDR"),
                        user_agent=request.META.get("HTTP_USER_AGENT", ""),
                        request_method=request.method,
                        request_path=request.path,
                    )

                    return redirect("accounting:invoice_detail", invoice_id=invoice.id)

            except Exception as e:
                logger.error(f"Failed to record payment: {str(e)}")
                messages.error(request, f"Failed to record payment: {str(e)}")
    else:
        form = InvoicePaymentForm(tenant=request.user.tenant, user=request.user, invoice=invoice)

    context = {
        "form": form,
        "invoice": invoice,
        "page_title": f"Record Payment - Invoice {invoice.invoice_number}",
        "submit_text": "Record Payment",
    }

    return render(request, "accounting/invoices/payment_form.html", context)


@login_required
@tenant_access_required
def credit_memo_create(request):
    """
    Create a credit memo for a customer.

    Automatically creates journal entry: DR Sales Returns, CR AR

    Implements Requirements: 3.6, 3.7
    """
    from .forms import CreditMemoForm
    from .services import InvoiceService

    if request.method == "POST":
        form = CreditMemoForm(request.POST, tenant=request.user.tenant, user=request.user)

        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create credit memo
                    credit_memo = form.save(commit=False)
                    credit_memo.tenant = request.user.tenant
                    credit_memo.created_by = request.user
                    credit_memo.save()

                    # Create automatic journal entry
                    try:
                        je = InvoiceService.create_credit_memo_journal_entry(
                            credit_memo, request.user
                        )
                        if je:
                            messages.success(
                                request,
                                f"Credit memo {credit_memo.credit_memo_number} created successfully with journal entry.",
                            )
                        else:
                            messages.warning(
                                request,
                                "Credit memo created but journal entry failed. Please create manually.",
                            )
                    except Exception as e:
                        logger.error(f"Failed to create journal entry for credit memo: {str(e)}")
                        messages.warning(
                            request,
                            f"Credit memo created but journal entry failed: {str(e)}. "
                            "Please create manually.",
                        )

                    # Audit logging
                    from apps.core.audit_models import AuditLog

                    AuditLog.objects.create(
                        tenant=request.user.tenant,
                        user=request.user,
                        category=AuditLog.CATEGORY_DATA,
                        action=AuditLog.ACTION_CREATE,
                        severity=AuditLog.SEVERITY_INFO,
                        description=f"Created credit memo {credit_memo.credit_memo_number} for {credit_memo.customer}",
                        after_value=f"Amount: ${credit_memo.amount:,.2f}, Reason: {credit_memo.reason}",
                        ip_address=request.META.get("REMOTE_ADDR"),
                        user_agent=request.META.get("HTTP_USER_AGENT", ""),
                        request_method=request.method,
                        request_path=request.path,
                    )

                    return redirect("accounting:credit_memo_detail", credit_memo_id=credit_memo.id)

            except Exception as e:
                logger.error(f"Failed to create credit memo: {str(e)}")
                messages.error(request, f"Failed to create credit memo: {str(e)}")
    else:
        form = CreditMemoForm(tenant=request.user.tenant, user=request.user)

    context = {
        "form": form,
        "page_title": "Create Credit Memo",
        "submit_text": "Create Credit Memo",
    }

    return render(request, "accounting/credit_memos/form.html", context)


@login_required
@tenant_access_required
def credit_memo_detail(request, credit_memo_id):
    """
    View credit memo details.

    Implements Requirements: 3.6, 3.7
    """
    from .invoice_models import CreditMemo

    # Get credit memo with tenant filtering
    credit_memo = get_object_or_404(
        CreditMemo.objects.select_related(
            "customer", "created_by", "original_invoice", "applied_to_invoice", "journal_entry"
        ),
        id=credit_memo_id,
        tenant=request.user.tenant,
    )

    # Audit logging
    from apps.core.audit_models import AuditLog

    AuditLog.objects.create(
        tenant=request.user.tenant,
        user=request.user,
        category=AuditLog.CATEGORY_DATA,
        action=AuditLog.ACTION_API_GET,
        severity=AuditLog.SEVERITY_INFO,
        description=f"Viewed credit memo {credit_memo.credit_memo_number}",
        ip_address=request.META.get("REMOTE_ADDR"),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        request_method=request.method,
        request_path=request.path,
    )

    context = {
        "credit_memo": credit_memo,
        "page_title": f"Credit Memo {credit_memo.credit_memo_number}",
    }

    return render(request, "accounting/credit_memos/detail.html", context)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def credit_memo_apply(request, credit_memo_id, invoice_id):
    """
    Apply a credit memo to an invoice.

    Implements Requirements: 3.6, 3.7
    """
    from .invoice_models import CreditMemo, Invoice

    # Get credit memo and invoice with tenant filtering
    credit_memo = get_object_or_404(
        CreditMemo.objects.select_related("customer"),
        id=credit_memo_id,
        tenant=request.user.tenant,
    )

    invoice = get_object_or_404(
        Invoice.objects.select_related("customer"),
        id=invoice_id,
        tenant=request.user.tenant,
    )

    # Validate
    if credit_memo.customer != invoice.customer:
        messages.error(request, "Credit memo and invoice must be for the same customer.")
        return redirect("accounting:credit_memo_detail", credit_memo_id=credit_memo.id)

    if credit_memo.status == "VOID":
        messages.error(request, "Cannot apply a void credit memo.")
        return redirect("accounting:credit_memo_detail", credit_memo_id=credit_memo.id)

    if invoice.status == "VOID":
        messages.error(request, "Cannot apply credit to a void invoice.")
        return redirect("accounting:invoice_detail", invoice_id=invoice.id)

    # Get amount to apply (use remaining credit or invoice balance, whichever is smaller)
    amount_to_apply = min(credit_memo.amount_available, invoice.amount_due)

    if amount_to_apply <= Decimal("0.00"):
        messages.error(request, "No credit available to apply or invoice is fully paid.")
        return redirect("accounting:credit_memo_detail", credit_memo_id=credit_memo.id)

    try:
        with transaction.atomic():
            # Apply credit memo to invoice
            credit_memo.apply_to_invoice(invoice, amount_to_apply, request.user)

            messages.success(
                request,
                f"Applied ${amount_to_apply:,.2f} from credit memo {credit_memo.credit_memo_number} "
                f"to invoice {invoice.invoice_number}.",
            )

            # Audit logging
            from apps.core.audit_models import AuditLog

            AuditLog.objects.create(
                tenant=request.user.tenant,
                user=request.user,
                category=AuditLog.CATEGORY_DATA,
                action=AuditLog.ACTION_UPDATE,
                severity=AuditLog.SEVERITY_INFO,
                description=f"Applied credit memo {credit_memo.credit_memo_number} to invoice {invoice.invoice_number}",
                after_value=f"Amount: ${amount_to_apply:,.2f}",
                ip_address=request.META.get("REMOTE_ADDR"),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
                request_method=request.method,
                request_path=request.path,
            )

    except Exception as e:
        logger.error(f"Failed to apply credit memo: {str(e)}")
        messages.error(request, f"Failed to apply credit memo: {str(e)}")

    return redirect("accounting:invoice_detail", invoice_id=invoice.id)


# ============================================================================
# Bank Account Management Views (Task 4.3)
# ============================================================================


@login_required
@tenant_access_required
def bank_account_list(request):
    """
    Display list of bank accounts with filtering and status information.

    Shows current balance, reconciled balance, and unreconciled transactions
    for each account.

    Implements Requirements: 6.1, 6.3, 6.6, 6.7
    """
    from .bank_models import BankAccount

    # Get all bank accounts for tenant
    bank_accounts = (
        BankAccount.objects.filter(tenant=request.user.tenant)
        .select_related("created_by")
        .order_by("-is_default", "account_name")
    )

    # Apply filters
    status_filter = request.GET.get("status", "")
    account_type_filter = request.GET.get("account_type", "")

    if status_filter == "active":
        bank_accounts = bank_accounts.filter(is_active=True)
    elif status_filter == "inactive":
        bank_accounts = bank_accounts.filter(is_active=False)

    if account_type_filter:
        bank_accounts = bank_accounts.filter(account_type=account_type_filter)

    # Calculate totals
    total_balance = sum(account.current_balance for account in bank_accounts)
    total_reconciled = sum(account.reconciled_balance for account in bank_accounts)
    total_unreconciled = total_balance - total_reconciled

    # Get account types for filter dropdown
    account_types = BankAccount.ACCOUNT_TYPE_CHOICES

    # Audit logging
    from apps.core.audit_models import AuditLog

    AuditLog.objects.create(
        tenant=request.user.tenant,
        user=request.user,
        category=AuditLog.CATEGORY_DATA,
        action=AuditLog.ACTION_API_GET,
        severity=AuditLog.SEVERITY_INFO,
        description=f"Viewed bank accounts list (filters: status={status_filter}, type={account_type_filter})",
        ip_address=request.META.get("REMOTE_ADDR"),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        request_method=request.method,
        request_path=request.path,
    )

    context = {
        "bank_accounts": bank_accounts,
        "total_balance": total_balance,
        "total_reconciled": total_reconciled,
        "total_unreconciled": total_unreconciled,
        "status_filter": status_filter,
        "account_type_filter": account_type_filter,
        "account_types": account_types,
        "page_title": "Bank Accounts",
    }

    return render(request, "accounting/bank_accounts/list.html", context)


@login_required
@tenant_access_required
def bank_account_create(request):
    """
    Create a new bank account.

    Handles form submission and creates bank account with tenant isolation
    and audit logging.

    Implements Requirements: 6.1, 6.7
    """
    from .forms import BankAccountForm

    if request.method == "POST":
        form = BankAccountForm(request.POST, tenant=request.user.tenant, user=request.user)

        if form.is_valid():
            try:
                with transaction.atomic():
                    bank_account = form.save()

                    # Audit logging
                    from apps.core.audit_models import AuditLog

                    AuditLog.objects.create(
                        tenant=request.user.tenant,
                        user=request.user,
                        category=AuditLog.CATEGORY_DATA,
                        action=AuditLog.ACTION_CREATE,
                        severity=AuditLog.SEVERITY_INFO,
                        description=f"Created bank account: {bank_account.account_name}",
                        new_values={
                            "account_name": bank_account.account_name,
                            "account_number": bank_account.masked_account_number,
                            "bank_name": bank_account.bank_name,
                            "opening_balance": str(bank_account.opening_balance),
                            "current_balance": str(bank_account.current_balance),
                        },
                        ip_address=request.META.get("REMOTE_ADDR"),
                        user_agent=request.META.get("HTTP_USER_AGENT", ""),
                        request_method=request.method,
                        request_path=request.path,
                    )

                    messages.success(
                        request,
                        f"Bank account '{bank_account.account_name}' created successfully.",
                    )
                    return redirect("accounting:bank_account_detail", account_id=bank_account.id)

            except Exception as e:
                logger.error(f"Failed to create bank account: {str(e)}")
                messages.error(request, f"Failed to create bank account: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = BankAccountForm(tenant=request.user.tenant, user=request.user)

    context = {
        "form": form,
        "page_title": "Create Bank Account",
    }

    return render(request, "accounting/bank_accounts/form.html", context)


@login_required
@tenant_access_required
def bank_account_detail(request, account_id):
    """
    Display bank account details with transaction history.

    Shows current balance, reconciled balance, unreconciled transactions,
    and recent transaction history with running balance.

    Implements Requirements: 6.1, 6.3, 6.6, 6.7
    """
    from .bank_models import BankAccount, BankTransaction

    # Get bank account with tenant filtering
    bank_account = get_object_or_404(
        BankAccount.objects.select_related("created_by"),
        id=account_id,
        tenant=request.user.tenant,
    )

    # Get date range from request or default to current month
    end_date = date.today()
    start_date = date(end_date.year, end_date.month, 1)

    if request.GET.get("start_date"):
        start_date = datetime.strptime(request.GET["start_date"], "%Y-%m-%d").date()
    if request.GET.get("end_date"):
        end_date = datetime.strptime(request.GET["end_date"], "%Y-%m-%d").date()

    # Get transactions for this account
    transactions = (
        BankTransaction.objects.filter(
            bank_account=bank_account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
        )
        .select_related("reconciled_by", "created_by")
        .order_by("-transaction_date", "-created_at")
    )

    # Apply reconciliation filter
    reconciliation_filter = request.GET.get("reconciliation", "")
    if reconciliation_filter == "reconciled":
        transactions = transactions.filter(is_reconciled=True)
    elif reconciliation_filter == "unreconciled":
        transactions = transactions.filter(is_reconciled=False)

    # Calculate running balance
    transactions_with_balance = []
    running_balance = bank_account.current_balance

    for txn in transactions:
        transactions_with_balance.append(
            {
                "transaction": txn,
                "balance": running_balance,
            }
        )
        # Adjust running balance backwards (since we're going newest to oldest)
        running_balance -= txn.signed_amount

    # Get reconciliation statistics
    unreconciled_count = BankTransaction.objects.filter(
        bank_account=bank_account,
        is_reconciled=False,
    ).count()

    unreconciled_amount = sum(
        txn.signed_amount
        for txn in BankTransaction.objects.filter(
            bank_account=bank_account,
            is_reconciled=False,
        )
    )

    # Get recent reconciliations
    from .bank_models import BankReconciliation

    recent_reconciliations = (
        BankReconciliation.objects.filter(bank_account=bank_account)
        .select_related("created_by", "completed_by")
        .order_by("-reconciliation_date")[:5]
    )

    # Audit logging
    from apps.core.audit_models import AuditLog

    AuditLog.objects.create(
        tenant=request.user.tenant,
        user=request.user,
        category=AuditLog.CATEGORY_DATA,
        action=AuditLog.ACTION_API_GET,
        severity=AuditLog.SEVERITY_INFO,
        description=f"Viewed bank account detail: {bank_account.account_name}",
        ip_address=request.META.get("REMOTE_ADDR"),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        request_method=request.method,
        request_path=request.path,
    )

    context = {
        "bank_account": bank_account,
        "transactions_with_balance": transactions_with_balance,
        "unreconciled_count": unreconciled_count,
        "unreconciled_amount": unreconciled_amount,
        "recent_reconciliations": recent_reconciliations,
        "start_date": start_date,
        "end_date": end_date,
        "reconciliation_filter": reconciliation_filter,
        "page_title": f"Bank Account: {bank_account.account_name}",
    }

    return render(request, "accounting/bank_accounts/detail.html", context)


@login_required
@tenant_access_required
def bank_account_edit(request, account_id):
    """
    Edit an existing bank account.

    Allows updating account details while maintaining audit trail.

    Implements Requirements: 6.1, 6.7
    """
    from .bank_models import BankAccount
    from .forms import BankAccountForm

    # Get bank account with tenant filtering
    bank_account = get_object_or_404(
        BankAccount,
        id=account_id,
        tenant=request.user.tenant,
    )

    if request.method == "POST":
        form = BankAccountForm(
            request.POST,
            instance=bank_account,
            tenant=request.user.tenant,
            user=request.user,
        )

        if form.is_valid():
            try:
                with transaction.atomic():
                    # Capture before values for audit
                    before_values = {
                        "account_name": bank_account.account_name,
                        "account_number": bank_account.masked_account_number,
                        "bank_name": bank_account.bank_name,
                        "account_type": bank_account.get_account_type_display(),
                        "is_active": bank_account.is_active,
                        "is_default": bank_account.is_default,
                    }

                    bank_account = form.save()

                    # Capture after values for audit
                    after_values = {
                        "account_name": bank_account.account_name,
                        "account_number": bank_account.masked_account_number,
                        "bank_name": bank_account.bank_name,
                        "account_type": bank_account.get_account_type_display(),
                        "is_active": bank_account.is_active,
                        "is_default": bank_account.is_default,
                    }

                    # Audit logging
                    from apps.core.audit_models import AuditLog

                    AuditLog.objects.create(
                        tenant=request.user.tenant,
                        user=request.user,
                        category=AuditLog.CATEGORY_DATA,
                        action=AuditLog.ACTION_UPDATE,
                        severity=AuditLog.SEVERITY_INFO,
                        description=f"Updated bank account: {bank_account.account_name}",
                        before_value=str(before_values),
                        after_value=str(after_values),
                        ip_address=request.META.get("REMOTE_ADDR"),
                        user_agent=request.META.get("HTTP_USER_AGENT", ""),
                        request_method=request.method,
                        request_path=request.path,
                    )

                    messages.success(
                        request,
                        f"Bank account '{bank_account.account_name}' updated successfully.",
                    )
                    return redirect("accounting:bank_account_detail", account_id=bank_account.id)

            except Exception as e:
                logger.error(f"Failed to update bank account: {str(e)}")
                messages.error(request, f"Failed to update bank account: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = BankAccountForm(
            instance=bank_account,
            tenant=request.user.tenant,
            user=request.user,
        )

    context = {
        "form": form,
        "bank_account": bank_account,
        "page_title": f"Edit Bank Account: {bank_account.account_name}",
    }

    return render(request, "accounting/bank_accounts/form.html", context)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def bank_account_deactivate(request, account_id):
    """
    Deactivate a bank account.

    Prevents new transactions but preserves historical data.

    Implements Requirements: 6.7
    """
    from .bank_models import BankAccount

    # Get bank account with tenant filtering
    bank_account = get_object_or_404(
        BankAccount,
        id=account_id,
        tenant=request.user.tenant,
    )

    try:
        with transaction.atomic():
            bank_account.deactivate()

            # Audit logging
            from apps.core.audit_models import AuditLog

            AuditLog.objects.create(
                tenant=request.user.tenant,
                user=request.user,
                category=AuditLog.CATEGORY_DATA,
                action=AuditLog.ACTION_UPDATE,
                severity=AuditLog.SEVERITY_WARNING,
                description=f"Deactivated bank account: {bank_account.account_name}",
                ip_address=request.META.get("REMOTE_ADDR"),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
                request_method=request.method,
                request_path=request.path,
            )

            messages.success(
                request,
                f"Bank account '{bank_account.account_name}' has been deactivated.",
            )

    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        logger.error(f"Failed to deactivate bank account: {str(e)}")
        messages.error(request, f"Failed to deactivate bank account: {str(e)}")

    return redirect("accounting:bank_account_list")


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def bank_account_set_default(request, account_id):
    """
    Set a bank account as the default account for transactions.

    Implements Requirements: 6.1, 6.7
    """
    from .bank_models import BankAccount

    # Get bank account with tenant filtering
    bank_account = get_object_or_404(
        BankAccount,
        id=account_id,
        tenant=request.user.tenant,
    )

    try:
        with transaction.atomic():
            bank_account.set_as_default()

            # Audit logging
            from apps.core.audit_models import AuditLog

            AuditLog.objects.create(
                tenant=request.user.tenant,
                user=request.user,
                category=AuditLog.CATEGORY_DATA,
                action=AuditLog.ACTION_UPDATE,
                severity=AuditLog.SEVERITY_INFO,
                description=f"Set bank account as default: {bank_account.account_name}",
                ip_address=request.META.get("REMOTE_ADDR"),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
                request_method=request.method,
                request_path=request.path,
            )

            messages.success(
                request,
                f"Bank account '{bank_account.account_name}' is now the default account.",
            )

    except Exception as e:
        logger.error(f"Failed to set default bank account: {str(e)}")
        messages.error(request, f"Failed to set default bank account: {str(e)}")

    return redirect("accounting:bank_account_detail", account_id=bank_account.id)


# ============================================================================
# Bank Reconciliation Views
# ============================================================================


@login_required
@tenant_access_required
def bank_reconciliation_start(request):
    """
    Start a new bank reconciliation session.

    GET: Display form to select bank account and enter statement details
    POST: Create reconciliation and redirect to detail view

    Implements Requirements: 4.1, 4.7
    """
    from .bank_models import BankAccount
    from .forms import BankReconciliationStartForm
    from .services import BankReconciliationService

    if request.method == "POST":
        form = BankReconciliationStartForm(request.POST, tenant=request.user.tenant)

        if form.is_valid():
            try:
                with transaction.atomic():
                    bank_account = form.cleaned_data["bank_account"]
                    statement_date = form.cleaned_data["statement_date"]
                    ending_balance = form.cleaned_data["ending_balance"]
                    beginning_balance = form.cleaned_data.get("beginning_balance")

                    # Start reconciliation
                    reconciliation = BankReconciliationService.start_reconciliation(
                        bank_account=bank_account,
                        statement_date=statement_date,
                        ending_balance=ending_balance,
                        user=request.user,
                        beginning_balance=beginning_balance,
                    )

                    messages.success(
                        request, f"Bank reconciliation started for {bank_account.account_name}"
                    )

                    return redirect("accounting:bank_reconciliation_detail", pk=reconciliation.id)

            except Exception as e:
                logger.error(f"Failed to start reconciliation: {str(e)}")
                messages.error(request, f"Failed to start reconciliation: {str(e)}")
    else:
        form = BankReconciliationStartForm(tenant=request.user.tenant)

    # Get bank accounts for context
    bank_accounts = BankAccount.objects.filter(tenant=request.user.tenant, is_active=True)

    context = {
        "form": form,
        "bank_accounts": bank_accounts,
        "page_title": "Start Bank Reconciliation",
    }

    return render(request, "accounting/bank_reconciliation/start.html", context)


@login_required
@tenant_access_required
def bank_reconciliation_detail(request, pk):
    """
    Display bank reconciliation interface with transactions to reconcile.

    GET: Show reconciliation interface
    POST: Handle marking transactions as reconciled

    Implements Requirements: 4.1, 4.2, 4.4, 4.7
    """
    from .bank_models import BankReconciliation, BankTransaction
    from .services import BankReconciliationService

    # Get reconciliation with tenant filtering
    reconciliation = get_object_or_404(BankReconciliation, id=pk, tenant=request.user.tenant)

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "mark_reconciled":
            # Get selected transaction IDs
            transaction_ids = request.POST.getlist("transaction_ids")

            if transaction_ids:
                try:
                    count = BankReconciliationService.mark_reconciled(
                        transaction_ids=transaction_ids,
                        reconciliation=reconciliation,
                        user=request.user,
                    )

                    messages.success(request, f"Marked {count} transaction(s) as reconciled")

                except Exception as e:
                    logger.error(f"Failed to mark transactions as reconciled: {str(e)}")
                    messages.error(request, f"Failed to mark transactions: {str(e)}")
            else:
                messages.warning(request, "No transactions selected")

        elif action == "unreconcile":
            transaction_id = request.POST.get("transaction_id")
            reason = request.POST.get("reason", "User requested unreconcile")

            if transaction_id:
                try:
                    txn = BankTransaction.objects.get(id=transaction_id, tenant=request.user.tenant)

                    BankReconciliationService.unreconcile_transaction(
                        transaction=txn, reason=reason, user=request.user
                    )

                    messages.success(request, "Transaction unreconciled")

                except Exception as e:
                    logger.error(f"Failed to unreconcile transaction: {str(e)}")
                    messages.error(request, f"Failed to unreconcile: {str(e)}")

        return redirect("accounting:bank_reconciliation_detail", pk=pk)

    # Get unreconciled transactions
    unreconciled_transactions = BankTransaction.objects.filter(
        bank_account=reconciliation.bank_account,
        is_reconciled=False,
        transaction_date__lte=reconciliation.reconciliation_date,
    ).order_by("transaction_date")

    # Get reconciled transactions for this reconciliation
    reconciled_transactions = BankTransaction.objects.filter(
        reconciliation=reconciliation, is_reconciled=True
    ).order_by("transaction_date")

    # Calculate summary
    unreconciled_deposits = sum(
        txn.amount for txn in unreconciled_transactions if txn.transaction_type == "CREDIT"
    )
    unreconciled_withdrawals = sum(
        txn.amount for txn in unreconciled_transactions if txn.transaction_type == "DEBIT"
    )

    context = {
        "reconciliation": reconciliation,
        "unreconciled_transactions": unreconciled_transactions,
        "reconciled_transactions": reconciled_transactions,
        "unreconciled_deposits": unreconciled_deposits,
        "unreconciled_withdrawals": unreconciled_withdrawals,
        "page_title": f"Bank Reconciliation - {reconciliation.bank_account.account_name}",
    }

    return render(request, "accounting/bank_reconciliation/detail.html", context)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def bank_reconciliation_complete(request, pk):
    """
    Complete a bank reconciliation.

    Implements Requirements: 4.4, 4.7
    """
    from .bank_models import BankReconciliation
    from .services import BankReconciliationService

    # Get reconciliation with tenant filtering
    reconciliation = get_object_or_404(BankReconciliation, id=pk, tenant=request.user.tenant)

    try:
        BankReconciliationService.complete_reconciliation(
            reconciliation=reconciliation, user=request.user
        )

        if reconciliation.is_balanced:
            messages.success(request, "Reconciliation completed successfully. Account is balanced.")
        else:
            messages.warning(
                request,
                f"Reconciliation completed with variance of {reconciliation.variance}. "
                f"Please review and create adjusting entries if needed.",
            )

        return redirect("accounting:bank_reconciliation_report", pk=pk)

    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        logger.error(f"Failed to complete reconciliation: {str(e)}")
        messages.error(request, f"Failed to complete reconciliation: {str(e)}")

    return redirect("accounting:bank_reconciliation_detail", pk=pk)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def bank_reconciliation_cancel(request, pk):
    """
    Cancel a bank reconciliation.

    Implements Requirements: 4.7, 4.8
    """
    from .bank_models import BankReconciliation
    from .services import BankReconciliationService

    # Get reconciliation with tenant filtering
    reconciliation = get_object_or_404(BankReconciliation, id=pk, tenant=request.user.tenant)

    reason = request.POST.get("reason", "User cancelled reconciliation")

    try:
        BankReconciliationService.cancel_reconciliation(
            reconciliation=reconciliation, reason=reason, user=request.user
        )

        messages.success(request, "Reconciliation cancelled")
        return redirect("accounting:bank_account_detail", account_id=reconciliation.bank_account.id)

    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        logger.error(f"Failed to cancel reconciliation: {str(e)}")
        messages.error(request, f"Failed to cancel reconciliation: {str(e)}")

    return redirect("accounting:bank_reconciliation_detail", pk=pk)


@login_required
@tenant_access_required
def bank_reconciliation_report(request, pk):
    """
    Display completed bank reconciliation report.

    Implements Requirements: 4.4, 4.6
    """
    from .bank_models import BankReconciliation, BankTransaction

    # Get reconciliation with tenant filtering
    reconciliation = get_object_or_404(BankReconciliation, id=pk, tenant=request.user.tenant)

    # Get all reconciled transactions
    reconciled_transactions = BankTransaction.objects.filter(
        reconciliation=reconciliation, is_reconciled=True
    ).order_by("transaction_date")

    # Separate deposits and withdrawals
    deposits = [txn for txn in reconciled_transactions if txn.transaction_type == "CREDIT"]
    withdrawals = [txn for txn in reconciled_transactions if txn.transaction_type == "DEBIT"]

    context = {
        "reconciliation": reconciliation,
        "deposits": deposits,
        "withdrawals": withdrawals,
        "page_title": f"Reconciliation Report - {reconciliation.bank_account.account_name}",
    }

    return render(request, "accounting/bank_reconciliation/report.html", context)


@login_required
@tenant_access_required
def bank_reconciliation_list(request):
    """
    Display list of all bank reconciliations.

    Implements Requirements: 4.6, 4.7
    """
    from .bank_models import BankAccount, BankReconciliation

    # Get filter parameters
    bank_account_id = request.GET.get("bank_account")
    status = request.GET.get("status")

    # Base queryset with tenant filtering
    reconciliations = BankReconciliation.objects.filter(tenant=request.user.tenant).select_related(
        "bank_account", "created_by", "completed_by"
    )

    # Apply filters
    if bank_account_id:
        reconciliations = reconciliations.filter(bank_account_id=bank_account_id)

    if status:
        reconciliations = reconciliations.filter(status=status)

    # Order by date (most recent first)
    reconciliations = reconciliations.order_by("-reconciliation_date", "-created_at")

    # Get bank accounts for filter dropdown
    bank_accounts = BankAccount.objects.filter(tenant=request.user.tenant, is_active=True)

    context = {
        "reconciliations": reconciliations,
        "bank_accounts": bank_accounts,
        "selected_bank_account": bank_account_id,
        "selected_status": status,
        "page_title": "Bank Reconciliations",
    }

    return render(request, "accounting/bank_reconciliation/list.html", context)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def bank_reconciliation_create_adjustment(request, pk):
    """
    Create an adjusting journal entry during reconciliation.

    Implements Requirements: 4.5, 4.7
    """
    from .bank_models import BankReconciliation
    from .services import BankReconciliationService

    # Get reconciliation with tenant filtering
    reconciliation = get_object_or_404(BankReconciliation, id=pk, tenant=request.user.tenant)

    # Get form data
    description = request.POST.get("description")
    amount = request.POST.get("amount")
    account_code = request.POST.get("account_code")
    is_debit = request.POST.get("is_debit") == "true"

    if not all([description, amount, account_code]):
        messages.error(request, "All fields are required")
        return redirect("accounting:bank_reconciliation_detail", pk=pk)

    try:
        amount = Decimal(amount)

        BankReconciliationService.create_adjusting_entry(
            reconciliation=reconciliation,
            description=description,
            amount=amount,
            account_code=account_code,
            is_debit=is_debit,
            user=request.user,
        )

        messages.success(request, f"Adjusting entry created: {description} for {amount}")

    except ValueError:
        messages.error(request, "Invalid amount")
    except Exception as e:
        logger.error(f"Failed to create adjusting entry: {str(e)}")
        messages.error(request, f"Failed to create adjusting entry: {str(e)}")

    return redirect("accounting:bank_reconciliation_detail", pk=pk)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def bank_reconciliation_auto_match(request, pk):
    """
    Automatically match bank transactions with journal entries.

    Implements Requirements: 4.3
    """
    from .bank_models import BankReconciliation
    from .services import BankReconciliationService

    # Get reconciliation with tenant filtering
    reconciliation = get_object_or_404(BankReconciliation, id=pk, tenant=request.user.tenant)

    try:
        result = BankReconciliationService.auto_match_transactions(reconciliation)

        matched_count = result["matched_count"]
        suggestions_count = len(result["suggestions"])

        if matched_count > 0:
            messages.success(request, f"Automatically matched {matched_count} transaction(s)")

        if suggestions_count > 0:
            messages.info(request, f"Found {suggestions_count} potential matches for manual review")

        if matched_count == 0 and suggestions_count == 0:
            messages.info(request, "No automatic matches found")

    except Exception as e:
        logger.error(f"Failed to auto-match transactions: {str(e)}")
        messages.error(request, f"Failed to auto-match transactions: {str(e)}")

    return redirect("accounting:bank_reconciliation_detail", pk=pk)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def bank_transaction_toggle_reconcile(request, transaction_id):
    """
    Toggle reconciliation status of a bank transaction (HTMX endpoint).

    This view is called via HTMX when a user checks/unchecks a transaction
    in the bank reconciliation interface.

    Implements Requirements: 4.2, 4.7
    """
    from django.http import HttpResponse

    from .bank_models import BankReconciliation, BankTransaction
    from .services import BankReconciliationService

    # Get transaction with tenant filtering
    transaction = get_object_or_404(BankTransaction, id=transaction_id, tenant=request.user.tenant)

    # Get the active reconciliation for this bank account
    reconciliation = BankReconciliation.objects.filter(
        bank_account=transaction.bank_account, tenant=request.user.tenant, status="IN_PROGRESS"
    ).first()

    if not reconciliation:
        return HttpResponse(status=400)

    try:
        if transaction.is_reconciled:
            # Unreconcile the transaction
            BankReconciliationService.unreconcile_transaction(
                transaction=transaction,
                reason="User toggled in reconciliation interface",
                user=request.user,
            )
        else:
            # Mark as reconciled
            BankReconciliationService.mark_reconciled(
                transaction_ids=[str(transaction.id)],
                reconciliation=reconciliation,
                user=request.user,
            )

        # Return success (HTMX will handle the UI update via JavaScript)
        return HttpResponse(status=200)

    except Exception as e:
        logger.error(f"Failed to toggle transaction reconciliation: {str(e)}")
        return HttpResponse(status=500)


@login_required
@tenant_access_required
@require_http_methods(["POST"])
def bank_transaction_create_adjustment(request, account_id):
    """
    Create an adjustment transaction during bank reconciliation.

    Allows users to add missing transactions or adjustments when reconciling.
    Requirement: 4.5
    """
    from django.contrib.contenttypes.models import ContentType

    from apps.core.audit_models import AuditLog

    from .bank_models import BankAccount, BankTransaction

    try:
        bank_account = get_object_or_404(BankAccount, id=account_id, tenant=request.user.tenant)

        # Get form data
        description = request.POST.get("description", "").strip()
        transaction_date = request.POST.get("transaction_date")
        transaction_type = request.POST.get("transaction_type")  # DEBIT or CREDIT
        amount = request.POST.get("amount")
        reference_number = request.POST.get("reference_number", "").strip()

        # Validate
        if not all([description, transaction_date, transaction_type, amount]):
            messages.error(request, "All fields are required.")
            return redirect(f"{reverse('accounting:bank_reconciliation')}?account={account_id}")

        try:
            amount = Decimal(amount)
            if amount <= 0:
                raise ValueError("Amount must be positive")
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f"Invalid amount: {str(e)}")
            return redirect(f"{reverse('accounting:bank_reconciliation')}?account={account_id}")

        try:
            transaction_date = datetime.strptime(transaction_date, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect(f"{reverse('accounting:bank_reconciliation')}?account={account_id}")

        if transaction_type not in ["DEBIT", "CREDIT"]:
            messages.error(request, "Invalid transaction type.")
            return redirect(f"{reverse('accounting:bank_reconciliation')}?account={account_id}")

        # Create the adjustment transaction
        with transaction.atomic():
            bank_transaction = BankTransaction.objects.create(
                tenant=request.user.tenant,
                bank_account=bank_account,
                transaction_date=transaction_date,
                description=f"[ADJUSTMENT] {description}",
                amount=amount,
                transaction_type=transaction_type,
                reference_number=reference_number
                or f"ADJ-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                is_reconciled=False,
                created_by=request.user,
            )

            # Audit log
            content_type = ContentType.objects.get_for_model(BankTransaction)
            AuditLog.objects.create(
                tenant=request.user.tenant,
                user=request.user,
                category="DATA",
                action="CREATE",
                severity="INFO",
                description=f"Created adjustment transaction: {description}",
                content_type=content_type,
                object_id=str(bank_transaction.id),
                new_values={
                    "description": description,
                    "amount": str(amount),
                    "type": transaction_type,
                    "date": str(transaction_date),
                },
                ip_address=request.META.get("REMOTE_ADDR"),
                user_agent=request.META.get("HTTP_USER_AGENT", "")[:200],
            )

            messages.success(request, f"Adjustment transaction created: {description}")
            return redirect(f"{reverse('accounting:bank_reconciliation')}?account={account_id}")

    except Exception as e:
        logger.error(f"Error creating adjustment transaction: {str(e)}")
        messages.error(request, f"Error creating adjustment: {str(e)}")
        return redirect("accounting:bank_reconciliation")
