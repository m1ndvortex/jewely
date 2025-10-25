"""
Data management services for export, import, and backup operations.

Implements Requirement 20: Settings and Configuration
- Data export functionality (CSV/Excel)
- Data import with validation
- Backup trigger interface
"""

import csv
import json
import logging
import os
import tempfile
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Optional

from django.apps import apps
from django.db import transaction
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill

from apps.core.data_models import BackupTrigger, DataActivity
from apps.core.models import Tenant

logger = logging.getLogger(__name__)


class DataExportService:
    """
    Service for exporting tenant data to various formats.
    """

    def __init__(self, tenant: Tenant):
        self.tenant = tenant

    def export_data(
        self,
        data_types: List[str],
        format: str,
        user=None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> DataActivity:
        """
        Export tenant data to specified format.

        Args:
            data_types: List of data types to export
            format: Export format (csv, excel, json)
            user: User initiating the export
            date_from: Start date for filtering (for sales data)
            date_to: End date for filtering (for sales data)

        Returns:
            DataActivity: Created activity record
        """
        # Create activity record
        activity = DataActivity.objects.create(
            tenant=self.tenant,
            activity_type="EXPORT",
            data_type=",".join(data_types) if len(data_types) > 1 else data_types[0],
            format=format,
            initiated_by=user,
            parameters={
                "data_types": data_types,
                "date_from": date_from.isoformat() if date_from else None,
                "date_to": date_to.isoformat() if date_to else None,
            },
        )

        try:
            activity.mark_started()

            # Collect data for export
            export_data = {}
            total_records = 0

            for data_type in data_types:
                data = self._get_data_for_export(data_type, date_from, date_to)
                export_data[data_type] = data
                total_records += len(data)

            # Generate export file
            if format == "csv":
                file_path = self._export_to_csv(export_data, data_types)
            elif format == "excel":
                file_path = self._export_to_excel(export_data, data_types)
            elif format == "json":
                file_path = self._export_to_json(export_data, data_types)
            else:
                raise ValueError(f"Unsupported export format: {format}")

            # Update activity with results
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            activity.file_path = file_path
            activity.file_name = os.path.basename(file_path)
            activity.file_size = file_size
            activity.mark_completed(
                records_processed=total_records, records_successful=total_records
            )

            logger.info(f"Export completed: {activity.file_name} ({total_records} records)")

        except Exception as e:
            logger.exception(f"Export failed for tenant {self.tenant.id}: {e}")
            activity.mark_failed(str(e))
            raise

        return activity

    def _get_data_for_export(
        self, data_type: str, date_from: Optional[datetime], date_to: Optional[datetime]
    ) -> List[Dict]:
        """Get data for specific data type."""
        if data_type == "inventory":
            return self._get_inventory_data()
        elif data_type == "customers":
            return self._get_customers_data()
        elif data_type == "sales":
            return self._get_sales_data(date_from, date_to)
        elif data_type == "suppliers":
            return self._get_suppliers_data()
        elif data_type == "settings":
            return self._get_settings_data()
        else:
            raise ValueError(f"Unsupported data type: {data_type}")

    def _get_inventory_data(self) -> List[Dict]:
        """Get inventory data for export."""
        InventoryItem = apps.get_model("inventory", "InventoryItem")

        items = InventoryItem.objects.filter(tenant=self.tenant).select_related(
            "category", "branch"
        )

        data = []
        for item in items:
            data.append(
                {
                    "SKU": item.sku,
                    "Name": item.name,
                    "Description": getattr(item, "description", ""),
                    "Category": item.category.name if item.category else "",
                    "Karat": item.karat,
                    "Weight (grams)": float(item.weight_grams),
                    "Cost Price": float(item.cost_price),
                    "Selling Price": float(item.selling_price),
                    "Quantity": item.quantity,
                    "Branch": item.branch.name if item.branch else "",
                    "Serial Number": item.serial_number or "",
                    "Lot Number": item.lot_number or "",
                    "Barcode": item.barcode or "",
                    "Is Active": item.is_active,
                    "Created At": item.created_at.isoformat(),
                    "Updated At": item.updated_at.isoformat(),
                }
            )

        return data

    def _get_customers_data(self) -> List[Dict]:
        """Get customers data for export."""
        Customer = apps.get_model("crm", "Customer")

        customers = Customer.objects.filter(tenant=self.tenant)

        data = []
        for customer in customers:
            data.append(
                {
                    "Customer Number": customer.customer_number,
                    "First Name": customer.first_name,
                    "Last Name": customer.last_name,
                    "Email": customer.email or "",
                    "Phone": customer.phone,
                    "Address": getattr(customer, "address", ""),
                    "City": getattr(customer, "city", ""),
                    "State": getattr(customer, "state", ""),
                    "Postal Code": getattr(customer, "postal_code", ""),
                    "Country": getattr(customer, "country", ""),
                    "Date of Birth": (
                        customer.date_of_birth.isoformat()
                        if getattr(customer, "date_of_birth", None)
                        else ""
                    ),
                    "Loyalty Tier": customer.loyalty_tier,
                    "Loyalty Points": customer.loyalty_points,
                    "Store Credit": float(customer.store_credit),
                    "Total Purchases": float(customer.total_purchases),
                    "Notes": getattr(customer, "notes", ""),
                    "Created At": customer.created_at.isoformat(),
                }
            )

        return data

    def _get_sales_data(
        self, date_from: Optional[datetime], date_to: Optional[datetime]
    ) -> List[Dict]:
        """Get sales data for export."""
        Sale = apps.get_model("sales", "Sale")

        sales_qs = (
            Sale.objects.filter(tenant=self.tenant)
            .select_related("customer", "branch", "terminal", "employee")
            .prefetch_related("items__inventory_item")
        )

        if date_from:
            sales_qs = sales_qs.filter(created_at__gte=date_from)
        if date_to:
            sales_qs = sales_qs.filter(created_at__lte=date_to)

        data = []
        for sale in sales_qs:
            # Add sale header
            sale_data = {
                "Sale Number": sale.sale_number,
                "Customer": (
                    f"{sale.customer.first_name} {sale.customer.last_name}" if sale.customer else ""
                ),
                "Customer Number": sale.customer.customer_number if sale.customer else "",
                "Branch": sale.branch.name,
                "Terminal": sale.terminal.terminal_id if sale.terminal else "",
                "Employee": f"{sale.employee.first_name} {sale.employee.last_name}",
                "Subtotal": float(sale.subtotal),
                "Tax": float(sale.tax),
                "Discount": float(sale.discount),
                "Total": float(sale.total),
                "Payment Method": sale.payment_method,
                "Status": sale.status,
                "Created At": sale.created_at.isoformat(),
            }

            # Add line items
            for item in sale.items.all():
                item_data = sale_data.copy()
                item_data.update(
                    {
                        "Item SKU": item.inventory_item.sku,
                        "Item Name": item.inventory_item.name,
                        "Item Quantity": item.quantity,
                        "Item Unit Price": float(item.unit_price),
                        "Item Subtotal": float(item.subtotal),
                    }
                )
                data.append(item_data)

        return data

    def _get_suppliers_data(self) -> List[Dict]:
        """Get suppliers data for export."""
        try:
            Supplier = apps.get_model("procurement", "Supplier")

            suppliers = Supplier.objects.filter(tenant=self.tenant)

            data = []
            for supplier in suppliers:
                data.append(
                    {
                        "Company Name": supplier.name,
                        "Contact Person": supplier.contact_person,
                        "Email": supplier.email,
                        "Phone": supplier.phone,
                        "Address": getattr(supplier, "address", ""),
                        "City": getattr(supplier, "city", ""),
                        "State": getattr(supplier, "state", ""),
                        "Postal Code": getattr(supplier, "postal_code", ""),
                        "Country": getattr(supplier, "country", ""),
                        "Website": getattr(supplier, "website", ""),
                        "Rating": supplier.rating,
                        "Notes": getattr(supplier, "notes", ""),
                        "Created At": supplier.created_at.isoformat(),
                    }
                )

            return data
        except LookupError:
            # Supplier model doesn't exist yet
            return []

    def _get_settings_data(self) -> List[Dict]:
        """Get settings data for export."""
        from apps.core.models import TenantSettings

        data = []

        # Tenant settings
        try:
            tenant_settings = TenantSettings.objects.get(tenant=self.tenant)
            data.append(
                {
                    "Setting Type": "Tenant Settings",
                    "Company Name": tenant_settings.company_name,
                    "Contact Email": tenant_settings.contact_email or "",
                    "Contact Phone": tenant_settings.contact_phone or "",
                    "Address": tenant_settings.address or "",
                    "City": tenant_settings.city or "",
                    "State": tenant_settings.state or "",
                    "Postal Code": tenant_settings.postal_code or "",
                    "Country": tenant_settings.country or "",
                    "Currency": tenant_settings.currency,
                    "Timezone": tenant_settings.timezone,
                    "Business Hours": json.dumps(tenant_settings.business_hours),
                    "Holidays": json.dumps(tenant_settings.holidays),
                }
            )
        except TenantSettings.DoesNotExist:
            pass

        return data

    def _export_to_csv(self, export_data: Dict[str, List[Dict]], data_types: List[str]) -> str:
        """Export data to CSV format."""
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{self.tenant.slug}_{timestamp}.csv"
        filepath = os.path.join(tempfile.gettempdir(), filename)

        with open(filepath, "w", newline="", encoding="utf-8") as csvfile:
            if len(data_types) == 1:
                # Single data type - simple CSV
                data = export_data[data_types[0]]
                if data:
                    fieldnames = data[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(data)
            else:
                # Multiple data types - add data type column
                writer = None
                for data_type, data in export_data.items():
                    if data:
                        # Add data type to each row
                        for row in data:
                            row["Data Type"] = data_type.title()

                        if writer is None:
                            fieldnames = ["Data Type"] + list(data[0].keys())
                            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                            writer.writeheader()

                        writer.writerows(data)

        return filepath

    def _export_to_excel(self, export_data: Dict[str, List[Dict]], data_types: List[str]) -> str:
        """Export data to Excel format."""
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{self.tenant.slug}_{timestamp}.xlsx"
        filepath = os.path.join(tempfile.gettempdir(), filename)

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        for data_type, data in export_data.items():
            if data:
                self._create_excel_worksheet(workbook, data_type, data)

        # If no sheets were created, add an empty one
        if not workbook.worksheets:
            workbook.create_sheet(title="No Data")

        workbook.save(filepath)
        return filepath

    def _create_excel_worksheet(self, workbook, data_type: str, data: List[Dict]):
        """Create a worksheet for a specific data type."""
        worksheet = workbook.create_sheet(title=data_type.title())
        headers = list(data[0].keys())

        # Add headers with formatting
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Add data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        self._adjust_excel_column_widths(worksheet)

    def _adjust_excel_column_widths(self, worksheet):
        """Auto-adjust column widths in Excel worksheet."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _export_to_json(self, export_data: Dict[str, List[Dict]], data_types: List[str]) -> str:
        """Export data to JSON format."""
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{self.tenant.slug}_{timestamp}.json"
        filepath = os.path.join(tempfile.gettempdir(), filename)

        # Prepare export structure
        export_structure = {
            "tenant": {
                "id": str(self.tenant.id),
                "company_name": self.tenant.company_name,
                "slug": self.tenant.slug,
            },
            "export_date": timezone.now().isoformat(),
            "data_types": data_types,
            "data": export_data,
        }

        with open(filepath, "w", encoding="utf-8") as jsonfile:
            json.dump(export_structure, jsonfile, indent=2, ensure_ascii=False)

        return filepath


class DataImportService:
    """
    Service for importing data from various formats.
    """

    def __init__(self, tenant: Tenant):
        self.tenant = tenant

    def import_data(
        self,
        data_type: str,
        file_path: str,
        user=None,
        update_existing: bool = False,
        validate_only: bool = False,
    ) -> DataActivity:
        """
        Import data from file.

        Args:
            data_type: Type of data to import
            file_path: Path to import file
            user: User initiating the import
            update_existing: Whether to update existing records
            validate_only: Only validate, don't import

        Returns:
            DataActivity: Created activity record
        """
        # Create activity record
        activity = DataActivity.objects.create(
            tenant=self.tenant,
            activity_type="IMPORT",
            data_type=data_type,
            file_name=os.path.basename(file_path),
            file_path=file_path,
            file_size=os.path.getsize(file_path) if os.path.exists(file_path) else 0,
            initiated_by=user,
            parameters={
                "update_existing": update_existing,
                "validate_only": validate_only,
            },
        )

        try:
            activity.mark_started()

            # Parse file
            data = self._parse_import_file(file_path)

            # Validate data
            validation_errors = self._validate_import_data(data_type, data)

            if validation_errors:
                activity.mark_failed("Validation failed", {"validation_errors": validation_errors})
                return activity

            if validate_only:
                activity.mark_completed(
                    records_processed=len(data), records_successful=len(data), records_failed=0
                )
                return activity

            # Import data
            results = self._import_data_records(data_type, data, update_existing)

            activity.mark_completed(
                records_processed=results["total"],
                records_successful=results["successful"],
                records_failed=results["failed"],
            )

            if results["errors"]:
                activity.error_details = {"import_errors": results["errors"]}
                activity.save(update_fields=["error_details"])

            logger.info(f"Import completed: {results['successful']}/{results['total']} records")

        except Exception as e:
            logger.exception(f"Import failed for tenant {self.tenant.id}: {e}")
            activity.mark_failed(str(e))
            raise

        return activity

    def _parse_import_file(self, file_path: str) -> List[Dict]:
        """Parse import file based on extension."""
        _, ext = os.path.splitext(file_path.lower())

        if ext == ".csv":
            return self._parse_csv_file(file_path)
        elif ext in [".xlsx", ".xls"]:
            return self._parse_excel_file(file_path)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    def _parse_csv_file(self, file_path: str) -> List[Dict]:
        """Parse CSV file."""
        data = []
        with open(file_path, "r", encoding="utf-8") as csvfile:
            # Try to detect delimiter
            sample = csvfile.read(1024)
            csvfile.seek(0)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter

            reader = csv.DictReader(csvfile, delimiter=delimiter)
            for row in reader:
                # Clean up row data
                cleaned_row = {k.strip(): v.strip() if v else "" for k, v in row.items()}
                data.append(cleaned_row)

        return data

    def _parse_excel_file(self, file_path: str) -> List[Dict]:
        """Parse Excel file."""
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value.strip() if cell.value else "")

        # Get data rows
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = str(value).strip() if value is not None else ""

            # Skip empty rows
            if any(row_data.values()):
                data.append(row_data)

        return data

    def _validate_import_data(self, data_type: str, data: List[Dict]) -> List[str]:
        """Validate import data."""
        errors = []

        if not data:
            errors.append("No data found in file")
            return errors

        # Get required fields for data type
        required_fields = self._get_required_fields(data_type)

        # Check headers
        headers = set(data[0].keys())
        missing_fields = set(required_fields) - headers
        if missing_fields:
            errors.append(f"Missing required fields: {', '.join(missing_fields)}")

        # Validate each row
        for i, row in enumerate(data[:100]):  # Limit validation to first 100 rows
            row_errors = self._validate_row(
                data_type, row, i + 2
            )  # +2 for header and 1-based indexing
            errors.extend(row_errors)

        return errors

    def _get_required_fields(self, data_type: str) -> List[str]:
        """Get required fields for data type."""
        if data_type == "inventory":
            return [
                "SKU",
                "Name",
                "Karat",
                "Weight (grams)",
                "Cost Price",
                "Selling Price",
                "Quantity",
            ]
        elif data_type == "customers":
            return ["First Name", "Last Name", "Phone"]
        elif data_type == "suppliers":
            return ["Company Name", "Contact Person", "Email", "Phone"]
        else:
            return []

    def _validate_row(self, data_type: str, row: Dict, row_num: int) -> List[str]:
        """Validate a single row."""
        errors = []

        if data_type == "inventory":
            # Validate SKU
            if not row.get("SKU"):
                errors.append(f"Row {row_num}: SKU is required")

            # Validate numeric fields
            for field in ["Karat", "Weight (grams)", "Cost Price", "Selling Price", "Quantity"]:
                value = row.get(field, "").strip()
                if value:
                    try:
                        float(value)
                    except ValueError:
                        errors.append(f"Row {row_num}: {field} must be a number")

        elif data_type == "customers":
            # Validate email format if provided
            email = row.get("Email", "").strip()
            if email and "@" not in email:
                errors.append(f"Row {row_num}: Invalid email format")

        return errors

    def _import_data_records(self, data_type: str, data: List[Dict], update_existing: bool) -> Dict:
        """Import data records."""
        results = {"total": len(data), "successful": 0, "failed": 0, "errors": []}

        with transaction.atomic():
            for i, row in enumerate(data):
                try:
                    if data_type == "inventory":
                        self._import_inventory_record(row, update_existing)
                    elif data_type == "customers":
                        self._import_customer_record(row, update_existing)
                    elif data_type == "suppliers":
                        self._import_supplier_record(row, update_existing)

                    results["successful"] += 1

                except Exception as e:
                    results["failed"] += 1
                    results["errors"].append(f"Row {i + 2}: {str(e)}")

        return results

    def _import_inventory_record(self, row: Dict, update_existing: bool):
        """Import inventory record."""
        InventoryItem = apps.get_model("inventory", "InventoryItem")
        ProductCategory = apps.get_model("inventory", "ProductCategory")
        Branch = apps.get_model("core", "Branch")

        # Get or create category
        category_name = row.get("Category", "").strip()
        category = None
        if category_name:
            category, _ = ProductCategory.objects.get_or_create(
                tenant=self.tenant, name=category_name
            )

        # Get default branch
        branch = Branch.objects.filter(tenant=self.tenant).first()
        if not branch:
            raise ValueError("No branch found for tenant")

        # Prepare data
        item_data = {
            "tenant": self.tenant,
            "sku": row["SKU"].strip(),
            "name": row["Name"].strip(),
            "category": category,
            "karat": int(float(row["Karat"])),
            "weight_grams": Decimal(row["Weight (grams)"]),
            "cost_price": Decimal(row["Cost Price"]),
            "selling_price": Decimal(row["Selling Price"]),
            "quantity": int(float(row["Quantity"])),
            "branch": branch,
            "serial_number": row.get("Serial Number", "").strip() or None,
            "lot_number": row.get("Lot Number", "").strip() or None,
        }

        if update_existing:
            item, created = InventoryItem.objects.update_or_create(
                tenant=self.tenant, sku=item_data["sku"], defaults=item_data
            )
        else:
            # Check if SKU already exists
            if InventoryItem.objects.filter(tenant=self.tenant, sku=item_data["sku"]).exists():
                raise ValueError(f"SKU {item_data['sku']} already exists")

            InventoryItem.objects.create(**item_data)

    def _import_customer_record(self, row: Dict, update_existing: bool):
        """Import customer record."""
        Customer = apps.get_model("crm", "Customer")

        # Generate customer number if not provided
        customer_number = row.get("Customer Number", "").strip()
        if not customer_number:
            # Generate unique customer number
            import uuid

            customer_number = f"CUST-{uuid.uuid4().hex[:8].upper()}"

        # Prepare data
        customer_data = {
            "tenant": self.tenant,
            "customer_number": customer_number,
            "first_name": row["First Name"].strip(),
            "last_name": row["Last Name"].strip(),
            "email": row.get("Email", "").strip() or None,
            "phone": row["Phone"].strip(),
            "loyalty_tier": row.get("Loyalty Tier", "BRONZE"),
            "loyalty_points": int(row.get("Loyalty Points", 0) or 0),
            "store_credit": Decimal(row.get("Store Credit", 0) or 0),
        }

        if update_existing:
            # Try to find existing customer by email or phone
            existing = None
            if customer_data["email"]:
                existing = Customer.objects.filter(
                    tenant=self.tenant, email=customer_data["email"]
                ).first()

            if not existing:
                existing = Customer.objects.filter(
                    tenant=self.tenant, phone=customer_data["phone"]
                ).first()

            if existing:
                for key, value in customer_data.items():
                    if key != "tenant":
                        setattr(existing, key, value)
                existing.save()
            else:
                Customer.objects.create(**customer_data)
        else:
            # Check for duplicates
            if Customer.objects.filter(
                tenant=self.tenant, customer_number=customer_number
            ).exists():
                raise ValueError(f"Customer number {customer_number} already exists")

            Customer.objects.create(**customer_data)

    def _import_supplier_record(self, row: Dict, update_existing: bool):
        """Import supplier record."""
        try:
            Supplier = apps.get_model("procurement", "Supplier")

            # Prepare data
            supplier_data = {
                "tenant": self.tenant,
                "name": row["Company Name"].strip(),
                "contact_person": row["Contact Person"].strip(),
                "email": row["Email"].strip(),
                "phone": row["Phone"].strip(),
                "rating": int(row.get("Rating", 0) or 0),
            }

            if update_existing:
                supplier, created = Supplier.objects.update_or_create(
                    tenant=self.tenant, email=supplier_data["email"], defaults=supplier_data
                )
            else:
                # Check if email already exists
                if Supplier.objects.filter(
                    tenant=self.tenant, email=supplier_data["email"]
                ).exists():
                    raise ValueError(f"Supplier with email {supplier_data['email']} already exists")

                Supplier.objects.create(**supplier_data)

        except LookupError:
            # Supplier model doesn't exist yet
            raise ValueError("Supplier management is not available")


class BackupTriggerService:
    """
    Service for triggering manual backups.
    """

    def __init__(self, tenant: Optional[Tenant] = None):
        self.tenant = tenant

    def trigger_backup(
        self,
        backup_type: str,
        user=None,
        priority: str = "NORMAL",
        reason: str = "",
        scheduled_at: Optional[datetime] = None,
        include_media: bool = True,
        compress_backup: bool = True,
        encrypt_backup: bool = True,
    ) -> BackupTrigger:
        """
        Trigger a manual backup.

        Args:
            backup_type: Type of backup (FULL, TENANT, INCREMENTAL)
            user: User triggering the backup
            priority: Backup priority
            reason: Reason for backup
            scheduled_at: When to execute (None for immediate)
            include_media: Include media files
            compress_backup: Compress the backup
            encrypt_backup: Encrypt the backup

        Returns:
            BackupTrigger: Created trigger record
        """
        # Create backup trigger
        trigger = BackupTrigger.objects.create(
            tenant=self.tenant,
            backup_type=backup_type,
            priority=priority,
            reason=reason,
            scheduled_at=scheduled_at,
            include_media=include_media,
            compress_backup=compress_backup,
            encrypt_backup=encrypt_backup,
            initiated_by=user,
        )

        # Queue backup task (this would integrate with the backup system)
        # For now, we'll just mark it as queued
        trigger.mark_queued()

        logger.info(
            f"Backup triggered: {backup_type} for tenant {self.tenant.id if self.tenant else 'ALL'}"
        )

        return trigger

    def get_backup_status(self, trigger_id: str) -> Dict:
        """Get status of a backup trigger."""
        try:
            trigger = BackupTrigger.objects.get(id=trigger_id)
            return {
                "id": str(trigger.id),
                "status": trigger.status,
                "backup_type": trigger.backup_type,
                "priority": trigger.priority,
                "created_at": trigger.created_at.isoformat(),
                "started_at": trigger.started_at.isoformat() if trigger.started_at else None,
                "completed_at": trigger.completed_at.isoformat() if trigger.completed_at else None,
                "duration": str(trigger.duration) if trigger.duration else None,
                "file_path": trigger.file_path,
                "file_size": trigger.file_size,
                "error_message": trigger.error_message,
            }
        except BackupTrigger.DoesNotExist:
            raise ValueError(f"Backup trigger {trigger_id} not found")
